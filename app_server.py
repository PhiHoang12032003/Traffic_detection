import datetime
import webbrowser
import cv2
import tempfile
import numpy as np
import pandas as pd
import json
import time
from decimal import Decimal
from PIL import Image
from ultralytics import YOLO

from flask import Flask, jsonify, url_for, request, session, send_file
from flask import render_template, Response
from flask_cors import CORS

# Import database modules
from db_config import get_database_connection, VideoDatabase, ViolationDatabase, StatisticsDatabase

import threading
import uuid

# Import Gemini Chatbot
try:
    from gemini_chatbot import GeminiChatbot
    CHATBOT_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ Gemini Chatbot not available: {e}")
    print("   Install: pip install google-generativeai python-dotenv")
    CHATBOT_AVAILABLE = False

from red_light_main import process_red_light_video_complete, generate_frames_red_light_new
import createBB
import createBB_lane  # Import module tạo PDF cho lane violation
from utils.helmet_pdf_utils import create_helmet_pdf_report, get_helmet_violation_info
from processHelmetVideo import process_helmet_video_complete  # Import function xử lý video bảo hiểm
from werkzeug.utils import secure_filename
import os

app = Flask(__name__, static_folder='static')
CORS(app)
app.secret_key = 'your-secret-key-here-change-in-production'

# Global database connections
db_connection = None
video_db = None
violation_db = None
stats_db = None
chatbot = None  # Gemini AI Chatbot instance

def convert_decimal_to_float(obj):
    """
    Recursively convert Decimal objects to float for JSON serialization
    Chuyển đổi Decimal thành float để có thể serialize JSON
    """
    if isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, dict):
        return {key: convert_decimal_to_float(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [convert_decimal_to_float(item) for item in obj]
    elif isinstance(obj, tuple):
        return tuple(convert_decimal_to_float(item) for item in obj)
    return obj

def init_database(password=''):
    """Initialize database connection"""
    global db_connection, video_db, violation_db, stats_db, chatbot
    
    try:
        db_connection = get_database_connection(password)
        
        if db_connection:
            video_db = VideoDatabase(db_connection)
            violation_db = ViolationDatabase(db_connection)
            stats_db = StatisticsDatabase(db_connection)
            
            # Initialize chatbot with database connection
            if CHATBOT_AVAILABLE:
                try:
                    chatbot = GeminiChatbot(db_connection)
                    print("✅ Gemini AI Chatbot initialized")
                except Exception as e:
                    print(f"⚠️ Chatbot initialization failed: {e}")
                    chatbot = None
            
            print("✅ Database initialized successfully")
            return True
        else:
            print("⚠️ Database connection failed - running without database")
            return False
    except Exception as e:
        print(f"⚠️ Database initialization error: {e}")
        print("⚠️ Running without database support")
        return False

# Configure upload
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'mp4', 'avi', 'mov', 'mkv'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB max

# Create upload folder if not exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Global variables for lane detection control
lane_detection_active = False
lane_detection_thread = None
lane_detection_data = {
    'violations': [],
    'motor_violations': 0,
    'car_violations': 0,
    'start_time': None,
    'video_writer': None,
    'output_path': None,
    'timestamp': None,
    'tracked_vehicles': {},  # Tracking để tránh duplicate
    'violation_cooldown': {},  # Cooldown để tránh spam detection
    'violation_frames': [],  # Lưu frames vi phạm để xuất video
    'frame_count': 0,
    'original_fps': None,
    'output_size': None,
    'vehicle_states': {}  # Track trạng thái lane của từng xe: {vehicle_id: {'current_lane': 'motor'/'car', 'last_seen': frame_number}}
}

# Global variables for helmet detection control
helmet_detection_active = False
helmet_detection_thread = None
helmet_detection_data = {
    'violations': [],
    'total_violations': 0,
    'with_helmet': 0,  # Đếm người có mũ
    'without_helmet': 0,  # Đếm người không mũ (vi phạm)
    'start_time': None,
    'video_writer': None,
    'output_path': None,
    'timestamp': None,
    'original_fps': None,
    'output_size': None,
    'detection_cooldown': {},  # Cooldown để tránh đếm trùng: {position_key: frame_number}
    'frame_detections': set(),  # Tracking detections trong frame hiện tại
    'temporal_votes': {},  # Bộ nhớ phiếu theo vị trí để chống nhiễu (position_key -> {window, last_frame})
    'state_map': {},  # Bộ nhớ trạng thái theo vị trí để giảm nhấp nháy: {pos_key: {state, window, last_change, lock_until}}
    'recent_plates': [],  # Danh sách plate ứng viên gần đây: [{text, bbox, frame}]
    'current_with': 0,
    'current_without': 0,
    'current_total': 0,
    'person_states': {},  # Track trạng thái của từng người: {person_id: {'has_violated': bool, 'last_seen': frame_number, 'last_bbox': bbox}}
    'next_person_id': 0,  # Counter để tạo person_id mới
    # Use None so generate_frames_helmet() knows to perform DB-initialization
    'frame_count': None
}

# MySQL Configuration - Commented out for easier setup
# app.config['MYSQL_HOST'] = 'localhost'
# app.config['MYSQL_USER'] = 'root'
# app.config['MYSQL_PASSWORD'] = '12345678'
# app.config['MYSQL_DB'] = 'datn'
# mysql = MySQL(app)


# Apply Flask CORSx`
# CORS(app)
# app.config['CORS_HEADERS'] = 'Content-Type'
#
@app.route('/test', methods=['GET'])
def get_violate():
    """Get violation statistics from database - ALL TIME"""
    try:
        if stats_db:
            # Get statistics from database
            overall = stats_db.get_overall_stats()
            
            if overall:
                # Format data for frontend
                result = []
                today = datetime.date.today().strftime("%Y-%m-%d")
                
                for camera in overall:
                    camera_type = camera['camera_type']
                    total_violations = camera['total_violations'] or 0
                    
                    # Map camera type to vehicle name - DÙNG TỔNG THỰC TẾ TỪ DATABASE
                    if camera_type == 'lane':
                        # Lấy chi tiết vi phạm làn đường
                        lane_stats = stats_db.get_lane_stats(camera['camera_id'])
                        motor_count = 0
                        car_count = 0
                        for stat in lane_stats:
                            if stat['violation_type'] == 'motor_in_car_lane':
                                motor_count = stat['count']
                            elif stat['violation_type'] == 'car_in_motor_lane':
                                car_count = stat['count']
                        result.append(["Xe May", today, motor_count])
                        result.append(["OTO", today, car_count])
                    elif camera_type == 'helmet':
                        result.append(["Xe May", today, total_violations])
                    elif camera_type == 'red_light':
                        result.append(["OTO", today, total_violations])
                
                return jsonify(result)
    except Exception as e:
        print(f"⚠️ Database error: {e}")
        import traceback
        traceback.print_exc()
    
    # Fallback: Empty data khi không có database
    return jsonify([
        ["OTO", datetime.date.today().strftime("%Y-%m-%d"), 0],
        ["Xe May", datetime.date.today().strftime("%Y-%m-%d"), 0]
    ])


@app.route('/test1', methods=['GET'])
def get_violate_current():
    """Get violation statistics for TODAY only from database"""
    try:
        if stats_db:
            today = datetime.date.today()
            result = []
            
            # Camera 1 - Lane violations TODAY
            lane_stats = stats_db.get_lane_stats(camera_id=1, date=today)
            motor_today = 0
            car_today = 0
            for stat in lane_stats:
                if stat['violation_type'] == 'motor_in_car_lane':
                    motor_today = stat['count']
                elif stat['violation_type'] == 'car_in_motor_lane':
                    car_today = stat['count']
            if motor_today > 0:
                result.append(["Xe May", today.strftime("%Y-%m-%d"), motor_today])
            if car_today > 0:
                result.append(["OTO", today.strftime("%Y-%m-%d"), car_today])
            
            # Camera 2 - Helmet violations TODAY
            helmet_stats = stats_db.get_helmet_stats(camera_id=2, date=today)
            if helmet_stats and helmet_stats['no_helmet_count'] > 0:
                result.append(["Xe May", today.strftime("%Y-%m-%d"), helmet_stats['no_helmet_count']])
            
            # Camera 3 - Red light violations TODAY
            redlight_stats = stats_db.get_red_light_stats(camera_id=3, date=today)
            if redlight_stats and redlight_stats['violation_count'] > 0:
                result.append(["OTO", today.strftime("%Y-%m-%d"), redlight_stats['violation_count']])
            
            return jsonify(result if result else [])
    except Exception as e:
        print(f"⚠️ Database error: {e}")
        import traceback
        traceback.print_exc()
    
    # Fallback: Empty data khi không có database
    return jsonify([])


# MySQL database insert function - Disabled for easier setup
# @app.route('/create', methods=['GET'])
def create(cls):
    # MySQL functionality disabled
    # with app.app_context():
    #     cur = mysql.connection.cursor()
    #     ngay_hien_tai = datetime.date.today()
    #     cur.execute("insert into transportationviolation(id_name , date_violate) values (%s, %s)",
    #                 (cls + 1, ngay_hien_tai))
    #     mysql.connection.commit()
    #     cur.close()
    #     return jsonify({'message': 'User created successfully'})
    
    # Simple logging instead of database insert
    ngay_hien_tai = datetime.date.today()
    print(f"Violation detected: Class {cls + 1} on {ngay_hien_tai}")
    return None


def call_route(cls):
    url_for('create', cls=cls)
    # return redirect(url_for('create', cls=cls))


def draw_text(img, text, pos=(10, 30), font_scale=0.7, text_color=(255, 255, 255), text_color_bg=(0, 0, 0)):
    """Vẽ text với background"""
    font = cv2.FONT_HERSHEY_SIMPLEX
    thickness = 2
    x, y = pos
    text_size, _ = cv2.getTextSize(text, font, font_scale, thickness)
    text_w, text_h = text_size
    # Vẽ hình chữ nhật làm nền
    cv2.rectangle(img, pos, (x + text_w + 5, y + text_h + 5), text_color_bg, -1)
    # Vẽ chữ lên trên
    cv2.putText(img, text, (x, y + text_h + 3), font, font_scale, text_color, thickness)

# Code chính phát hiện vi phạm làn đường
def video_detection_web(path_x=""):
    """Video detection cho lane violation với xuất kết quả"""
    global lane_detection_active, lane_detection_data, video_db, violation_db
    
    cap = None
    out = None
    
    try:
        from performance_config import PerformanceConfig, auto_detect_performance
        
        # Khởi tạo performance config
        performance_config = auto_detect_performance()
        
        cap = cv2.VideoCapture(path_x)
        if not cap.isOpened():
            print("❌ Không thể mở video")
            return
            
        model = YOLO('model_lane/vehicle.pt')
        
        # Get video properties
        original_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        original_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        original_fps = int(cap.get(cv2.CAP_PROP_FPS))
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        
        # Calculate new dimensions
        new_width, new_height = performance_config.get_video_dimensions(original_width, original_height)
        
        # Initialize video writer
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = "output"
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"lane_violations_{timestamp}.mp4")
        
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        out = cv2.VideoWriter(output_path, fourcc, original_fps, (new_width, new_height))
        
        if not out.isOpened():
            print("❌ Không thể tạo video writer")
            return
        
        # Giữ lại video_id nếu đã được set trước đó
        existing_video_id = lane_detection_data.get('video_id')
        
        # Update global data
        lane_detection_data.update({
            'start_time': time.time(),
            'video_writer': out,
            'output_path': output_path,
            'timestamp': timestamp,
            'original_fps': original_fps,
            'output_size': (new_width, new_height),
            'violations': [],
            'motor_violations': 0,
            'car_violations': 0,
            'tracked_vehicles': {},
            'violation_cooldown': {},
            'violation_frames': [],
            'frame_count': 0,
            'video_id': existing_video_id  # Giữ lại video_id đã set
        })
        
        frame_count = 0
        processed_count = 0
        
        # Update video status to 'processing'
        video_id = lane_detection_data.get('video_id')
        print(f"🔍 [VIDEO_ID CHECK] video_id in lane_detection_data: {video_id}")
        
        if video_id and video_db:
            try:
                video_db.update_video_status(video_id, 'processing')
                print(f"📹 Video {video_id} status: processing")
            except Exception as e:
                print(f"⚠️ Failed to update video status: {e}")
        
        print(f"🚀 Bắt đầu phân tích web - Performance: {performance_config.mode}")
        print(f"📹 Video output: {output_path}")
        
        while lane_detection_active and cap.isOpened():
            success, frame = cap.read()
            if not success:
                break

            frame_count += 1

            # Resize frame immediately
            frame = cv2.resize(frame, (new_width, new_height))
            h, w, _ = frame.shape

            # Update global frame count
            lane_detection_data['frame_count'] = frame_count

            # Vẽ UI elements cơ bản cho mọi frame
            roi_start = (0, int(0.2 * h))
            roi_end = (w, int(0.8 * h))
            # Đã bỏ vẽ ROI xanh dương
            
            # Làn xe máy: 17% - 42% width (trong vùng video)
            start_line_motor = (int(0.14 * w), int(0.15 * h))
            end_line_motor = (int(0.42 * w), int(0.85 * h))
            cv2.rectangle(frame, start_line_motor, end_line_motor, (255, 0, 255), 2)
            draw_text(frame, "LANE XE MAY", (int(0.18 * w), int(0.18 * h)), text_color=(255, 255, 255), text_color_bg=(255, 0, 255))
            
            # Làn ô tô: 42% - 88% width (trong vùng video)
            start_line_car = (int(0.42 * w), int(0.15 * h))
            end_line_car = (int(0.85 * w), int(0.85 * h))
            cv2.rectangle(frame, start_line_car, end_line_car, (0, 255, 0), 2)
            draw_text(frame, "LANE O TO", (int(0.50 * w), int(0.18 * h)), text_color=(255, 255, 255), text_color_bg=(0, 255, 0))
            
            # If frame should be skipped, still write with basic UI (chỉ lanes, không có text thống kê)
            if not performance_config.should_process_frame(frame_count):
                try:
                    out.write(frame)
                except Exception:
                    pass

                yield frame
                continue

            processed_count += 1
            
            # Run YOLO
            results = model(frame, 
                          conf=performance_config.yolo_conf_threshold,
                          imgsz=performance_config.yolo_img_size,
                          verbose=False)
            
            # Debug: In class names một lần để kiểm tra
            if frame_count == 1:
                print(f"🔍 [MODEL DEBUG] Class names: {results[0].names}")
            
            # Process detections
            for result in results:
                boxes = result.boxes
                if boxes is not None:
                    for box in boxes:
                        try:
                            x1, y1, x2, y2 = map(int, box.xyxy[0])
                            cls = int(box.cls[0])
                            conf = float(box.conf[0])
                            
                            center_x = (x1 + x2) // 2
                            center_y = (y1 + y2) // 2
                            
                            # Chỉ xử lý trong ROI
                            if not (roi_start[1] < center_y < roi_end[1]):
                                continue
                            
                            # LOGIC VI PHẠM CẢI TIẾN - Global vehicle tracking để tránh đếm trùng
                            violation_detected = False
                            violation_type = ""
                            
                            lane_boundary = w // 2
                            
                            # Tìm vehicle ID dựa trên IoU với các xe đã track
                            # Sử dụng khoảng cách và IoU để match với xe hiện có
                            vehicle_id = None
                            best_match_score = 0
                            current_bbox = (x1, y1, x2, y2)
                            
                            # Helper function tính IoU
                            def calculate_iou(box1, box2):
                                x1_1, y1_1, x2_1, y2_1 = box1
                                x1_2, y1_2, x2_2, y2_2 = box2
                                
                                inter_x1 = max(x1_1, x1_2)
                                inter_y1 = max(y1_1, y1_2)
                                inter_x2 = min(x2_1, x2_2)
                                inter_y2 = min(y2_1, y2_2)
                                
                                inter_area = max(0, inter_x2 - inter_x1) * max(0, inter_y2 - inter_y1)
                                
                                box1_area = (x2_1 - x1_1) * (y2_1 - y1_1)
                                box2_area = (x2_2 - x1_2) * (y2_2 - y1_2)
                                union_area = box1_area + box2_area - inter_area
                                
                                return inter_area / union_area if union_area > 0 else 0
                            
                            # Tìm xe match với detection hiện tại
                            best_iou = 0
                            best_dist = 999999
                            for vid, vstate in lane_detection_data['vehicle_states'].items():
                                # Chỉ xem xét xe cùng class và được thấy trong 30 frames gần đây
                                if vstate.get('class') != cls:
                                    continue
                                if frame_count - vstate.get('last_seen', 0) > 30:
                                    continue
                                    
                                last_bbox = vstate.get('last_bbox')
                                if last_bbox is None:
                                    continue
                                
                                iou = calculate_iou(current_bbox, last_bbox)
                                
                                # Tính khoảng cách center
                                last_cx = (last_bbox[0] + last_bbox[2]) // 2
                                last_cy = (last_bbox[1] + last_bbox[3]) // 2
                                dist = ((center_x - last_cx)**2 + (center_y - last_cy)**2) ** 0.5
                                
                                # Match score: IoU cao hoặc khoảng cách gần
                                match_score = iou + (1.0 / (1.0 + dist / 100))
                                
                                if match_score > best_match_score and (iou > 0.3 or dist < 150):
                                    best_match_score = match_score
                                    vehicle_id = vid
                                    best_iou = iou
                                    best_dist = dist
                            
                            # Nếu không tìm thấy match, tạo ID mới
                            if vehicle_id is None:
                                # Tạo unique ID mới
                                if 'next_vehicle_id' not in lane_detection_data:
                                    lane_detection_data['next_vehicle_id'] = 0
                                vehicle_id = f"vehicle_{cls}_{lane_detection_data['next_vehicle_id']}"
                                lane_detection_data['next_vehicle_id'] += 1
                                print(f"🆕 [NEW VEHICLE] {vehicle_id} | Class {cls} | Frame {frame_count} | Pos ({center_x}, {center_y})")
                            else:
                                # Matched existing vehicle - in chi tiết khi debug mode
                                if frame_count % 50 == 0:  # Log mỗi 50 frames
                                    print(f"🔗 [MATCHED] {vehicle_id} | IoU {best_iou:.2f} | Dist {best_dist:.0f}px | Frame {frame_count}")
                            
                            # Xác định lane hiện tại của vehicle (kiểm tra cả center và bbox)
                            # Sử dụng majority rule: nếu >60% bbox nằm trong lane thì tính là trong lane đó
                            bbox_width = x2 - x1
                            left_in_motor = max(0, min(x2, lane_boundary) - x1)
                            right_in_car = max(0, x2 - max(x1, lane_boundary))
                            
                            if left_in_motor > bbox_width * 0.6:
                                current_lane = "motor"
                            elif right_in_car > bbox_width * 0.6:
                                current_lane = "car"
                            else:
                                # Xe đang ở giữa 2 làn - dùng center để quyết định
                                current_lane = "motor" if center_x < lane_boundary else "car"
                            
                            vehicle_type = "xe_may" if cls == 1 else "oto" if cls in [0, 3, 4] else f"unknown_cls_{cls}"
                            
                            # Khởi tạo hoặc update vehicle state
                            if vehicle_id not in lane_detection_data['vehicle_states']:
                                lane_detection_data['vehicle_states'][vehicle_id] = {
                                    'current_lane': current_lane,
                                    'lane_history': [current_lane],  # Track lane over frames
                                    'last_seen': frame_count,
                                    'has_violated': False,
                                    'violation_frame': None,
                                    'class': cls,
                                    'confidence_sum': conf,
                                    'detection_count': 1,
                                    'last_bbox': current_bbox  # Thêm bbox để tracking
                                }
                            else:
                                # Update existing vehicle
                                vehicle_state = lane_detection_data['vehicle_states'][vehicle_id]
                                vehicle_state['last_seen'] = frame_count
                                vehicle_state['current_lane'] = current_lane
                                vehicle_state['confidence_sum'] += conf
                                vehicle_state['detection_count'] += 1
                                vehicle_state['last_bbox'] = current_bbox  # Cập nhật bbox
                                
                                # Update lane history (keep last 10 frames)
                                vehicle_state['lane_history'].append(current_lane)
                                if len(vehicle_state['lane_history']) > 10:
                                    vehicle_state['lane_history'] = vehicle_state['lane_history'][-10:]
                            
                            vehicle_state = lane_detection_data['vehicle_states'][vehicle_id]
                            
                            # Multi-frame validation: xe phải ở sai lane ít nhất 3 trong 5 frames gần nhất
                            lane_history = vehicle_state['lane_history']
                            if len(lane_history) >= 3:
                                recent_5 = lane_history[-5:] if len(lane_history) >= 5 else lane_history
                                
                                # Đếm số lần xe ở sai lane
                                wrong_lane_count = 0
                                for lane in recent_5:
                                    # XE MÁY ở LANE Ô TÔ = sai
                                    if cls == 1 and lane == "car":
                                        wrong_lane_count += 1
                                    # Ô TÔ ở LANE XE MÁY = sai
                                    elif cls in [0, 3, 4] and lane == "motor":
                                        wrong_lane_count += 1
                                
                                # Nếu xe ở sai lane >= 3 frames trong 5 frames gần nhất
                                # VÀ chưa bị đếm vi phạm
                                # VÀ confidence trung bình đủ cao (>= 0.4)
                                avg_conf = vehicle_state['confidence_sum'] / vehicle_state['detection_count']
                                
                                if (wrong_lane_count >= 3 and 
                                    not vehicle_state['has_violated'] and 
                                    avg_conf >= 0.4):
                                    
                                    # XE MÁY ở LANE Ô TÔ
                                    if cls == 1 and current_lane == "car":
                                        violation_detected = True
                                        violation_type = "xe_may_vi_pham_lan_oto"
                                        vehicle_state['has_violated'] = True
                                        vehicle_state['violation_frame'] = frame_count
                                        
                                        old_count = lane_detection_data['motor_violations']
                                        lane_detection_data['motor_violations'] += 1
                                        new_count = lane_detection_data['motor_violations']
                                        print(f"🚨🏍️ [MOTOR VIOLATION] {old_count} → {new_count} | Vehicle {vehicle_id} | Frame {frame_count} | History: {recent_5} | Conf: {avg_conf:.2f}")
                                    
                                    # Ô TÔ ở LANE XE MÁY
                                    elif cls in [0, 3, 4] and current_lane == "motor":
                                        violation_detected = True
                                        violation_type = "oto_vi_pham_lan_xe_may"
                                        vehicle_state['has_violated'] = True
                                        vehicle_state['violation_frame'] = frame_count
                                        
                                        old_count = lane_detection_data['car_violations']
                                        lane_detection_data['car_violations'] += 1
                                        new_count = lane_detection_data['car_violations']
                                        print(f"🚨🚗 [CAR VIOLATION] {old_count} → {new_count} | Vehicle {vehicle_id} | Frame {frame_count} | History: {recent_5} | Conf: {avg_conf:.2f}")
                            
                            # Violation key cho backward compatibility
                            violation_key = f"{vehicle_id}_{current_lane}"
                            frame_violation_key = f"{vehicle_id}_frame_{frame_count}"
                            cooldown_frames = 150
                            
                            
                            # Xử lý hiển thị vi phạm trên video
                            # Kiểm tra xem xe có đang ở sai làn không (hiển thị ngay lập tức)
                            is_in_wrong_lane = False
                            if cls == 1 and current_lane == "car":  # Xe máy ở làn ô tô
                                is_in_wrong_lane = True
                            elif cls in [0, 3, 4] and current_lane == "motor":  # Ô tô ở làn xe máy
                                is_in_wrong_lane = True
                            
                            # Hoặc xe đã bị đánh dấu vi phạm trước đó
                            if violation_detected or vehicle_state.get('has_violated', False) or is_in_wrong_lane:
                                # VẼ BOUNDING BOX ĐỎ CHO XE VI PHẠM hoặc đang ở sai làn
                                cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 2)
                                label = f"{result.names[cls]} {conf:.2f}"
                                draw_text(frame, label, pos=(x1, y1 - 20), font_scale=0.5, text_color_bg=(0, 0, 255))
                                
                                # Vẽ text vi phạm với thông tin chi tiết
                                if violation_detected or vehicle_state.get('has_violated', False):
                                    draw_text(frame, "VI PHAM!", (x1, y1 - 40), 
                                            text_color=(255, 255, 255), text_color_bg=(255, 0, 0))
                                else:
                                    draw_text(frame, "SAI LAN!", (x1, y1 - 40), 
                                            text_color=(255, 255, 255), text_color_bg=(255, 165, 0))
                                draw_text(frame, f"Lane: {current_lane.upper()}", (x1, y1 - 60), 
                                        font_scale=0.5, text_color=(255, 255, 255), text_color_bg=(255, 0, 0))
                                    
                            else:
                                # VẼ BOUNDING BOX XANH LÁ CHO XE KHÔNG VI PHẠM
                                color = (0, 255, 0)
                                cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                                label = f"{result.names[cls]} {conf:.2f}"
                                draw_text(frame, label, pos=(x1, y1 - 20), font_scale=0.5, text_color_bg=color)
                                
                                # Hiển thị lane hiện tại cho xe không vi phạm (để debug)
                                if frame_count % 15 == 0:  # Mỗi 0.5 giây
                                    lane_color = (255, 0, 255) if current_lane == "motor" else (0, 255, 0)
                                    draw_text(frame, current_lane.upper(), (x1, y2 + 5), 
                                            font_scale=0.4, text_color=(255, 255, 255), text_color_bg=lane_color)
                            
                            # Lưu thông tin vi phạm
                            if violation_detected:
                                violation_info = {
                                    'violation_id': len(lane_detection_data['violations']) + 1,
                                    'type': violation_type,
                                    'frame_number': frame_count,
                                    'time_seconds': frame_count / original_fps,
                                    'time_formatted': f"{int((frame_count / original_fps) // 60):02d}:{int((frame_count / original_fps) % 60):02d}",
                                    'confidence': conf,
                                    'bbox': [x1, y1, x2, y2],
                                    'center': [center_x, center_y],
                                    'vehicle_class': result.names[cls],
                                    'detected_at': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                }
                                lane_detection_data['violations'].append(violation_info)
                                
                                # TẠO PDF BIÊN BẢN VI PHẠM
                                try:
                                    # Tạo thư mục lưu ảnh và PDF vi phạm
                                    if violation_type == "xe_may_vi_pham_lan_oto":
                                        # Xe máy vi phạm
                                        os.makedirs("data_xe_may_vi_pham", exist_ok=True)
                                        os.makedirs("BienBanNopPhatXeMay", exist_ok=True)
                                        
                                        stt = lane_detection_data['motor_violations']
                                        violation_img_path = f"data_xe_may_vi_pham/{stt}.jpg"
                                        pdf_report_path = f"BienBanNopPhatXeMay/{stt}.pdf"
                                        
                                        # Lưu ảnh vi phạm
                                        cv2.imwrite(violation_img_path, frame)
                                        
                                        # Tạo ảnh tạm từ frame hiện tại
                                        frame_pil = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
                                        temp_image = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
                                        frame_pil.save(temp_image.name)
                                        
                                        # Cập nhật thông tin biên bản
                                        examBB_motor = createBB_lane.infoObject_motor()
                                        examBB_motor['date'] = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
                                        
                                        # Tạo PDF biên bản
                                        createBB_lane.bienBanNopPhat(examBB_motor, temp_image.name, 
                                                                    violation_img_path, pdf_report_path)
                                        temp_image.close()
                                        
                                        print(f"✅ Đã tạo PDF biên bản xe máy: {pdf_report_path}")
                                        violation_info['pdf_report_path'] = pdf_report_path
                                        violation_info['image_path'] = violation_img_path
                                        
                                    elif violation_type == "oto_vi_pham_lan_xe_may":
                                        # Ô tô vi phạm
                                        os.makedirs("data_oto_vi_pham", exist_ok=True)
                                        os.makedirs("BienBanNopPhatXeOTo", exist_ok=True)
                                        
                                        stt = lane_detection_data['car_violations']
                                        violation_img_path = f"data_oto_vi_pham/{stt}.jpg"
                                        pdf_report_path = f"BienBanNopPhatXeOTo/{stt}.pdf"
                                        
                                        # Lưu ảnh vi phạm
                                        cv2.imwrite(violation_img_path, frame)
                                        
                                        # Tạo ảnh tạm từ frame hiện tại
                                        frame_pil = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
                                        temp_image = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
                                        frame_pil.save(temp_image.name)
                                        
                                        # Cập nhật thông tin biên bản
                                        examBB_car = createBB_lane.infoObject_car()
                                        examBB_car['date'] = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
                                        
                                        # Tạo PDF biên bản
                                        createBB_lane.bienBanNopPhat(examBB_car, temp_image.name, 
                                                                    violation_img_path, pdf_report_path)
                                        temp_image.close()
                                        
                                        print(f"✅ Đã tạo PDF biên bản ô tô: {pdf_report_path}")
                                        violation_info['pdf_report_path'] = pdf_report_path
                                        violation_info['image_path'] = violation_img_path
                                        
                                except Exception as pdf_error:
                                    print(f"⚠️ Lỗi tạo PDF cho vi phạm: {pdf_error}")
                                    import traceback
                                    traceback.print_exc()
                                
                                # LƯU VÀO DATABASE
                                video_id = lane_detection_data.get('video_id')
                                print(f"🔍 [DB DEBUG] video_id={video_id}, violation_db={violation_db is not None}")
                                
                                if video_id and violation_db:
                                    try:
                                        # Map violation type to DB format
                                        db_violation_type = 'motor_in_car_lane' if violation_type == 'xe_may_vi_pham_lan_oto' else 'car_in_motor_lane'
                                        
                                        db_violation_id = violation_db.insert_lane_violation(
                                            video_id=video_id,
                                            frame_number=frame_count,
                                            time_in_video=frame_count / original_fps,
                                            violation_type=db_violation_type,
                                            vehicle_type=result.names[cls],
                                            confidence=conf,
                                            bbox=[x1, y1, x2, y2],
                                            image_path=violation_info.get('image_path'),  # Lưu đường dẫn ảnh
                                            pdf_report_path=violation_info.get('pdf_report_path')  # Lưu đường dẫn PDF
                                        )
                                        
                                        if db_violation_id:
                                            print(f"💾 Saved to DB: violation_id={db_violation_id}, type={db_violation_type}, PDF={violation_info.get('pdf_report_path')}")
                                    except Exception as e:
                                        print(f"⚠️ Failed to save violation to DB: {e}")
                                        import traceback
                                        traceback.print_exc()
                                else:
                                    if not video_id:
                                        print(f"⚠️ No video_id in lane_detection_data")
                                    if not violation_db:
                                        print(f"⚠️ violation_db is None")
                                
                        except Exception as e:
                            print(f"Lỗi xử lý detection: {e}")
                            continue
            
            # Cleanup vehicle states và frame violations
            if frame_count % 100 == 0:  # Cleanup mỗi 100 frames (3-4 giây)
                current_frame = frame_count
                
                # Cleanup vehicle states - xóa vehicles không thấy > 100 frames (3-4 giây)
                expired_vehicles = [vid for vid, state in lane_detection_data['vehicle_states'].items() 
                                  if current_frame - state['last_seen'] > 100]
                for vid in expired_vehicles:
                    del lane_detection_data['vehicle_states'][vid]
                
                # Cleanup violation cooldown
                expired_keys = [k for k, v in lane_detection_data['violation_cooldown'].items() 
                              if current_frame - v > 200]
                for k in expired_keys:
                    del lane_detection_data['violation_cooldown'][k]
                    
                # Cleanup frame violations cũ
                if 'frame_violations' in lane_detection_data:
                    recent_violations = {v for v in lane_detection_data['frame_violations'] 
                                       if any(f"_frame_{current_frame - i}" in v for i in range(10))}
                    lane_detection_data['frame_violations'] = recent_violations
                    
                if len(expired_vehicles) > 0:
                    print(f"🧹 [FRAME {frame_count}] Cleaned {len(expired_vehicles)} expired vehicles")
                print(f"📊 Active vehicles: {len(lane_detection_data['vehicle_states'])}, Violations: Motor={lane_detection_data['motor_violations']}, Car={lane_detection_data['car_violations']}")
            
            # Debug log ít hơn để tránh spam
            if frame_count % 200 == 0:
                active_vehicles = len(lane_detection_data['vehicle_states'])
                violated_vehicles = sum(1 for state in lane_detection_data['vehicle_states'].values() if state.get('has_violated', False))
                print(f"🔍 [DEBUG FRAME {frame_count}] Motor violations: {lane_detection_data['motor_violations']}, Car violations: {lane_detection_data['car_violations']}")
                print(f"🚗 Active vehicles: {active_vehicles}, Violated: {violated_vehicles}")
                
                # Debug top 3 vehicles với lane history
                for vid, state in list(lane_detection_data['vehicle_states'].items())[:3]:
                    history_str = ''.join(['M' if l == 'motor' else 'C' for l in state.get('lane_history', [])])
                    print(f"   Vehicle {vid}: lane={state['current_lane']}, violated={state['has_violated']}, history={history_str}")
            
            # Video sạch - không hiển thị text thống kê trên video
            # Chỉ thống kê thời gian thực ở web UI bên phải
            
            # Ghi frame
            out.write(frame)
            
            # Throttle
            time.sleep(0.033)
            
            yield frame
        
        # End of while loop - video đã xử lý xong hoặc bị dừng
        print(f"✅ Video processing loop ended - {processed_count}/{frame_count} frames processed")
        
    except Exception as e:
        print(f"❌ Lỗi trong video_detection_web: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        # QUAN TRỌNG: Đảm bảo cap và out luôn được release TRƯỚC KHI export
        print("🧹 Cleanup resources...")
        if cap is not None:
            try:
                cap.release()
                print("✅ Camera released")
            except Exception as e:
                print(f"⚠️ Error releasing camera: {e}")
        
        if out is not None:
            try:
                out.release()
                print("✅ Video writer released - Video đã được finalize")
                # Đánh dấu đã release để stop_lane_detection biết
                lane_detection_data['video_writer'] = None
            except Exception as e:
                print(f"⚠️ Error releasing video writer: {e}")
        
        cv2.destroyAllWindows()
        
        # SAU KHI release video writer, mới export kết quả
        print("📦 Tự động xuất kết quả CSV/Excel/JSON TỪ DATABASE...")
        try:
            timestamp = lane_detection_data.get('timestamp', datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
            output_path = lane_detection_data.get('output_path', '')
            video_id = lane_detection_data.get('video_id')
            
            output_dir = "output"
            os.makedirs(output_dir, exist_ok=True)
            
            # ========== LẤY DỮ LIỆU TỪ DATABASE THAY VÌ MEMORY ==========
            violations_data = []
            motor_count_db = 0
            car_count_db = 0
            
            if video_id and violation_db:
                print(f"📊 Đang truy vấn dữ liệu vi phạm từ database (video_id={video_id})...")
                try:
                    # Truy vấn tất cả vi phạm của video này từ database
                    query = """
                        SELECT 
                            violation_id,
                            frame_number,
                            time_in_video,
                            violation_type,
                            vehicle_type,
                            confidence,
                            bbox_x1, bbox_y1, bbox_x2, bbox_y2,
                            detected_at
                        FROM lane_violations
                        WHERE video_id = %s
                        ORDER BY frame_number ASC
                    """
                    db_results = violation_db.db.execute_query(query, (video_id,), fetch=True)
                    
                    if db_results:
                        print(f"✅ Lấy được {len(db_results)} vi phạm từ database cho video_id={video_id}")
                        
                        # Chuyển đổi sang format cho export
                        for row in db_results:
                            time_seconds = row.get('time_in_video', 0)
                            # Convert Decimal to float for JSON serialization
                            if hasattr(time_seconds, '__float__'):
                                time_seconds = float(time_seconds)
                            
                            confidence = row.get('confidence')
                            if hasattr(confidence, '__float__'):
                                confidence = float(confidence)
                            
                            violation_info = {
                                'violation_id': int(row.get('violation_id')) if row.get('violation_id') else None,
                                'frame_number': int(row.get('frame_number')) if row.get('frame_number') else None,
                                'time_seconds': time_seconds,
                                'time_formatted': f"{int(time_seconds // 60):02d}:{int(time_seconds % 60):02d}",
                                'violation_type': row.get('violation_type'),
                                'vehicle_type': row.get('vehicle_type'),
                                'confidence': confidence,
                                'bbox': [
                                    int(row.get('bbox_x1')) if row.get('bbox_x1') is not None else None,
                                    int(row.get('bbox_y1')) if row.get('bbox_y1') is not None else None,
                                    int(row.get('bbox_x2')) if row.get('bbox_x2') is not None else None,
                                    int(row.get('bbox_y2')) if row.get('bbox_y2') is not None else None
                                ],
                                'detected_at': row.get('detected_at').strftime('%Y-%m-%d %H:%M:%S') if row.get('detected_at') else ''
                            }
                            violations_data.append(violation_info)
                            
                            # Đếm theo loại
                            if row.get('violation_type') == 'motor_in_car_lane':
                                motor_count_db += 1
                            elif row.get('violation_type') == 'car_in_motor_lane':
                                car_count_db += 1
                        
                        print(f"📊 Thống kê từ DB (video_id={video_id}): Xe máy={motor_count_db}, Ô tô={car_count_db}, Tổng={len(violations_data)}")
                    else:
                        print(f"⚠️ Không tìm thấy vi phạm cho video_id={video_id}, đang lấy TẤT CẢ dữ liệu từ database...")
                        
                        # FALLBACK: Lấy TẤT CẢ vi phạm từ database (không filter video_id)
                        query_all = """
                            SELECT 
                                v.violation_id,
                                v.video_id,
                                vid.video_filename,
                                v.frame_number,
                                v.time_in_video,
                                v.violation_type,
                                v.vehicle_type,
                                v.confidence,
                                v.bbox_x1, v.bbox_y1, v.bbox_x2, v.bbox_y2,
                                v.detected_at
                            FROM lane_violations v
                            LEFT JOIN videos vid ON v.video_id = vid.video_id
                            WHERE v.camera_id = 1
                            ORDER BY v.detected_at DESC
                            LIMIT 100
                        """
                        db_results_all = violation_db.db.execute_query(query_all, fetch=True)
                        
                        if db_results_all:
                            print(f"✅ Lấy được {len(db_results_all)} vi phạm TỪ TẤT CẢ VIDEO trong database")
                            
                            for row in db_results_all:
                                time_seconds = row.get('time_in_video', 0)
                                # Convert Decimal to float for JSON serialization
                                if hasattr(time_seconds, '__float__'):
                                    time_seconds = float(time_seconds)
                                
                                confidence = row.get('confidence')
                                if hasattr(confidence, '__float__'):
                                    confidence = float(confidence)
                                
                                violation_info = {
                                    'violation_id': int(row.get('violation_id')) if row.get('violation_id') else None,
                                    'video_id': int(row.get('video_id')) if row.get('video_id') else None,
                                    'video_filename': row.get('video_filename', 'N/A'),
                                    'frame_number': int(row.get('frame_number')) if row.get('frame_number') else None,
                                    'time_seconds': time_seconds,
                                    'time_formatted': f"{int(time_seconds // 60):02d}:{int(time_seconds % 60):02d}",
                                    'violation_type': row.get('violation_type'),
                                    'vehicle_type': row.get('vehicle_type'),
                                    'confidence': confidence,
                                    'bbox': [
                                        int(row.get('bbox_x1')) if row.get('bbox_x1') is not None else None,
                                        int(row.get('bbox_y1')) if row.get('bbox_y1') is not None else None,
                                        int(row.get('bbox_x2')) if row.get('bbox_x2') is not None else None,
                                        int(row.get('bbox_y2')) if row.get('bbox_y2') is not None else None
                                    ],
                                    'detected_at': row.get('detected_at').strftime('%Y-%m-%d %H:%M:%S') if row.get('detected_at') else ''
                                }
                                violations_data.append(violation_info)
                                
                                # Đếm theo loại
                                if row.get('violation_type') == 'motor_in_car_lane':
                                    motor_count_db += 1
                                elif row.get('violation_type') == 'car_in_motor_lane':
                                    car_count_db += 1
                            
                            print(f"📊 Thống kê từ DB (TẤT CẢ): Xe máy={motor_count_db}, Ô tô={car_count_db}, Tổng={len(violations_data)}")
                        else:
                            print(f"❌ KHÔNG CÓ DỮ LIỆU TRONG DATABASE!")
                        
                except Exception as db_error:
                    print(f"❌ Lỗi khi truy vấn database: {db_error}")
                    import traceback
                    traceback.print_exc()
                    # Fallback về memory nếu database lỗi
                    violations_data = lane_detection_data.get('violations', [])
                    motor_count_db = lane_detection_data.get('motor_violations', 0)
                    car_count_db = lane_detection_data.get('car_violations', 0)
                    print(f"⚠️ Fallback: Sử dụng dữ liệu từ memory")
            else:
                # Không có database hoặc video_id, dùng memory
                print(f"⚠️ Không có database connection hoặc video_id, sử dụng dữ liệu từ memory")
                violations_data = lane_detection_data.get('violations', [])
                motor_count_db = lane_detection_data.get('motor_violations', 0)
                car_count_db = lane_detection_data.get('car_violations', 0)
            
            # ========== XUẤT CSV/EXCEL/JSON TỪ DỮ LIỆU DATABASE ==========
            
            # Xuất CSV
            csv_filename = os.path.join(output_dir, f"lane_violations_stats_{timestamp}.csv")
            if len(violations_data) > 0:
                df = pd.DataFrame(violations_data)
                # Map violation_type sang tiếng Việt
                violation_type_map = {
                    'motor_in_car_lane': 'Xe máy vi phạm làn ô tô',
                    'car_in_motor_lane': 'Ô tô vi phạm làn xe máy'
                }
                if 'violation_type' in df.columns:
                    df['violation_type_vn'] = df['violation_type'].map(violation_type_map)
            else:
                # Tạo DataFrame rỗng với các columns chuẩn
                df = pd.DataFrame(columns=['violation_id', 'frame_number', 'time_seconds', 'time_formatted', 
                                          'violation_type', 'vehicle_type', 'confidence', 'detected_at'])
            df.to_csv(csv_filename, index=False, encoding='utf-8-sig')
            print(f"✅ Đã xuất lane CSV từ DB: {csv_filename} ({len(violations_data)} vi phạm)")
            
            # Lưu filename vào global data để endpoint stop có thể trả về
            lane_detection_data['csv_filename'] = csv_filename
            
            # Xuất Excel
            xlsx_filename = os.path.join(output_dir, f"lane_violations_stats_{timestamp}.xlsx")
            try:
                with pd.ExcelWriter(xlsx_filename, engine='openpyxl') as writer:
                    # Sheet 1: Chi tiết vi phạm
                    df_export = df.copy()
                    if len(df_export) > 0:
                        # Chọn columns để export
                        export_cols = ['violation_id', 'frame_number', 'time_formatted', 
                                      'violation_type_vn', 'vehicle_type', 'confidence', 'detected_at']
                        export_cols = [col for col in export_cols if col in df_export.columns]
                        df_export = df_export[export_cols]
                        
                        # Đổi tên columns sang tiếng Việt
                        df_export.columns = ['ID Vi Phạm', 'Frame', 'Thời Gian', 
                                            'Loại Vi Phạm', 'Loại Xe', 'Độ Tin Cậy', 'Thời Điểm']
                    
                    df_export.to_excel(writer, sheet_name='Vi Phạm', index=False)
                    
                    # Sheet 2: Thống kê
                    summary_data = {
                        'Chỉ Số': [
                            'Tổng Vi Phạm',
                            'Xe Máy Vi Phạm Làn Ô Tô',
                            'Ô Tô Vi Phạm Làn Xe Máy',
                            'Video ID',
                            'Timestamp',
                            'Nguồn Dữ Liệu'
                        ],
                        'Giá Trị': [
                            len(violations_data),
                            motor_count_db,
                            car_count_db,
                            video_id or 'N/A',
                            timestamp,
                            'MySQL Database' if video_id and violation_db else 'Memory (No DB)'
                        ]
                    }
                    df_summary = pd.DataFrame(summary_data)
                    df_summary.to_excel(writer, sheet_name='Thống Kê', index=False)
                    
                    # Auto-adjust column width
                    for sheet_name in writer.sheets:
                        worksheet = writer.sheets[sheet_name]
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                
                print(f"✅ Đã xuất lane Excel từ DB: {xlsx_filename} ({len(violations_data)} vi phạm)")
                lane_detection_data['xlsx_filename'] = xlsx_filename
            except Exception as e:
                print(f"⚠️ Không thể xuất Excel: {e}")
                import traceback
                traceback.print_exc()
                lane_detection_data['xlsx_filename'] = None
            
            # Xuất JSON chi tiết
            json_filename = os.path.join(output_dir, f"lane_violations_details_{timestamp}.json")
            with open(json_filename, 'w', encoding='utf-8') as f:
                # Convert Decimal to float before JSON serialization
                violations_data_converted = convert_decimal_to_float(violations_data)
                json.dump({
                    'video_info': {
                        'video_id': video_id,
                        'input_path': path_x,
                        'output_path': output_path,
                        'timestamp': timestamp,
                        'total_frames': frame_count,
                        'frames_processed': processed_count,
                        'fps': original_fps,
                        'data_source': 'MySQL Database' if video_id and violation_db else 'Memory'
                    },
                    'violations': violations_data_converted,
                    'summary': {
                        'total_violations': len(violations_data),
                        'motor_violations': motor_count_db,
                        'car_violations': car_count_db,
                        'processing_time': time.time() - lane_detection_data.get('start_time', time.time()),
                        'frames_processed': frame_count
                    }
                }, f, indent=2, ensure_ascii=False)
            print(f"✅ Đã xuất lane JSON từ DB: {json_filename}")
            lane_detection_data['json_filename'] = json_filename
            
            # Update video status to 'completed'
            if video_id and video_db:
                try:
                    video_db.update_video_status(video_id, 'completed')
                    print(f"✅ Video {video_id} status: completed")
                except Exception as e:
                    print(f"⚠️ Failed to update video status: {e}")
            
            print(f"📊 Tổng kết vi phạm (từ DB): Xe máy={motor_count_db}, Ô tô={car_count_db}, Tổng={len(violations_data)}")
            print(f"🎉 Đã xuất tất cả file thành công từ DATABASE!")
            
            # Verify video có thể mở được không
            if output_path and os.path.exists(output_path):
                test_cap = cv2.VideoCapture(output_path)
                if test_cap.isOpened():
                    frames = int(test_cap.get(cv2.CAP_PROP_FRAME_COUNT))
                    print(f"✅ Video có thể xem được: {frames} frames")
                    test_cap.release()
                else:
                    print(f"⚠️ Video không mở được - có thể bị lỗi!")
        except Exception as export_error:
            print(f"⚠️ Lỗi khi xuất kết quả tự động: {export_error}")
            import traceback
            traceback.print_exc()


def reset_lane_detection_data():
    """Reset toàn bộ dữ liệu lane detection cho phiên mới"""
    global lane_detection_data
    
    print("🔄 RESET lane detection data for new session...")
    
    # Giữ lại video_id nếu đã được set
    current_video_id = lane_detection_data.get('video_id')
    
    lane_detection_data.update({
        'violations': [],
        'motor_violations': 0,
        'car_violations': 0,
        'start_time': time.time(),
        'video_writer': None,
        'output_path': None,
        'timestamp': datetime.datetime.now().strftime("%Y%m%d_%H%M%S"),
        'tracked_vehicles': {},
        'violation_cooldown': {},
        'violation_frames': [],
        'frame_count': 0,
        'original_fps': None,
        'output_size': None,
        'vehicle_states': {},  # Reset vehicle tracking states
        'video_id': current_video_id  # Giữ lại video_id
    })
    
    print(f"✅ Lane detection data reset complete - Timestamp: {lane_detection_data['timestamp']}, video_id: {current_video_id}")


def generate_frames_lane(path_x):
    """Generate frames cho lane detection"""
    global lane_detection_active
    
    # RESET DỮ LIỆU
    reset_lane_detection_data()
    lane_detection_active = True
    
    print(f"🎬 Starting new lane detection session for: {path_x}")
    
    try:
        yolo_output = video_detection_web(path_x)
        for detection_ in yolo_output:
            if not lane_detection_active:
                print("🛑 Lane detection stopped by user request")
                break
            
            try:
                # Encode với quality cao
                encode_params = [cv2.IMWRITE_JPEG_QUALITY, 85]
                ref, buffer = cv2.imencode('.jpg', detection_, encode_params)
                
                if ref:
                    frame = buffer.tobytes()
                    yield (b'--frame\r\n'
                           b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
                else:
                    print("⚠️ Failed to encode frame, skipping...")
                    
            except Exception as e:
                print(f"⚠️ Error encoding frame: {e}")
                continue
        
        # Video đã hết hoặc bị dừng - set flag về False
        lane_detection_active = False
        print("✅ Lane detection stream ended")
                
    except Exception as e:
        lane_detection_active = False
        print(f"❌ Error in video detection: {e}")
        import traceback
        traceback.print_exc()


# OLD FUNCTION - DISABLED TO PREVENT CONFLICT WITH NEW LOGIC
# def video_detection(path_x=""):
#     cap = cv2.VideoCapture(path_x)
#     model = YOLO('best_new/vehicle.pt')
#     stt_m = 0
#     stt_ctb = 0
#     examBB = createBB.infoObject()
#     dataBienBan_M = 'BienBanNopPhatXeMay/'
#     dataBienBan_CTB = 'BienBanNopPhatXeOTo/'

#     # results = model.track(source="Videos/test4.mp4", show=True, stream=True)
#     while cap.isOpened():
#         success, frame = cap.read()
#         if success:
#             #  Dự đoán
#             results = model(frame)

#             # lấy ra frame sau khi đc gắn nhãn
#             annotated_frame = results[0].plot()

#             # lấy kích thước (height , width , _ )
#             # print("kích thước frame : ", annotated_frame.shape)

#             # Hiển thị lên
#             # cv2.imshow("Display ", annotated_frame)
#             # results = model.track(source="Videos/test4.mp4", show=True, tracker="bytetrack.yaml", stream=True)
#             for result in results:
#                 boxes = result.boxes.numpy()

#                 # Lấy tên class
#                 name = result.names

#                 # lấy tất cả các thông số trong một list tọa độ các đối tượng (x0 ,y0, x1, y1, )
#                 # print("list 1 ", boxes.xyxy)
#                 list_2 = []

#                 # Lấy tất các các thông số của nhiều đối tượng (x0, y0 , x1 , y1 , id ,độ chính xác , loại class)
#                 # print("Boxes ", boxes)

#                 for box in boxes:
#                     # TOÀN BỘ LOGIC CŨ COMMENTED TO PREVENT CONFLICT
#                     pass
#         else:
#             break
#     cv2.destroyAllWindows()


# OLD FUNCTION - DISABLED  
# def generate_frames(path_x):
#     yolo_output = video_detection(path_x)
#     for detection_ in yolo_output:
#         ref, buffer = cv2.imencode('.jpg', detection_)

#         frame = buffer.tobytes()
#         yield (b'--frame\r\n'
#                b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')


def _resolve_helmet_weights_path():
    """Find a valid helmet model weights path - CHỈ DÙNG MODEL_HELMET_V2."""
    # CHỈ SỬ DỤNG MODEL HELMET V2
    model_path = os.path.join('model_helmet_v2', 'best.pt')
    if os.path.exists(model_path):
        return model_path
    return None


def generate_frames_helmet(path_x, video_id=None):
    """
    Generate frames for helmet detection streaming with stop control
    
    Args:
        path_x: Path to video file
        video_id: Database video ID (from session, passed in to avoid request context issues)
    """
    global helmet_detection_active, helmet_detection_data
    
    try:
        cap = cv2.VideoCapture(path_x)
        if not cap.isOpened():
            print(f"Error: Could not open video {path_x}")
            return
        
        # Load YOLO model for helmet detection - CHỈ DÙNG MODEL_HELMET_V2
        weights_path = _resolve_helmet_weights_path()
        if not weights_path:
            print("❌ Helmet model weights not found! Please place 'best.pt' in 'model_helmet_v2/' folder.")
            model = None
        else:
            print(f"✅ Loading helmet model from: {weights_path}")
            model = YOLO(weights_path)
            # In thông tin model để debug
            if hasattr(model, 'names'):
                print(f"📊 Model classes: {model.names}")
                print(f"   Class mapping (THỰC TẾ): 0=head (không mũ - VI PHẠM), 1=helmet (có mũ), 2=person (người)")
        
        # Optional vehicle model for license plate assist (motorbike focus)
        vehicle_model = None
        try:
            vehicle_weights = os.path.join('model_lane', 'vehicle.pt')
            if os.path.exists(vehicle_weights):
                vehicle_model = YOLO(vehicle_weights)
                print(f"✅ Vehicle model loaded for plate assist: {vehicle_weights}")
            else:
                print("ℹ️ Vehicle model not found; plate extraction disabled")
        except Exception as e:
            print(f"⚠️ Could not load vehicle model: {e}")
        
        # Initialize EasyOCR for number plate text recognition
        try:
            import easyocr
            reader = easyocr.Reader(['en'])
            print("✅ EasyOCR initialized for helmet detection")
        except Exception as e:
            print(f"⚠️ EasyOCR initialization failed: {e}")
            reader = None
        
        # Ensure output directory exists for analyzed video
        os.makedirs('output', exist_ok=True)
        
        # ✅ LOAD DỮ LIỆU TỪ DATABASE TRƯỚC KHI RESET
        # Initialize helmet detection data ONCE khi bắt đầu
        if helmet_detection_data.get('frame_count') is None:
            print("🔧 [INIT] Initializing helmet detection data...")
            
            # TẢI DỮ LIỆU TỪ DATABASE (nếu có)
            existing_without_helmet = 0
            existing_with_helmet = 0
            existing_violations = []
            
            if violation_db is not None:
                try:
                    # Truy vấn TẤT CẢ vi phạm helmet từ database
                    query = """
                        SELECT 
                            violation_id,
                            video_id,
                            frame_number,
                            time_in_video,
                            has_helmet,
                            confidence,
                            license_plate,
                            bbox_x1, bbox_y1, bbox_x2, bbox_y2,
                            detected_at
                        FROM helmet_violations
                        ORDER BY detected_at DESC
                    """
                    db_results = violation_db.db.execute_query(query, fetch=True)
                    
                    if db_results:
                        print(f"📊 [HELMET INIT] Loaded {len(db_results)} violations from database")
                        
                        for row in db_results:
                            has_helmet = row.get('has_helmet', True)
                            
                            # Đếm theo loại
                            if has_helmet:
                                existing_with_helmet += 1
                            else:
                                existing_without_helmet += 1
                            
                            # Lưu vào violations list
                            time_seconds = row.get('time_in_video', 0)
                            violation_info = {
                                'violation_id': row.get('violation_id'),
                                'video_id': row.get('video_id'),
                                'frame_number': row.get('frame_number'),
                                'time_seconds': time_seconds,
                                'time_formatted': f"{int(time_seconds // 60):02d}:{int(time_seconds % 60):02d}",
                                'has_helmet': has_helmet,
                                'confidence': row.get('confidence'),
                                'license_plate': row.get('license_plate'),
                                'bbox': [
                                    row.get('bbox_x1'),
                                    row.get('bbox_y1'),
                                    row.get('bbox_x2'),
                                    row.get('bbox_y2')
                                ],
                                'detected_at': row.get('detected_at').strftime('%Y-%m-%d %H:%M:%S') if row.get('detected_at') else ''
                            }
                            existing_violations.append(violation_info)
                        
                        print(f"✅ [HELMET INIT] Loaded: With helmet={existing_with_helmet}, Without helmet={existing_without_helmet}")
                    else:
                        print("ℹ️ [HELMET INIT] No existing data in database")
                        
                except Exception as db_error:
                    print(f"⚠️ [HELMET INIT] Failed to load from database: {db_error}")
                    import traceback
                    traceback.print_exc()
            else:
                print("ℹ️ [HELMET INIT] No database connection - starting fresh")
            
            # SET DỮ LIỆU - GIỮ NGUYÊN SỐ LIỆU CŨ, CHỈ RESET CÁC BIẾN TẠM THỜI
            helmet_detection_data['frame_count'] = 0
            helmet_detection_data['violations'] = existing_violations  # ✅ GIỮ DỮ LIỆU CŨ
            helmet_detection_data['total_violations'] = existing_without_helmet  # ✅ GIỮ SỐ LIỆU CŨ
            helmet_detection_data['with_helmet'] = existing_with_helmet  # ✅ GIỮ SỐ LIỆU CŨ
            helmet_detection_data['without_helmet'] = existing_without_helmet  # ✅ GIỮ SỐ LIỆU CŨ
            
            # RESET CÁC BIẾN TẠM THỜI (cho phiên phân tích mới)
            helmet_detection_data['detection_cooldown'] = {}
            helmet_detection_data['frame_detections'] = set()
            helmet_detection_data['temporal_votes'] = {}
            helmet_detection_data['state_map'] = {}
            helmet_detection_data['recent_plates'] = []
            helmet_detection_data['person_states'] = {}
            helmet_detection_data['next_person_id'] = 0
            
            print(f"✅ [INIT COMPLETE] Starting counts: With={existing_with_helmet}, Without={existing_without_helmet}, Total violations={len(existing_violations)}")
        
        frame_number = 0
        
        while cap.isOpened():  # ✅ STREAM VIDEO
            # Kiểm tra nếu detection đã bị dừng
            if not helmet_detection_active:
                print("🛑 Helmet detection stopped, closing video...")
                break
            
            ret, frame = cap.read()
            if not ret:
                break
            
            frame_number += 1
            
            # Chỉ chạy detection khi helmet_detection_active = True
            if helmet_detection_active and model is not None:
                helmet_detection_data['frame_count'] = frame_number
                if helmet_detection_data.get('start_time') is None:
                    helmet_detection_data['start_time'] = time.time()
                # Lazy init video writer on first active frame
                if helmet_detection_data.get('video_writer') is None:
                    try:
                        h, w = frame.shape[:2]
                        fps = cap.get(cv2.CAP_PROP_FPS) or 25.0
                        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
                        timestamp = helmet_detection_data.get('timestamp') or datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                        out_path = os.path.join('output', f"helmet_analysis_{timestamp}.mp4")
                        writer = cv2.VideoWriter(out_path, fourcc, fps, (w, h))
                        helmet_detection_data['video_writer'] = writer
                        helmet_detection_data['output_path'] = out_path
                        helmet_detection_data['original_fps'] = fps
                        helmet_detection_data['output_size'] = (w, h)
                        print(f"🎥 [HELMET] Recording analyzed video to: {out_path} @ {fps:.2f}fps, size {(w, h)}")
                    except Exception as e:
                        print(f"⚠️ Failed to initialize video writer: {e}")
                
                # Run helmet detection với tham số tối ưu (giống Colab)
                # conf: confidence threshold (0.25 theo Colab)
                # imgsz: image size 800 (theo Colab - model được train với imgsz=800)
                # classes: chỉ detect class 0 (head) và class 1 (helmet), bỏ class 2 (person)
                results = model(frame, 
                              conf=0.50,  # ✅ TĂNG từ 0.25 lên 0.50 để giảm false positive khi frame không rõ
                              imgsz=800,  # Image size 800 - GIỐNG COLAB để nhận diện chính xác
                              classes=[0, 1],  # Chỉ detect class 0 (helmet) và 1 (head), bỏ class 2 (person)
                              verbose=False)  # Không in log chi tiết
                
                # SỬ DỤNG PLOT() MẶC ĐỊNH CỦA YOLO GIỐNG COLAB - KHÔNG VẼ THỦ CÔNG
                # YOLO plot() sẽ tự động vẽ boxes với màu sắc và labels phù hợp
                if results and len(results) > 0:
                    # Dùng plot() mặc định của YOLO - giống Colab
                    annotated_frame = results[0].plot(conf=True, labels=True, boxes=True)
                    frame = annotated_frame
                
                # QUAN TRỌNG: Cleanup frame detections và state map mỗi 50 frames
                if frame_number % 50 == 0:
                    # Cleanup frame detections - chỉ giữ 10 frames gần nhất
                    current_frame_detections = {d for d in helmet_detection_data.get('frame_detections', set()) 
                                               if any(f"_frame_{frame_number - i}" in d for i in range(10))}
                    helmet_detection_data['frame_detections'] = current_frame_detections
                    
                    # ✅ TĂNG LÊN 50 GIÂY: Cleanup temporal_votes và state_map - xóa các entry không active > 1500 frames (~50 giây)
                    active_keys = set()
                    for key in list(helmet_detection_data.get('temporal_votes', {}).keys()):
                        mem = helmet_detection_data['temporal_votes'][key]
                        if frame_number - mem.get('last_frame', 0) > 1500:
                            del helmet_detection_data['temporal_votes'][key]
                        else:
                            active_keys.add(key)
                    
                    for key in list(helmet_detection_data.get('state_map', {}).keys()):
                        if key not in active_keys:
                            state_rec = helmet_detection_data['state_map'][key]
                            if frame_number - state_rec.get('last_frame', 0) > 1500:
                                del helmet_detection_data['state_map'][key]
                    
                    print(f"🧹 [FRAME {frame_number}] Cleaned up: temporal_votes={len(helmet_detection_data.get('temporal_votes', {}))}, state_map={len(helmet_detection_data.get('state_map', {}))}")
                
                # Draw detections on frame with label-based mapping and conflict resolution
                def _norm_label(s: str) -> str:
                    return s.lower().replace('-', ' ').strip()

                def _iou(a, b) -> float:
                    ax1, ay1, ax2, ay2 = a
                    bx1, by1, bx2, by2 = b
                    inter_x1 = max(ax1, bx1)
                    inter_y1 = max(ay1, by1)
                    inter_x2 = min(ax2, bx2)
                    inter_y2 = min(ay2, by2)
                    iw = max(0, inter_x2 - inter_x1)
                    ih = max(0, inter_y2 - inter_y1)
                    inter = iw * ih
                    area_a = max(0, ax2 - ax1) * max(0, ay2 - ay1)
                    area_b = max(0, bx2 - bx1) * max(0, by2 - by1)
                    union = area_a + area_b - inter
                    return inter / union if union > 0 else 0.0

                frame_with = 0
                frame_without = 0
                for r in results:
                    boxes = r.boxes
                    if boxes is None:
                        continue

                    # First pass: collect detections with normalized labels
                    # MODEL HELMET V2 CLASSES (THỰC TẾ): 0=head (không mũ - VI PHẠM), 1=helmet (có mũ), 2=person (người)
                    dets = []  # {type: 'no_helmet'|'with_helmet'|'person'|'other', bbox, conf, cls}
                    names = getattr(r, 'names', {}) or {}
                    
                    for box in boxes:
                        cls = int(box.cls[0])
                        conf = float(box.conf[0])
                        
                        # Lấy tọa độ và đảm bảo hợp lệ
                        x1, y1, x2, y2 = box.xyxy[0].cpu().numpy()
                        x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                        
                        # Clamp to frame boundaries
                        h_frame, w_frame = frame.shape[:2]
                        x1 = max(0, min(x1, w_frame - 1))
                        y1 = max(0, min(y1, h_frame - 1))
                        x2 = max(0, min(x2, w_frame - 1))
                        y2 = max(0, min(y2, h_frame - 1))
                        
                        w = max(0, x2 - x1)
                        h = max(0, y2 - y1)
                        area = w * h
                        
                        # ✅ MAPPING THEO CLASS ID THỰC TẾ của model_helmet_v2
                        # Model thực tế: {0: 'helmet' (có mũ), 1: 'head' (không mũ - VI PHẠM), 2: 'person'}
                        # ❌ TRƯỚC ĐÂY SAI: cls==0 → no_helmet, cls==1 → with_helmet
                        # ✅ BÂY GIỜ ĐÚNG: cls==0 → with_helmet (helmet), cls==1 → no_helmet (head)
                        if cls == 0:
                            det_type = 'with_helmet'  # helmet = có mũ (KHÔNG VI PHẠM)
                        elif cls == 1:
                            det_type = 'no_helmet'  # head = không mũ = VI PHẠM
                        elif cls == 2:
                            det_type = 'person'  # person = người lái
                        else:
                            det_type = 'other'

                        # Class-specific confidence and area gates
                        # ✅ TĂNG AREA THRESHOLD: Tăng lên để loại bỏ detection quá nhỏ khi frame không rõ
                        min_area_head = 2000     # Tăng từ 1200 lên 2000 để tránh false positive
                        min_area_helmet = 1200   # Tăng từ 800 lên 1200
                        
                        # THÊM: Kiểm tra aspect ratio để loại bỏ detection không hợp lý
                        aspect_ratio = w / h if h > 0 else 0
                        
                        # ✅ SỬA LẠI: cls==0 là helmet (có mũ), cls==1 là head (không mũ)
                        if det_type == 'no_helmet':  # cls == 1 (head - không mũ - VI PHẠM)
                            # ✅ TĂNG THRESHOLD: Tăng lên để giảm false positive khi frame không rõ
                            # Tăng confidence từ 0.45 lên 0.60 để chỉ nhận diện khi chắc chắn
                            if conf < 0.7 or area < min_area_head:  # Tăng từ 0.45 lên 0.60
                                continue
                            # Kiểm tra aspect ratio hợp lý cho đầu người (0.5 - 2.0)
                            if aspect_ratio < 0.5 or aspect_ratio > 2.0:
                                continue
                        elif det_type == 'with_helmet':  # cls == 0 (helmet - có mũ)
                            # TĂNG THRESHOLD: Tăng lên để cân bằng với no_helmet
                            if conf < 0.40 or area < min_area_helmet:  # Tăng từ 0.35 lên 0.40
                                continue
                            # Aspect ratio cho helmet (0.4 - 2.5)
                            if aspect_ratio < 0.4 or aspect_ratio > 2.5:
                                continue
                        elif det_type == 'person':  # cls == 2
                            # BỎ QUA person detection - không cần vẽ
                            continue
                        else:
                            if conf < 0.3:
                                continue

                        dets.append({'type': det_type, 'bbox': (x1, y1, x2, y2), 'conf': conf, 'cls': cls})

                    # Conflict resolution: chỉ suppress khi with_helmet có confidence cao hơn RÕ RÀNG
                    # CẢI THIỆN: Chỉ suppress no_helmet khi with_helmet có confidence cao hơn ít nhất 0.2
                    suppressed_no = set()
                    for i, d_with in enumerate(d for d in dets if d['type'] == 'with_helmet'):
                        for j, d_no in enumerate(d for d in dets if d['type'] == 'no_helmet'):
                            iou = _iou(d_with['bbox'], d_no['bbox'])
                            if iou > 0.5:  # Tăng lại lên 0.5 để chỉ suppress khi overlap nhiều
                                # Chỉ suppress 'no_helmet' nếu 'with_helmet' có confidence CAO HƠN RÕ RÀNG (ít nhất 0.2)
                                if d_with['conf'] > d_no['conf'] + 0.2:  # Yêu cầu chênh lệch lớn hơn
                                    # Find absolute index of this no_helmet det to mark suppressed
                                    idx = [k for k, dd in enumerate(dets) if dd is d_no][0]
                                    suppressed_no.add(idx)
                                    print(f"🔀 [CONFLICT] Suppressed no_helmet (conf={d_no['conf']:.2f}) by with_helmet (conf={d_with['conf']:.2f}) | IoU={iou:.2f}")

                    # Same-type de-duplication (NMS-style) to avoid multiple boxes for one person
                    def _dedup_type_indices(det_type: str, iou_thresh: float = 0.6):
                        idxs = [i for i, d in enumerate(dets) if d['type'] == det_type]
                        # Sort by confidence desc
                        idxs_sorted = sorted(idxs, key=lambda i: dets[i]['conf'], reverse=True)
                        dropped = set()
                        for a in range(len(idxs_sorted)):
                            i = idxs_sorted[a]
                            if i in dropped:
                                continue
                            for b in range(a + 1, len(idxs_sorted)):
                                j = idxs_sorted[b]
                                if j in dropped:
                                    continue
                                if _iou(dets[i]['bbox'], dets[j]['bbox']) > iou_thresh:
                                    # Drop the lower-confidence duplicate j
                                    dropped.add(j)
                        return dropped

                    drop_indices = set()
                    drop_indices |= _dedup_type_indices('no_helmet', iou_thresh=0.6)
                    drop_indices |= _dedup_type_indices('with_helmet', iou_thresh=0.6)

                    # Second pass: draw and count with cooldown/grid + temporal vote logic + state hysteresis
                    counted_keys = set()  # ensure per-frame visible count is unique per position_key
                    for idx, det in enumerate(dets):
                        # Skip same-type duplicates determined above
                        if idx in drop_indices:
                            continue
                        det_type = det['type']
                        x1, y1, x2, y2 = det['bbox']
                        conf = det['conf']
                        cls = det.get('cls', -1)  # Lấy class ID để kiểm tra chính xác
                        current_bbox = (x1, y1, x2, y2)
                        
                        # ✅ DEBUG: Log để kiểm tra mapping class ID
                        if frame_number % 30 == 0:  # Log mỗi 30 frames để tránh spam
                            print(f"🔍 [DEBUG DETECTION] Frame {frame_number}: det_type={det_type}, cls={cls}, conf={conf:.2f}")

                        # compute grid and keys
                        center_x = (x1 + x2) // 2
                        center_y = (y1 + y2) // 2
                        grid_x = center_x // 300
                        grid_y = center_y // 300
                        position_key = f"{det_type}_{grid_x}_{grid_y}"
                        frame_detection_key = f"{position_key}_frame_{frame_number}"

                        # ✅ PERSON TRACKING: Match với người đã track (giống vehicle tracking trong lane detection)
                        person_id = None
                        if det_type in ('no_helmet', 'with_helmet'):
                            # Tính IoU với các người đã track
                            person_states = helmet_detection_data.get('person_states', {})
                            best_iou = 0.3  # Threshold tối thiểu để match
                            best_person_id = None
                            
                            for pid, pstate in person_states.items():
                                # ✅ TĂNG LÊN 50 GIÂY: Chỉ match với người đã thấy trong 1500 frames gần nhất (~50 giây với FPS=30)
                                # Với video 1 phút, đảm bảo cùng 1 người không bị gán ID mới
                                if frame_number - pstate.get('last_seen', 0) > 1500:
                                    continue
                                
                                last_bbox = pstate.get('last_bbox')
                                if last_bbox:
                                    # Tính IoU
                                    px1, py1, px2, py2 = last_bbox
                                    inter_x1 = max(x1, px1)
                                    inter_y1 = max(y1, py1)
                                    inter_x2 = min(x2, px2)
                                    inter_y2 = min(y2, py2)
                                    
                                    if inter_x2 > inter_x1 and inter_y2 > inter_y1:
                                        inter_area = (inter_x2 - inter_x1) * (inter_y2 - inter_y1)
                                        area1 = (x2 - x1) * (y2 - y1)
                                        area2 = (px2 - px1) * (py2 - py1)
                                        union_area = area1 + area2 - inter_area
                                        
                                        if union_area > 0:
                                            iou = inter_area / union_area
                                            if iou > best_iou:
                                                best_iou = iou
                                                best_person_id = pid
                            
                            # Nếu tìm thấy match, dùng person_id đó
                            if best_person_id is not None:
                                person_id = best_person_id
                                # Update person state
                                person_state = person_states[person_id]
                                person_state['last_seen'] = frame_number
                                person_state['last_bbox'] = current_bbox
                            else:
                                # Tạo person_id mới
                                if 'next_person_id' not in helmet_detection_data:
                                    helmet_detection_data['next_person_id'] = 0
                                person_id = f"person_{helmet_detection_data['next_person_id']}"
                                helmet_detection_data['next_person_id'] += 1
                                
                                # Khởi tạo person state
                                person_states[person_id] = {
                                    'has_violated': False,
                                    'last_seen': frame_number,
                                    'last_bbox': current_bbox,
                                    'violation_frame': None
                                }
                                helmet_detection_data['person_states'] = person_states

                        cooldown_frames = 250
                        can_count = True
                        skip_reason = ""

                        # ✅ CHECK has_violated FLAG PER PERSON (giống lane violations)
                        if det_type == 'no_helmet' and person_id is not None:
                            person_state = helmet_detection_data.get('person_states', {}).get(person_id)
                            if person_state and person_state.get('has_violated', False):
                                can_count = False
                                skip_reason = f"PERSON_ALREADY_VIOLATED (person_id={person_id})"

                        # Per-frame duplicate check (giữ lại để tránh đếm nhiều lần trong cùng frame)
                        if frame_detection_key in helmet_detection_data.get('frame_detections', set()):
                            can_count = False
                            skip_reason = "FRAME_DUP"

                        # Temporal voting: accumulate evidence across recent frames per position_key
                        if det_type in ('no_helmet', 'with_helmet'):
                            mem = helmet_detection_data.get('temporal_votes', {}).get(position_key)
                            if mem is None:
                                mem = {'window': [], 'last_frame': -1}
                                helmet_detection_data['temporal_votes'][position_key] = mem
                            # Only record one vote per frame
                            if mem.get('last_frame') != frame_number:
                                mem['last_frame'] = frame_number
                                # maintain a small sliding window
                                wdw = mem['window']
                                wdw.append('no' if det_type == 'no_helmet' else 'with')
                                # cap window size
                                if len(wdw) > 8:
                                    mem['window'] = wdw[-8:]

                            # Require stronger consensus and anti-flip for 'no_helmet'
                            wdw = mem['window']
                            last3 = wdw[-3:] if len(wdw) >= 3 else wdw
                            last4 = wdw[-4:] if len(wdw) >= 4 else wdw
                            no_votes = sum(1 for v in wdw[-6:] if v == 'no')
                            with_votes_last3 = sum(1 for v in last3 if v == 'with')
                            with_votes_last4 = sum(1 for v in last4 if v == 'with')
                            if det_type == 'no_helmet':
                                # ✅ TĂNG YÊU CẦU: Cần >=5 'no' trong last 7 frames và KHÔNG CÓ 'with' trong last 3 frames
                                # Tăng yêu cầu để giảm false positive khi frame không rõ
                                no_votes_last7 = sum(1 for v in wdw[-7:] if v == 'no') if len(wdw) >= 7 else no_votes
                                with_votes_last3 = sum(1 for v in wdw[-3:] if v == 'with') if len(wdw) >= 3 else 0
                                if len(wdw) < 7 or no_votes_last7 < 5 or with_votes_last3 > 0:  # Tăng yêu cầu để chắc chắn hơn
                                    can_count = False
                                    skip_reason = f'VOTE_INSUFFICIENT(no={no_votes_last7}/7, with={with_votes_last3}/3)'

                            # State machine with hysteresis per position
                            state_rec = helmet_detection_data.get('state_map', {}).get(position_key)
                            if state_rec is None:
                                state_rec = {'state': 'unknown', 'window': [], 'last_change': -1, 'lock_until': -1}
                                helmet_detection_data['state_map'][position_key] = state_rec
                            # Update state window (one vote per frame)
                            if (not state_rec['window']) or (state_rec.get('last_frame') != frame_number):
                                state_rec['window'].append('no' if det_type == 'no_helmet' else 'with')
                                if len(state_rec['window']) > 8:
                                    state_rec['window'] = state_rec['window'][-8:]
                                state_rec['last_frame'] = frame_number
                            # Update last bbox for stable rendering
                            state_rec['last_bbox'] = (x1, y1, x2, y2)
                            # Hysteresis transitions
                            cur_state = state_rec['state']
                            wdw_state = state_rec['window']
                            no6 = sum(1 for v in wdw_state[-6:] if v == 'no')
                            with6 = sum(1 for v in wdw_state[-6:] if v == 'with')
                            with3 = sum(1 for v in wdw_state[-3:] if v == 'with')
                            no3 = sum(1 for v in wdw_state[-3:] if v == 'no')
                            no4 = sum(1 for v in wdw_state[-4:] if v == 'no')
                            with4 = sum(1 for v in wdw_state[-4:] if v == 'with')
                            
                            if cur_state in ('unknown', 'with'):
                                # GIẢM YÊU CẦU: Chỉ cần ít nhất 3 'no' trong 4 frames và KHÔNG CÓ 'with' trong 2 frames gần nhất
                                no4 = sum(1 for v in wdw_state[-4:] if v == 'no') if len(wdw_state) >= 4 else no6
                                with2 = sum(1 for v in wdw_state[-2:] if v == 'with') if len(wdw_state) >= 2 else with4
                                if no4 >= 3 and with2 == 0:  # Giảm yêu cầu để phản ứng nhanh hơn
                                    # Transition to 'no'
                                    state_rec['state'] = 'no'
                                    state_rec['last_change'] = frame_number
                                    state_rec['lock_until'] = frame_number + 200  # Giảm lock xuống 200 frames (~8 giây)
                                    print(f"🔄 [STATE] {position_key}: 'with' → 'no' | F{frame_number} | no4={no4}, with2={with2} | Lock until {state_rec['lock_until']}")
                            elif cur_state == 'no':
                                # KHÓ CHUYỂN VỀ 'with': Cần bằng chứng RẤT MẠNH và sau khi hết lock
                                if frame_number >= state_rec.get('lock_until', -1):
                                    # Cần ít nhất 6/6 frames 'with' và hoàn toàn KHÔNG CÓ 'no' trong 4 frames gần nhất
                                    if with6 >= 6 and no4 == 0:  # Tăng từ 4 lên 6, yêu cầu 100% 'with'
                                        state_rec['state'] = 'with'
                                        state_rec['last_change'] = frame_number
                                        state_rec['lock_until'] = -1
                                        print(f"🔄 [STATE] {position_key}: 'no' → 'with' | F{frame_number}")
                                    else:
                                        # GIỮ NGUYÊN state 'no' - chỉ extend lock
                                        if with4 == 0 and no4 >= 3:
                                            # Vẫn thấy 'no' mạnh, extend lock thêm
                                            state_rec['lock_until'] = frame_number + 150
                                else:
                                    # Đang trong lock period - TUYỆT ĐỐI GIỮ state 'no'
                                    pass

                        if det_type == 'no_helmet':
                            # Skip if suppressed due to overlap with a helmet detection
                            if idx in suppressed_no:
                                # Do not draw or count suppressed red box to match visual logic
                                can_count = False
                                skip_reason = "SUPPRESSED_BY_HELMET"
                                continue
                            
                            # KHÔNG VẼ THỦ CÔNG - ĐÃ DÙNG plot() Ở TRÊN
                            # Chỉ đếm để tracking
                            if position_key not in counted_keys:
                                frame_without += 1
                                counted_keys.add(position_key)
                            
                            # ✅ TĂNG YÊU CẦU: Kiểm tra điều kiện để COUNT vi phạm vào database
                            # Tăng yêu cầu để giảm false positive khi frame không rõ
                            state_ok = helmet_detection_data['state_map'].get(position_key, {}).get('state') == 'no'
                            # Tăng yêu cầu: cần >=5 'no' trong last 7 frames và 0 'with' trong last 3 frames
                            no_votes_last7 = sum(1 for v in wdw[-7:] if v == 'no') if len(wdw) >= 7 else no_votes
                            with_votes_last3 = sum(1 for v in wdw[-3:] if v == 'with') if len(wdw) >= 3 else with_votes_last4
                            votes_ok = (no_votes_last7 >= 5 and with_votes_last3 == 0)  # Tăng yêu cầu
                            lock_active = frame_number < helmet_detection_data['state_map'].get(position_key, {}).get('lock_until', -1)
                            
                            # Chỉ COUNT khi đủ điều kiện (nhưng VẪN VẼ bbox ở trên)
                            if not (state_ok and votes_ok and lock_active):
                                # Không đủ chắc chắn - KHÔNG count vào DB
                                can_count = False
                                skip_reason = skip_reason or f'NOT_CONFIDENT(state={state_ok},votes={votes_ok},lock={lock_active})'
                            
                            # 🚗 LUÔN LUÔN chạy OCR biển số cho detection "không mũ" (dù có count hay không)
                            # Mục đích: Hiển thị biển số trên video để user nhìn thấy
                            if reader is not None:
                                try:
                                    box_w = x2 - x1
                                    box_h = y2 - y1
                                    
                                    # Extract region below detection for plate (lower 40-110% of box)
                                    plate_y1 = max(0, y1 + int(box_h * 0.4))  # Bắt đầu từ 40% box (gần cuối người)
                                    plate_y2 = min(frame.shape[0], y2 + int(box_h * 0.4))  # Extend xuống dưới thêm 40%
                                    plate_x1 = max(0, x1 - int(box_w * 0.2))  # Mở rộng trái 20%
                                    plate_x2 = min(frame.shape[1], x2 + int(box_w * 0.2))  # Mở rộng phải 20%
                                    
                                    if plate_y2 > plate_y1 and plate_x2 > plate_x1:
                                        plate_region = frame[plate_y1:plate_y2, plate_x1:plate_x2]
                                        
                                        # Debug: Vẽ vùng tìm kiếm biển số (BỎ VẼ BOX MÀU CAM)
                                        # cv2.rectangle(frame, (plate_x1, plate_y1), (plate_x2, plate_y2), (0, 165, 255), 2)
                                        
                                        if plate_region.size > 100:  # Đảm bảo vùng đủ lớn
                                            # Preprocess cho OCR tốt hơn
                                            gray_plate = cv2.cvtColor(plate_region, cv2.COLOR_BGR2GRAY)
                                            # Tăng contrast
                                            clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
                                            enhanced = clahe.apply(gray_plate)
                                            # Threshold
                                            _, thresh = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                                            
                                            # Run OCR
                                            ocr_results = reader.readtext(thresh, detail=1)
                                            
                                            if ocr_results:
                                                best_text = None
                                                best_prob = 0
                                                best_bbox_coords = None
                                                
                                                for (bbox_ocr, text, prob) in ocr_results:
                                                    # Clean text: giữ chữ và số, loại ký tự đặc biệt
                                                    cleaned = ''.join(c for c in text if c.isalnum() or c.isspace()).strip().upper()
                                                    # Biển số VN: thường 6-10 ký tự (VD: "29A12345", "DL15AE0190")
                                                    if len(cleaned.replace(' ', '')) >= 5 and prob > best_prob:
                                                        best_prob = prob
                                                        best_text = cleaned
                                                        best_bbox_coords = bbox_ocr
                                                
                                                if best_text and best_prob > 0.2:  # Lowered threshold
                                                    # Lưu biển số để dùng khi count violation
                                                    if lp_text is None:
                                                        lp_text = best_text
                                                    
                                                    print(f"🚗 [OCR] F{frame_number} | Plate: '{best_text}' | Conf: {best_prob:.2f}")
                                                    
                                                    # 🎨 VẼ BOUNDING BOX TEXT BIỂN SỐ chính xác (màu xanh lá nổi bật)
                                                    if best_bbox_coords:
                                                        try:
                                                            pts = best_bbox_coords
                                                            xs = [int(p[0]) + plate_x1 for p in pts]
                                                            ys = [int(p[1]) + plate_y1 for p in pts]
                                                            ocr_x1 = min(xs)
                                                            ocr_y1 = min(ys)
                                                            ocr_x2 = max(xs)
                                                            ocr_y2 = max(ys)
                                                            
                                                            # Vẽ bbox chính xác của text (màu xanh lá, dày)
                                                            cv2.rectangle(frame, (ocr_x1, ocr_y1), (ocr_x2, ocr_y2), (0, 255, 0), 4)
                                                        except:
                                                            pass
                                                    
                                                    # 🎨 VẼ TEXT BIỂN SỐ lớn và nổi bật
                                                    plate_label = f"{best_text}"
                                                    label_size, _ = cv2.getTextSize(plate_label, cv2.FONT_HERSHEY_SIMPLEX, 1.2, 3)
                                                    label_w, label_h = label_size
                                                    
                                                    # Vị trí text (dưới vùng tìm kiếm)
                                                    text_x = plate_x1
                                                    text_y = plate_y2 + label_h + 15
                                                    
                                                    # Vẽ nền đen cho text
                                                    cv2.rectangle(frame, 
                                                                (text_x - 5, text_y - label_h - 10), 
                                                                (text_x + label_w + 10, text_y + 5), 
                                                                (0, 0, 0), -1)
                                                    
                                                    # Vẽ viền vàng
                                                    cv2.rectangle(frame, 
                                                                (text_x - 5, text_y - label_h - 10), 
                                                                (text_x + label_w + 10, text_y + 5), 
                                                                (0, 255, 255), 2)
                                                    
                                                    # Vẽ text màu xanh lá sáng
                                                    cv2.putText(frame, plate_label, (text_x, text_y - 5),
                                                              cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 255, 0), 3)
                                except Exception as ocr_err:
                                    if frame_number % 100 == 0:  # Log ít hơn
                                        print(f"⚠️ [OCR] F{frame_number} Error: {ocr_err}")
                            
                            if can_count:
                                # ✅ SET has_violated FLAG PER PERSON (giống lane violations)
                                if person_id is not None:
                                    person_state = helmet_detection_data.get('person_states', {}).get(person_id)
                                    if person_state:
                                        person_state['has_violated'] = True
                                        person_state['violation_frame'] = frame_number
                                
                                old_count = helmet_detection_data['without_helmet']
                                helmet_detection_data['without_helmet'] += 1
                                new_count = helmet_detection_data['without_helmet']
                                helmet_detection_data['total_violations'] = helmet_detection_data['without_helmet']
                                helmet_detection_data['detection_cooldown'][position_key] = frame_number
                                helmet_detection_data['frame_detections'].add(frame_detection_key)
                                # Extend lockout when a no-helmet count happens
                                sm = helmet_detection_data['state_map'].get(position_key)
                                if sm:
                                    sm['state'] = 'no'
                                    sm['lock_until'] = max(sm.get('lock_until', -1), frame_number + 250)
                                
                                # ✅ IMPROVED: Try to extract license plate directly from detection box
                                lp_text = None
                                
                                # Method 1: OCR directly from lower region of no-helmet detection
                                if reader is not None:
                                    try:
                                        # Expand box downward to capture license plate area (typically below rider)
                                        box_h = y2 - y1
                                        box_w = x2 - x1
                                        
                                        # Extract region below detection for plate (lower 60-100% of full bike)
                                        plate_y1 = max(0, y1 + int(box_h * 0.6))
                                        plate_y2 = min(frame.shape[0], y2 + int(box_h * 0.3))  # Extend below
                                        plate_x1 = max(0, x1 - int(box_w * 0.1))  # Slightly wider
                                        plate_x2 = min(frame.shape[1], x2 + int(box_w * 0.1))
                                        
                                        if plate_y2 > plate_y1 and plate_x2 > plate_x1:
                                            plate_region = frame[plate_y1:plate_y2, plate_x1:plate_x2]
                                            
                                            if plate_region.size > 0:
                                                # Preprocess for better OCR
                                                gray_plate = cv2.cvtColor(plate_region, cv2.COLOR_BGR2GRAY)
                                                # Enhance contrast
                                                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                                                enhanced = clahe.apply(gray_plate)
                                                # Threshold
                                                _, thresh = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                                                
                                                # Run OCR
                                                ocr_results = reader.readtext(thresh)
                                                
                                                best_text = None
                                                best_prob = 0
                                                best_bbox_coords = None
                                                
                                                for (bbox_ocr, text, prob) in ocr_results:
                                                    # Clean text
                                                    cleaned = ''.join(c for c in text if c.isalnum() or c.isspace()).strip()
                                                    # Vietnamese plates: typically 4-9 characters (e.g., "YB 6433", "29A12345")
                                                    if len(cleaned) >= 4 and prob > best_prob:
                                                        best_prob = prob
                                                        best_text = cleaned
                                                        best_bbox_coords = bbox_ocr
                                                
                                                if best_text and best_prob > 0.3:  # Confidence threshold
                                                    lp_text = best_text
                                                    print(f"🚗 [OCR DIRECT] Plate detected: '{lp_text}' (conf: {best_prob:.2f})")
                                                    
                                                    # 🎨 VẼ BOUNDING BOX BIỂN SỐ (màu vàng nổi bật, dày hơn)
                                                    cv2.rectangle(frame, (plate_x1, plate_y1), (plate_x2, plate_y2), (0, 255, 255), 4)
                                                    
                                                    # Nếu có bbox chính xác từ OCR, vẽ thêm
                                                    if best_bbox_coords:
                                                        try:
                                                            # Convert bbox_ocr to absolute coordinates
                                                            pts = best_bbox_coords
                                                            xs = [int(p[0]) + plate_x1 for p in pts]
                                                            ys = [int(p[1]) + plate_y1 for p in pts]
                                                            ocr_x1 = min(xs)
                                                            ocr_y1 = min(ys)
                                                            ocr_x2 = max(xs)
                                                            ocr_y2 = max(ys)
                                                            
                                                            # Vẽ bbox chính xác của text biển số (màu xanh lá)
                                                            cv2.rectangle(frame, (ocr_x1, ocr_y1), (ocr_x2, ocr_y2), (0, 255, 0), 3)
                                                        except:
                                                            pass
                                                    
                                                    # 🎨 VẼ TEXT BIỂN SỐ với nền đen nổi bật
                                                    plate_label = f"LP: {lp_text} ({best_prob:.2f})"
                                                    label_size, _ = cv2.getTextSize(plate_label, cv2.FONT_HERSHEY_SIMPLEX, 0.8, 2)
                                                    label_w, label_h = label_size
                                                    
                                                    # Vị trí text (trên bbox)
                                                    text_x = plate_x1
                                                    text_y = plate_y1 - 10 if plate_y1 - 10 > label_h else plate_y2 + label_h + 10
                                                    
                                                    # Vẽ nền đen cho text
                                                    cv2.rectangle(frame, 
                                                                (text_x, text_y - label_h - 8), 
                                                                (text_x + label_w + 10, text_y + 5), 
                                                                (0, 0, 0), -1)
                                                    
                                                    # Vẽ text màu vàng
                                                    cv2.putText(frame, plate_label, (text_x + 5, text_y - 3),
                                                              cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2)
                                    except Exception as ocr_err:
                                        # Silent fail for OCR errors
                                        if frame_number % 50 == 0:
                                            print(f"⚠️ [OCR DIRECT] Error: {ocr_err}")
                                
                                # Method 2: Fallback to recent_plates if direct OCR failed
                                if lp_text is None:
                                    try:
                                        candidates = [p for p in helmet_detection_data.get('recent_plates', []) if frame_number - p.get('frame', 0) <= 15]
                                        if candidates:
                                            cx = (x1 + x2) // 2
                                            cy = (y1 + y2) // 2
                                            def dist2(p):
                                                bb = p.get('bbox') or (cx, cy, cx, cy)
                                                pcx = (bb[0] + bb[2]) // 2
                                                pcy = (bb[1] + bb[3]) // 2
                                                return (pcx - cx)*(pcx - cx) + (pcy - cy)*(pcy - cy)
                                            best = min(candidates, key=dist2)
                                            bb = best.get('bbox')
                                            if bb:
                                                pcx = (bb[0] + bb[2]) // 2
                                                pcy = (bb[1] + bb[3]) // 2
                                                if (pcx - cx)**2 + (pcy - cy)**2 < (250*250):
                                                    lp_text = best.get('text')
                                                    print(f"🚗 [OCR FALLBACK] Using recent plate: '{lp_text}' (distance: {((pcx - cx)**2 + (pcy - cy)**2)**0.5:.0f}px)")
                                    except Exception as fallback_err:
                                        if frame_number % 50 == 0:
                                            print(f"⚠️ [OCR FALLBACK] Error: {fallback_err}")
                                violation_info = {
                                    'violation_id': len(helmet_detection_data['violations']) + 1,
                                    'type': 'no_helmet',
                                    'frame_number': helmet_detection_data['frame_count'],
                                    'confidence': conf,
                                    'bbox': [x1, y1, x2, y2],
                                    'detected_at': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                    'license_plate': lp_text
                                }
                                helmet_detection_data['violations'].append(violation_info)
                                
                                # ✅ LƯU VÀO DATABASE (dùng video_id từ parameter, không dùng session)
                                # ✅ CHỈ LƯU ẢNH KHI PHÁT HIỆN "head" (cls == 0, không mũ) - KHÔNG LƯU KHI PHÁT HIỆN "helmet" (cls == 1, có mũ)
                                # ✅ QUAN TRỌNG: Chỉ lưu khi detection THỰC TẾ là no_helmet (cls == 0), KHÔNG dựa vào state machine
                                
                                # ✅ DEBUG: Log chi tiết để kiểm tra
                                # ✅ QUAN TRỌNG: Model thực tế: cls==0 là helmet (có mũ), cls==1 là head (không mũ - VI PHẠM)
                                # ✅ Chỉ lưu khi detection THỰC TẾ là no_helmet (cls == 1, head), KHÔNG dựa vào state machine
                                should_save = (det_type == 'no_helmet' and cls == 1)
                                
                                # ✅ DEBUG: Log để kiểm tra
                                # ✅ QUAN TRỌNG: Log CHI TIẾT để debug khi lưu ảnh
                                if not should_save:
                                    if frame_number % 30 == 0:  # Log mỗi 30 frames để tránh spam
                                        print(f"⏭️ [SKIP SAVE IMAGE] det_type={det_type}, cls={cls} | Chỉ lưu khi det_type='no_helmet' và cls=1 (head). cls=0 là helmet (có mũ)")
                                else:
                                    # ✅ LOG CHI TIẾT: Kiểm tra lại cls và det_type trước khi lưu
                                    print(f"🔍 [BEFORE SAVE CHECK] det_type={det_type}, cls={cls}, conf={conf:.2f}, position_key={position_key}")
                                    print(f"   → det dictionary: {det}")
                                    # ✅ KIỂM TRA LẠI: Đảm bảo cls == 1 (head), không phải cls == 0 (helmet)
                                    # ✅ Model thực tế: cls==0 là helmet (có mũ), cls==1 là head (không mũ)
                                    if cls != 1:
                                        print(f"❌ [ERROR] cls={cls} != 1! Không lưu ảnh. Chỉ lưu khi cls=1 (head/no_helmet). cls=0 là helmet (có mũ)")
                                        should_save = False
                                    if det_type != 'no_helmet':
                                        print(f"❌ [ERROR] det_type={det_type} != 'no_helmet'! Không lưu ảnh. Chỉ lưu khi det_type='no_helmet'")
                                        should_save = False
                                    if cls == 0:
                                        print(f"🚨 [ALERT] cls=0 (helmet/có mũ) nhưng det_type={det_type}! TUYỆT ĐỐI KHÔNG được lưu ảnh khi có helmet!")
                                        should_save = False
                                
                                # ✅ KIỂM TRA LẠI: Đảm bảo cls == 1 (head/no_helmet), không phải cls == 0 (helmet/with_helmet)
                                # ✅ QUAN TRỌNG: Kiểm tra lại cls và det_type để đảm bảo không lưu nhầm khi phát hiện helmet (cls == 0)
                                # ✅ Model thực tế: cls==0 là helmet (có mũ), cls==1 là head (không mũ)
                                # ✅ THÊM KIỂM TRA: Không lưu nếu cls == 0 (helmet) hoặc det_type == 'with_helmet'
                                if violation_db is not None and video_id is not None and should_save and cls == 1 and det_type == 'no_helmet':
                                    try:
                                        fps = helmet_detection_data.get('original_fps', 25.0)
                                        time_in_video = frame_number / fps if fps > 0 else 0
                                        
                                        # ✅ KIỂM TRA LẠI LẦN CUỐI: Đảm bảo chỉ lưu khi cls == 1 (head, không mũ)
                                        # ✅ QUAN TRỌNG: Kiểm tra lại cls và det_type để đảm bảo không lưu nhầm khi phát hiện helmet (cls == 0)
                                        # ✅ Model thực tế: cls==0 là helmet (có mũ), cls==1 là head (không mũ)
                                        # ✅ TUYỆT ĐỐI KHÔNG LƯU nếu cls == 0 (helmet) hoặc det_type == 'with_helmet'
                                        if cls == 0 or det_type == 'with_helmet':
                                            print(f"🚨 [BLOCK SAVE] cls={cls}, det_type={det_type} | TUYỆT ĐỐI KHÔNG LƯU ảnh khi phát hiện helmet! Chỉ lưu khi cls=1 (head/no_helmet)")
                                            continue
                                        if cls != 1 or det_type != 'no_helmet':
                                            print(f"❌ [ERROR] Attempted to save image but cls={cls}, det_type={det_type}! Chỉ lưu khi cls=1 (head) và det_type='no_helmet'. Skipping save.")
                                            continue
                                        
                                        os.makedirs("data_xe_vp_bh", exist_ok=True)
                                        image_path = f"data_xe_vp_bh/helmet_{new_count}.jpg"
                                        cv2.imwrite(image_path, frame)
                                        print(f"📸 [SAVE IMAGE] ✅ Saved violation image: {image_path} | det_type={det_type}, cls={cls} (head/no_helmet) | Person_ID={person_id}")
                                        
                                        # Tạo PDF biên bản vi phạm mũ bảo hiểm (luôn tạo, không cần biển số)
                                        pdf_path = None
                                        try:
                                            # Tạo PDF biên bản vi phạm mũ bảo hiểm
                                            pdf_path = create_helmet_pdf_report(
                                                frame=frame,
                                                violation_count=new_count,
                                                license_plate=lp_text if lp_text else None,
                                                output_dir="BienBanNopPhatXeMayViPhamMuBaoHiem"
                                            )
                                            if lp_text:
                                                print(f"📄 [PDF CREATED] ✅ Created helmet violation PDF: {pdf_path} (License plate: {lp_text})")
                                            else:
                                                print(f"📄 [PDF CREATED] ✅ Created helmet violation PDF: {pdf_path} (No license plate)")
                                        except Exception as pdf_err:
                                            print(f"⚠️ PDF creation error: {pdf_err}")
                                            import traceback
                                            traceback.print_exc()
                                        
                                        v_id = violation_db.insert_helmet_violation(
                                            video_id=video_id,
                                            frame_number=frame_number,
                                            time_in_video=time_in_video,
                                            has_helmet=False,  # VI PHẠM = không mũ
                                            confidence=conf,
                                            license_plate=lp_text,
                                            bbox=[x1, y1, x2, y2],
                                            image_path=image_path,
                                            pdf_report_path=pdf_path
                                        )
                                        
                                        if v_id:
                                            print(f"✅✅✅ [HELMET DB] Violation saved! v_id={v_id}, plate={lp_text}")
                                        else:
                                            print(f"❌ [HELMET DB] insert returned None")
                                    except Exception as db_err:
                                        print(f"❌❌❌ [HELMET DB ERROR] {db_err}")
                                        import traceback
                                        traceback.print_exc()
                                else:
                                    if not violation_db:
                                        print(f"⚠️ [HELMET DB SKIP] violation_db is None")
                                    if not video_id:
                                        print(f"⚠️ [HELMET DB SKIP] video_id is None")
                                
                                cooldown_count = len(helmet_detection_data['detection_cooldown'])
                                print(f"🚨 [KHÔNG MŨ ➕] {old_count} → {new_count} | F{frame_number} | Person_ID={person_id} | G({grid_x},{grid_y}) | Cooldowns: {cooldown_count}")
                            else:
                                if frame_number % 30 == 0:
                                    print(f"⏭️ [KHÔNG MŨ SKIP] F{frame_number} | G({grid_x},{grid_y}) | {skip_reason}")
                        elif det_type == 'with_helmet':
                            # KHÔNG VẼ THỦ CÔNG - ĐÃ DÙNG plot() Ở TRÊN
                            # Chỉ đếm để tracking
                            if position_key not in counted_keys:
                                frame_with += 1
                                counted_keys.add(position_key)
                            if can_count:
                                old_count = helmet_detection_data['with_helmet']
                                helmet_detection_data['with_helmet'] += 1
                                new_count = helmet_detection_data['with_helmet']
                                helmet_detection_data['detection_cooldown'][position_key] = frame_number
                                helmet_detection_data['frame_detections'].add(frame_detection_key)
                                cooldown_count = len(helmet_detection_data['detection_cooldown'])
                                print(f"✅ [CÓ MŨ ➕] {old_count} → {new_count} | F{frame_number} | G({grid_x},{grid_y}) | Cooldowns: {cooldown_count}")
                            else:
                                if frame_number % 30 == 0:
                                    print(f"⏭️ [CÓ MŨ SKIP] F{frame_number} | G({grid_x},{grid_y}) | {skip_reason}")
                
                        elif det_type == 'person':
                            # BỎ VẼ BOUNDING BOX CHO PERSON - model_helmet_v2 class 2
                            pass

                # Optional: vehicle-based license plate extraction every few frames
                try:
                    if vehicle_model is not None and (frame_number % 5 == 0):
                        v_results = vehicle_model(frame, verbose=False)
                        for vr in v_results:
                            vboxes = getattr(vr, 'boxes', None)
                            if vboxes is None:
                                continue
                            for vbox in vboxes:
                                vcls = int(vbox.cls[0])
                                vconf = float(vbox.conf[0])
                                # Heuristic: class 1 used for motorbikes in lane logic
                                if vconf < 0.35 or vcls not in [1]:
                                    continue
                                vx1, vy1, vx2, vy2 = map(int, vbox.xyxy[0])
                                crop_h = max(0, vy2 - vy1)
                                if crop_h < 20:
                                    continue
                                region_y1 = vy1 + int(crop_h * 0.6)
                                plate_region = frame[region_y1:vy2, vx1:vx2]
                                if reader is None or plate_region.size == 0:
                                    continue
                                try:
                                    gray = cv2.cvtColor(plate_region, cv2.COLOR_BGR2GRAY)
                                    _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                                    ocr_results = reader.readtext(th)
                                except Exception:
                                    ocr_results = []
                                best_text = None
                                best_prob = 0
                                best_bbox = None
                                for (bbox_pts, text, prob) in (ocr_results or []):
                                    cleaned = ''.join(c for c in text if c.isalnum()).upper()
                                    if len(cleaned) >= 4 and prob > best_prob:
                                        best_prob = prob
                                        best_text = cleaned
                                        xs = [p[0] for p in bbox_pts]
                                        ys = [p[1] for p in bbox_pts]
                                        px1 = int(min(xs)) + vx1
                                        py1 = int(min(ys)) + region_y1
                                        px2 = int(max(xs)) + vx1
                                        py2 = int(max(ys)) + region_y1
                                        best_bbox = (px1, py1, px2, py2)
                                if best_text:
                                    # Draw on frame and store recent
                                    # cv2.rectangle(frame, (vx1, vy1), (vx2, vy2), (255, 255, 0), 1)  # BỎ VẼ BOX XE MÀU VÀNG
                                    if best_bbox:
                                        cv2.rectangle(frame, (best_bbox[0], best_bbox[1]), (best_bbox[2], best_bbox[3]), (0, 255, 255), 2)
                                    cv2.putText(frame, f"LP:{best_text}", (vx1, max(0, vy1-5)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)
                                    helmet_detection_data['recent_plates'].append({'text': best_text, 'bbox': best_bbox, 'frame': frame_number})
                        # Keep only last N
                        helmet_detection_data['recent_plates'] = helmet_detection_data['recent_plates'][-30:]
                except Exception as e:
                    if frame_number % 60 == 0:
                        print(f"⚠️ Plate extraction error: {e}")
                
                # KHÔNG VẼ THỦ CÔNG - ĐÃ DÙNG plot() Ở TRÊN
                # Chỉ đếm stable states để tracking
                try:
                    smap = helmet_detection_data.get('state_map', {})
                    recent_limit = 8  # frames to keep state visible
                    stable_keys = [k for k, m in smap.items() if m.get('state') in ('no', 'with') and (frame_number - m.get('last_frame', -1) <= recent_limit)]
                    # Chỉ đếm, không vẽ (đã vẽ bằng plot())
                    for k in stable_keys:
                        if k in counted_keys:
                            continue
                        m = smap[k]
                        if m.get('state') == 'no':
                            frame_without += 1
                        else:
                            frame_with += 1
                        counted_keys.add(k)
                except Exception:
                    pass

                # Update current counts based on stable states (non-flickering)
                helmet_detection_data['current_with'] = frame_with
                helmet_detection_data['current_without'] = frame_without
                helmet_detection_data['current_total'] = frame_with + frame_without

                # BỎ PHẦN THỐNG KÊ TÍCH LŨY TRÊN VIDEO - Chỉ hiển thị ở web UI bên phải

                # Cleanup cooldown mỗi 150 frames (5 giây) - XÓA COOLDOWN CŨ HƠN 250 FRAMES
                if frame_number % 150 == 0:
                    expired_keys = [k for k, v in helmet_detection_data.get('detection_cooldown', {}).items() 
                                  if frame_number - v > 250]  # Tăng lên 250 frames để cooldown bền vững hơn
                    for k in expired_keys:
                        del helmet_detection_data['detection_cooldown'][k]
                    
                    # ✅ CLEANUP person_states - TĂNG LÊN 50 GIÂY: xóa người không thấy > 1500 frames (~50 giây với FPS=30)
                    # Với video 1 phút, đảm bảo cùng 1 người không bị xóa và gán ID mới
                    person_states = helmet_detection_data.get('person_states', {})
                    expired_persons = [pid for pid, pstate in person_states.items() 
                                     if frame_number - pstate.get('last_seen', 0) > 1500]
                    for pid in expired_persons:
                        del person_states[pid]
                    
                    active_cooldowns = len(helmet_detection_data.get('detection_cooldown', {}))
                    active_persons = len(person_states)
                    print(f"🧹 [HELMET] Frame {frame_number}: Cleaned {len(expired_keys)} cooldowns, {len(expired_persons)} persons | Active: {active_cooldowns} cooldowns, {active_persons} persons")

                    # ✅ TĂNG LÊN 50 GIÂY: Cleanup temporal votes too (stale > 1500 frames)
                    tv = helmet_detection_data.get('temporal_votes', {})
                    stale = [k for k, m in tv.items() if frame_number - (m.get('last_frame', -1)) > 1500]
                    for k in stale:
                        del tv[k]

                    # ✅ TĂNG LÊN 50 GIÂY: Cleanup state_map entries not updated for > 1500 frames
                    smap = helmet_detection_data.get('state_map', {})
                    stale_states = [k for k, m in smap.items() if frame_number - m.get('last_frame', -1) > 1500]
                    for k in stale_states:
                        del smap[k]
                
                # Write analyzed frame to output video
                try:
                    vw = helmet_detection_data.get('video_writer')
                    if vw is not None:
                        vw.write(frame)
                except Exception as e:
                    if frame_number % 60 == 0:
                        print(f"⚠️ Error writing frame to video: {e}")
                        
            else:
                # Không chạy detection - chỉ hiển thị video thuần
                # Thêm text thông báo
                status_text = "CHUA BAT DAU PHAT HIEN" if model is not None else "THIEU MODEL MU BAO HIEM"
                cv2.putText(frame, status_text, (50, 50), 
                          cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 255, 255), 3)
            
            # Encode frame as JPEG
            ret, buffer = cv2.imencode('.jpg', frame)
            if ret:
                frame_bytes = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
        
        # Đảm bảo cap được đóng
        if cap is not None:
            try:
                cap.release()
                print("✅ VideoCapture released in generate_frames_helmet")
            except Exception as e:
                print(f"⚠️ Error releasing VideoCapture: {e}")
        
        cv2.destroyAllWindows()
        
    except Exception as e:
        print(f"Error in generate_frames_helmet: {e}")
        import traceback
        traceback.print_exc()
        
        # Đảm bảo cap được đóng ngay cả khi có lỗi
        if 'cap' in locals() and cap is not None:
            try:
                cap.release()
                print("✅ VideoCapture released in exception handler")
            except:
                pass
        
        # Return a simple error frame
        error_frame = np.zeros((480, 640, 3), dtype=np.uint8)
        cv2.putText(error_frame, "Error loading video", (50, 240), 
                   cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
        ret, buffer = cv2.imencode('.jpg', error_frame)
        if ret:
            frame_bytes = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/Hethongcamera2")
def camera_2():
    return render_template("HelmetViolate.html")


@app.route("/bb")
def bb():
    return render_template("bb.html")


@app.route("/thongke")
def tk():
    return render_template("thongke.html")


@app.route("/Hethongcamera1")
def camera_1():
    return render_template("LaneViolate.html")


@app.route("/camera1")
def video():
    global lane_detection_data
    
    # Check if video has been uploaded for lane detection
    uploaded_video = session.get('uploaded_video_lane')
    if not uploaded_video:
        return "No video uploaded for lane detection", 400
    
    # Get video_id from session
    video_id = session.get('current_video_id_lane')
    if video_id:
        lane_detection_data['video_id'] = video_id
        print(f"✅ Set video_id for lane detection: {video_id}")
    
    return Response(generate_frames_lane(path_x=uploaded_video),
                    mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route("/camera2")
def video_2():
    # Check if there's an uploaded video for helmet detection
    uploaded_video = session.get('uploaded_video_helmet', None)
    if not uploaded_video:
        print("DEBUG camera2: No uploaded helmet video in session")
        return Response(generate_placeholder_stream(),
                        mimetype='multipart/x-mixed-replace; boundary=frame')
    
    # Check if file exists
    if not os.path.exists(uploaded_video):
        print(f"DEBUG camera2: File {uploaded_video} does not exist")
        return Response(generate_placeholder_stream(),
                        mimetype='multipart/x-mixed-replace; boundary=frame')
    
    print(f"DEBUG camera2: Starting helmet video stream for {uploaded_video}")
    
    # Get video_id from session to pass to generator (avoid request context issues)
    video_id = session.get('uploaded_video_helmet_id', None)
    print(f"DEBUG camera2: video_id from session: {video_id}")
    
    # Use helmet detection streaming with video_id
    return Response(generate_frames_helmet(uploaded_video, video_id=video_id),
                    mimetype='multipart/x-mixed-replace; boundary=frame')


# Helper function to check allowed file
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Route for red light violation detection page
@app.route("/test_upload")
def test_upload():
    return send_file("test_upload.html")

@app.route("/Hethongcamera3")
def camera_3():
    return render_template("RedLightViolate.html")


# Route for red light video stream
@app.route("/camera3")
def video_3():
    # Debug session info
    print(f"DEBUG camera3: Session keys: {list(session.keys())}")
    print(f"DEBUG camera3: uploaded_video = {session.get('uploaded_video', 'NOT_FOUND')}")
    
    # Check if there's an uploaded video
    uploaded_video = session.get('uploaded_video', None)
    if not uploaded_video:
        print("DEBUG: No uploaded video in session, showing placeholder")
        # Instead of returning 400 error, return a placeholder stream
        return Response(generate_placeholder_stream(),
                        mimetype='multipart/x-mixed-replace; boundary=frame')
    
    # Check if file exists
    if not os.path.exists(uploaded_video):
        print(f"DEBUG: File {uploaded_video} does not exist, showing placeholder")
        return Response(generate_placeholder_stream(),
                        mimetype='multipart/x-mixed-replace; boundary=frame')
    
    print(f"DEBUG: Starting video stream for {uploaded_video}")
    
    # Get video_id from session (if database is connected)
    video_id = session.get('uploaded_video_id', None)
    print(f"🔍 [STREAMING DEBUG] video_id from session: {video_id}")
    print(f"🔍 [STREAMING DEBUG] violation_db is None: {violation_db is None}")
    
    # Check if user wants advanced detection
    use_advanced = session.get('red_light_advanced', False)
    
    if use_advanced:
        # Use advanced streaming with license plate detection
        # Pass video_id and violation_db to save violations immediately
        return Response(generate_frames_red_light_new(uploaded_video, video_id=video_id, violation_db_instance=violation_db),
                        mimetype='multipart/x-mixed-replace; boundary=frame')
    else:
        # Use basic streaming
        return Response(generate_frames_red_light(path_x=uploaded_video),
                        mimetype='multipart/x-mixed-replace; boundary=frame')


# Route to process red light video and export results
@app.route("/process_red_light_video", methods=['POST'])
def process_red_light_video():
    """Process uploaded video and export analysis results"""
    print("DEBUG process_red_light_video: Processing request received")
    
    # Check if there's an uploaded video
    uploaded_video = session.get('uploaded_video', None)
    if not uploaded_video:
        return jsonify({'error': 'No video uploaded'}), 400
    
    # Check if file exists
    if not os.path.exists(uploaded_video):
        return jsonify({'error': 'Video file not found'}), 400
    
    # Get video_id from session (if database is connected)
    video_id = session.get('uploaded_video_id', None)
    print(f"🔍 [DEBUG - PROCESS] video_id from session: {video_id}")
    print(f"🔍 [DEBUG - PROCESS] video_db is None: {video_db is None}")
    print(f"🔍 [DEBUG - PROCESS] violation_db is None: {violation_db is None}")
    
    try:
        # Process the video using the new system with video_id and violation_db
        print(f"🔍 [DEBUG - PROCESS] Calling process_red_light_video_complete with video_id={video_id}")
        print(f"🔍 [DEBUG - PROCESS] Passing violation_db instance: {violation_db is not None}")
        
        output_video, violation_count = process_red_light_video_complete(
            uploaded_video, 
            output_dir="output",
            video_id=video_id,
            violation_db_instance=violation_db  # Pass violation_db instance
        )
        
        # Update video status to 'completed' in database
        if video_db is not None and video_id is not None:
            try:
                video_db.update_video_status(video_id, 'completed')
                print(f"✅ [CAMERA 3 - PROCESS] Video status updated to completed: video_id={video_id}")
            except Exception as e:
                print(f"⚠️ [CAMERA 3 - PROCESS] Failed to update video status: {e}")
        
        # Get the CSV file (should be the latest one in output directory)
        output_dir = "output"
        csv_files = [f for f in os.listdir(output_dir) if f.startswith('violations_data_') and f.endswith('.csv')]
        
        if csv_files:
            # Get the most recent CSV file
            latest_csv = max(csv_files, key=lambda x: os.path.getctime(os.path.join(output_dir, x)))
            csv_path = os.path.join(output_dir, latest_csv)
        else:
            csv_path = None
        
        return jsonify({
            'success': True,
            'output_video': output_video,
            'violation_count': violation_count,
            'csv_path': csv_path,
            'video_id': video_id,
            'message': f'Video processed successfully. Found {violation_count} violations.'
        })
        
    except Exception as e:
        print(f"ERROR processing video: {str(e)}")
        # Update video status to 'failed' in database
        if video_db is not None and video_id is not None:
            try:
                video_db.update_video_status(video_id, 'failed')
                print(f"⚠️ [CAMERA 3 - PROCESS] Video status updated to failed: video_id={video_id}")
            except Exception as db_error:
                print(f"⚠️ [CAMERA 3 - PROCESS] Failed to update video status to failed: {db_error}")
        return jsonify({'error': f'Processing failed: {str(e)}'}), 500

# Route to download processed video
@app.route("/download_processed_video")
def download_processed_video():
    """Download the processed video file"""
    output_dir = "output"
    video_files = [f for f in os.listdir(output_dir) if f.startswith('processed_video_') and f.endswith('.mp4')]
    
    if video_files:
        # Get the most recent video file
        latest_video = max(video_files, key=lambda x: os.path.getctime(os.path.join(output_dir, x)))
        video_path = os.path.join(output_dir, latest_video)
        return send_file(video_path, as_attachment=True, download_name=latest_video)
    else:
        return "No processed video found", 404

# Route to download violations CSV
@app.route("/download_violations_csv")
def download_violations_csv():
    """Download the violations CSV file"""
    output_dir = "output"
    csv_files = [f for f in os.listdir(output_dir) if f.startswith('violations_data_') and f.endswith('.csv')]
    
    if csv_files:
        # Get the most recent CSV file
        latest_csv = max(csv_files, key=lambda x: os.path.getctime(os.path.join(output_dir, x)))
        csv_path = os.path.join(output_dir, latest_csv)
        return send_file(csv_path, as_attachment=True, download_name=latest_csv)
    else:
        return "No violations data found", 404

# Route to upload video for analysis
@app.route("/upload_video", methods=['POST'])
def upload_video():
    print("DEBUG upload_video: Upload request received")
    
    if 'video' not in request.files:
        print("DEBUG upload_video: No video in request files")
        return jsonify({'error': 'No video file provided'}), 400
    
    file = request.files['video']
    if file.filename == '':
        print("DEBUG upload_video: Empty filename")
        return jsonify({'error': 'No selected file'}), 400
    
    print(f"DEBUG upload_video: File received: {file.filename}")
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        print(f"DEBUG upload_video: Saving to: {filepath}")
        
        file.save(filepath)
        
        # Store in session
        session['uploaded_video'] = filepath
        session['red_light_advanced'] = True
        
        # Save to database if connection is available (Camera 3 = Red Light)
        video_id = None
        print(f"🔍 [DEBUG - UPLOAD] video_db is None: {video_db is None}")
        
        if video_db is not None:
            try:
                print(f"🔍 [DEBUG - UPLOAD] Calling insert_video with camera_id=3, video_filename={filename}")
                video_id = video_db.insert_video(
                    camera_id=3,  # Camera 3 for red light detection
                    video_filename=filename,  # Changed from 'filename' to 'video_filename'
                    video_path=filepath,      # Changed from 'filepath' to 'video_path'
                    file_size_mb=None,
                    duration_seconds=None,
                    fps=None,
                    resolution=None
                )
                session['uploaded_video_id'] = video_id
                print(f"✅✅✅ [UPLOAD SUCCESS] Video saved to database: video_id={video_id}")
                print(f"✅ [CAMERA 3 - UPLOAD] Session updated with video_id={video_id}")
            except Exception as e:
                print(f"❌❌❌ [UPLOAD ERROR] Failed to save video to database: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("⚠️⚠️⚠️ [CAMERA 3 - UPLOAD] Database not connected (video_db is None)")
        
        print(f"DEBUG upload_video: Session updated - uploaded_video = {session.get('uploaded_video')}")
        print(f"DEBUG upload_video: Session keys: {list(session.keys())}")
        
        # Always use advanced method (since we removed basic option)
        detection_method = 'advanced'
        
        return jsonify({
            'success': True, 
            'filename': filename,
            'detection_method': detection_method,
            'filepath': filepath,  # Add this for debugging
            'video_id': video_id,  # Include video_id in response
            'message': f'Video uploaded successfully! Using {detection_method} detection method.'
        })
    
    print(f"DEBUG upload_video: Invalid file type for {file.filename}")
    return jsonify({'error': 'Invalid file type'}), 400


# Route to upload video for helmet detection
@app.route("/upload_video_helmet", methods=['POST'])
def upload_video_helmet():
    if 'video' not in request.files:
        return jsonify({'error': 'No video file provided'}), 400
    
    file = request.files['video']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'helmet_' + filename)
        file.save(filepath)
        
        # Generate unique ID for this processing job
        job_id = str(uuid.uuid4())
        
        # Create output path
        os.makedirs('processed_videos', exist_ok=True)
        output_path = os.path.join('processed_videos', f'{job_id}_processed.mp4')
        
        # Save to database if connection is available (Camera 2 = Helmet)
        video_id = None
        print(f"🔍 [DEBUG - HELMET UPLOAD] video_db is None: {video_db is None}")
        
        if video_db is not None:
            try:
                print(f"🔍 [DEBUG - HELMET UPLOAD] Calling insert_video with camera_id=2, video_filename={filename}")
                video_id = video_db.insert_video(
                    camera_id=2,  # Camera 2 for helmet detection
                    video_filename=filename,
                    video_path=filepath,
                    file_size_mb=None,
                    duration_seconds=None,
                    fps=None,
                    resolution=None
                )
                session['uploaded_video_helmet_id'] = video_id
                print(f"✅✅✅ [HELMET UPLOAD SUCCESS] Video saved to database: video_id={video_id}")
                print(f"✅ [CAMERA 2 - UPLOAD] Session updated with video_id={video_id}")
            except Exception as e:
                print(f"❌❌❌ [HELMET UPLOAD ERROR] Failed to save video to database: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("⚠️⚠️⚠️ [CAMERA 2 - UPLOAD] Database not connected (video_db is None)")
        
        # Store processing info in session
        session['helmet_processing'] = {
            'job_id': job_id,
            'status': 'processing',
            'input_path': filepath,
            'output_path': output_path,
            'video_id': video_id
        }
        
        # Store uploaded video path for streaming
        session['uploaded_video_helmet'] = filepath
        
        # For demonstration, process synchronously
        # In production, use a task queue like Celery for async processing
        session['helmet_processing']['message'] = 'Processing video, please wait...'
        
        # Return immediately to show processing status
        response = jsonify({
            'success': True, 
            'filename': filename,
            'job_id': job_id,
            'video_id': video_id,  # Include video_id in response
            'message': 'Video đang được xử lý. Vui lòng đợi...',
            'processing_url': '/process_helmet_now'
        })
        
        # Get detection method from form
        detection_method = request.form.get('detection_method', 'original')
        use_advanced = detection_method == 'advanced'
        
        # Store filepath for processing
        session['pending_helmet_process'] = {
            'input': filepath,
            'output': output_path,
            'job_id': job_id,
            'use_advanced': use_advanced,
            'video_id': video_id  # Pass video_id for database operations
        }
        
        return response
    
    return jsonify({'error': 'Invalid file type'}), 400


# Route to actually process the helmet video
@app.route("/process_helmet_now", methods=['GET'])
def process_helmet_now():
    pending = session.get('pending_helmet_process')
    if not pending:
        return jsonify({'error': 'No pending process'}), 400
    
    try:
        print("🚀 Starting helmet video processing...")
        
        # Process the video using selected detection method
        use_advanced = pending.get('use_advanced', False)
        result_path, stats = process_helmet_video_complete(
            pending['input'], 
            pending['output'],
            use_improved_detection=use_advanced
        )
        
        print(f"✅ Helmet video processing completed! Output: {result_path}")
        
        # Update session
        session['helmet_processing']['status'] = 'completed'
        session['helmet_processing']['stats'] = stats
        session['helmet_processing']['output_path'] = result_path
        session['uploaded_video_helmet'] = result_path
        
        # Clear pending
        session.pop('pending_helmet_process', None)
        
        return jsonify({
            'success': True,
            'stats': stats,
            'message': 'Processing completed!'
        })
    except Exception as e:
        print(f"❌ Error in helmet processing: {str(e)}")
        session['helmet_processing']['status'] = 'error'
        session['helmet_processing']['error'] = str(e)
        return jsonify({'error': str(e)}), 500


# Route to upload video for lane detection
@app.route("/upload_video_lane", methods=['POST'])
def upload_video_lane():
    global video_db
    
    if 'video' not in request.files:
        return jsonify({'error': 'No video file provided'}), 400
    
    file = request.files['video']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'lane_' + filename)
        file.save(filepath)
        
        # Store in session
        session['uploaded_video_lane'] = filepath
        
        # Insert video vào database
        video_id = None
        if video_db:
            try:
                # Get video info
                file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
                
                video_id = video_db.insert_video(
                    camera_id=1,  # Camera 1 for lane detection
                    video_filename=filename,
                    video_path=filepath,
                    file_size_mb=round(file_size_mb, 2)
                )
                
                if video_id:
                    session['current_video_id_lane'] = video_id
                    print(f"✅ Video inserted to DB: video_id={video_id}")
            except Exception as e:
                print(f"⚠️ Failed to insert video to DB: {e}")
        
        return jsonify({
            'success': True, 
            'filename': filename,
            'video_id': video_id,
            'message': 'Video uploaded successfully for lane detection!'
        })
    
    return jsonify({'error': 'Invalid file type'}), 400


# API để dừng lane detection và xuất kết quả
@app.route("/stop_lane_detection", methods=['POST'])
def stop_lane_detection():
    """Dừng lane detection và trả về kết quả đã được export tự động"""
    global lane_detection_active, lane_detection_data
    
    if not lane_detection_active:
        return jsonify({'error': 'Lane detection is not running'}), 400
    
    print("🛑 Nhận yêu cầu dừng từ web...")
    lane_detection_active = False
    
    # Đợi để video detection dừng hoàn toàn và export tự động chạy
    print("⏳ Đang đợi video detection dừng và export kết quả...")
    max_wait = 10  # Đợi tối đa 10 giây
    waited = 0
    while waited < max_wait:
        # Kiểm tra xem video writer đã được release chưa (đánh dấu export đã xong)
        if lane_detection_data.get('video_writer') is None and lane_detection_data.get('csv_filename'):
            print("✅ Video detection đã dừng và export hoàn tất!")
            break
        time.sleep(0.5)
        waited += 0.5
    
    if waited >= max_wait:
        print("⚠️ Timeout khi đợi export hoàn tất, trả về kết quả hiện tại...")
    
    try:
        violations_data = lane_detection_data.get('violations', [])
        output_path = lane_detection_data.get('output_path', '')
        csv_filename = lane_detection_data.get('csv_filename')
        xlsx_filename = lane_detection_data.get('xlsx_filename')
        json_filename = lane_detection_data.get('json_filename')
        
        # Kiểm tra file có tồn tại không
        video_exists = output_path and os.path.exists(output_path)
        csv_exists = csv_filename and os.path.exists(csv_filename)
        xlsx_exists = xlsx_filename and os.path.exists(xlsx_filename)
        json_exists = json_filename and os.path.exists(json_filename)
        
        # Inspect output video to report exact info
        actual_video_info = None
        try:
            if video_exists:
                cap_out = cv2.VideoCapture(output_path)
                out_frames = int(cap_out.get(cv2.CAP_PROP_FRAME_COUNT))
                out_fps = cap_out.get(cv2.CAP_PROP_FPS) or lane_detection_data.get('original_fps')
                out_duration = out_frames / (out_fps if out_fps > 0 else 1)
                cap_out.release()
                actual_video_info = {
                    'frames': out_frames,
                    'fps': out_fps,
                    'duration_seconds': out_duration,
                    'path': output_path,
                    'size_mb': os.path.getsize(output_path) / (1024 * 1024)
                }
                print(f"📹 Video info: {out_frames} frames, {out_fps:.2f} fps, {out_duration:.2f}s, {actual_video_info['size_mb']:.2f}MB")
        except Exception as e:
            print(f"⚠️ Unable to inspect output video: {e}")

        # Tạo response
        response_data = {
            'success': True,
            'message': 'Lane detection stopped successfully',
            'summary': {
                'total_violations': len(violations_data),
                'motor_violations': lane_detection_data.get('motor_violations', 0),
                'car_violations': lane_detection_data.get('car_violations', 0),
                'processing_time': time.time() - lane_detection_data.get('start_time', time.time()),
                'frames_processed': lane_detection_data.get('frame_count', 0),
            },
            'files': {
                'video': os.path.basename(output_path) if video_exists else None,
                'csv': os.path.basename(csv_filename) if csv_exists else None,
                'excel': os.path.basename(xlsx_filename) if xlsx_exists else None,
                'json': os.path.basename(json_filename) if json_exists else None
            },
            'files_exist': {
                'video': video_exists,
                'csv': csv_exists,
                'excel': xlsx_exists,
                'json': json_exists
            },
            'video_info': actual_video_info
        }
        
        print(f"✅ Trả về kết quả cho client:")
        print(f"   - Video: {os.path.basename(output_path) if video_exists else 'N/A'} ({'Có' if video_exists else 'Không'})")
        print(f"   - CSV: {os.path.basename(csv_filename) if csv_exists else 'N/A'} ({'Có' if csv_exists else 'Không'})")
        print(f"   - Excel: {os.path.basename(xlsx_filename) if xlsx_exists else 'N/A'} ({'Có' if xlsx_exists else 'Không'})")
        print(f"   - JSON: {os.path.basename(json_filename) if json_exists else 'N/A'} ({'Có' if json_exists else 'Không'})")
        
        return jsonify(response_data)
        
    except Exception as e:
        print(f"❌ Lỗi khi trả kết quả: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error retrieving results: {str(e)}'}), 500


# API để lấy trạng thái hiện tại
@app.route("/lane_detection_status", methods=['GET'])
def get_lane_detection_status():
    """Lấy thống kê từ DATABASE MySQL thay vì từ video đang chạy"""
    global lane_detection_active, lane_detection_data, violation_db
    
    # Tính runtime
    if lane_detection_data.get('start_time') and isinstance(lane_detection_data['start_time'], (int, float)):
        current_time = time.time()
        runtime_seconds = current_time - lane_detection_data['start_time']
    else:
        runtime_seconds = 0
    
    # ========== LẤY THỐNG KÊ TỪ DATABASE ==========
    motor_violations_db = 0
    car_violations_db = 0
    total_violations_db = 0
    data_source = "Memory"
    
    if violation_db:
        try:
            # ===== LUÔN LUÔN LẤY TỔNG TẤT CẢ VI PHẠM (TÍCH LŨY) =====
            # Không phân biệt video nào, đếm TỔNG từ database
            query_all = """
                SELECT 
                    violation_type,
                    COUNT(*) as count
                FROM lane_violations
                WHERE camera_id = 1
                GROUP BY violation_type
            """
            results_all = violation_db.db.execute_query(query_all, fetch=True)
            
            if results_all:
                for row in results_all:
                    if row['violation_type'] == 'motor_in_car_lane':
                        motor_violations_db = row['count']
                    elif row['violation_type'] == 'car_in_motor_lane':
                        car_violations_db = row['count']
                
                total_violations_db = motor_violations_db + car_violations_db
                data_source = "Database (TỔNG TÍCH LŨY)"
                print(f"📊 [STATUS API] TỔNG từ DB: Motor={motor_violations_db}, Car={car_violations_db}, Total={total_violations_db}")
                    
        except Exception as e:
            print(f"❌ [STATUS API] Lỗi khi truy vấn database: {e}")
            import traceback
            traceback.print_exc()
            # Fallback về memory
            motor_violations_db = lane_detection_data.get('motor_violations', 0)
            car_violations_db = lane_detection_data.get('car_violations', 0)
            total_violations_db = len(lane_detection_data.get('violations', []))
            data_source = "Memory (DB error)"
    else:
        # Không có database, dùng memory
        motor_violations_db = lane_detection_data.get('motor_violations', 0)
        car_violations_db = lane_detection_data.get('car_violations', 0)
        total_violations_db = len(lane_detection_data.get('violations', []))
        data_source = "Memory (No DB)"
    
    # Debug info
    debug_info = {
        'data_source': data_source,
        'active_cooldowns': len(lane_detection_data.get('violation_cooldown', {})),
        'active_vehicles': len(lane_detection_data.get('vehicle_states', {})),
        'violated_vehicles': sum(1 for state in lane_detection_data.get('vehicle_states', {}).values() if state.get('has_violated', False)),
        'frame_count': lane_detection_data.get('frame_count', 0),
        'timestamp': lane_detection_data.get('timestamp', 'N/A'),
        'video_id': lane_detection_data.get('video_id', 'N/A'),
        'db_available': violation_db is not None
    }
    
    # ✅ Lấy danh sách vi phạm gần đây (10 vi phạm mới nhất)
    recent_violations = []
    if violation_db:
        try:
            camera_id = 1  # Camera 1 = Lane
            query = """
                SELECT 
                    violation_id,
                    frame_number,
                    time_in_video,
                    violation_type,
                    vehicle_type,
                    confidence,
                    detected_at,
                    TIME(detected_at) as time_formatted,
                    DATE(detected_at) as violation_date
                FROM lane_violations
                WHERE camera_id = %s
                ORDER BY detected_at DESC
                LIMIT 10
            """
            print(f"🔍 Debug: Querying lane_violations for camera_id={camera_id}")
            results = violation_db.db.execute_query(query, (camera_id,), fetch=True)
            
            if results:
                print(f"✅ Found {len(results)} lane violations from database")
                for row in results:
                    # Format time_formatted properly (TIME() returns timedelta)
                    time_str = ''
                    if row.get('time_formatted'):
                        if isinstance(row['time_formatted'], datetime.timedelta):
                            total_seconds = int(row['time_formatted'].total_seconds())
                            hours = total_seconds // 3600
                            minutes = (total_seconds % 3600) // 60
                            seconds = total_seconds % 60
                            time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                        else:
                            time_str = str(row['time_formatted'])
                    
                    # Format violation date
                    date_str = ''
                    if row.get('violation_date'):
                        if isinstance(row['violation_date'], datetime.date):
                            date_str = row['violation_date'].strftime('%d/%m/%Y')
                        else:
                            date_str = str(row['violation_date'])
                    
                    # Format violation type text
                    violation_type_text = ''
                    if row.get('violation_type') == 'motor_in_car_lane':
                        violation_type_text = 'Xe máy vi phạm làn ô tô'
                    elif row.get('violation_type') == 'car_in_motor_lane':
                        violation_type_text = 'Ô tô vi phạm làn xe máy'
                    else:
                        violation_type_text = row.get('violation_type', 'Unknown')
                    
                    recent_violations.append({
                        'violation_id': row['violation_id'],
                        'frame_number': row['frame_number'],
                        'time_in_video': float(row['time_in_video']) if row['time_in_video'] else 0,
                        'violation_type': row['violation_type'],
                        'violation_type_text': violation_type_text,
                        'vehicle_type': row.get('vehicle_type', 'Unknown'),
                        'confidence': float(row['confidence']) if row['confidence'] else 0,
                        'detected_at': row['detected_at'].isoformat() if row['detected_at'] else '',
                        'time_formatted': time_str,
                        'date_formatted': date_str
                    })
            else:
                print(f"⚠️ No lane violations found in database")
        except Exception as e:
            print(f"⚠️ Error fetching recent lane violations: {e}")
            import traceback
            traceback.print_exc()
    
    return jsonify({
        'active': lane_detection_active,
        'motor_violations': motor_violations_db,
        'car_violations': car_violations_db, 
        'total_violations': total_violations_db,
        'runtime': runtime_seconds,
        'runtime_formatted': f"{int(runtime_seconds//3600):02d}:{int((runtime_seconds%3600)//60):02d}:{int(runtime_seconds%60):02d}",
        'debug': debug_info,
        'last_updated': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'recent_violations': recent_violations
    })


# Route để download file đã xuất - HỖ TRỢ CẢ PATH PARAM VÀ QUERY PARAM
@app.route("/download_lane_results", methods=['GET'])
@app.route("/download_lane_results/<file_type>", methods=['GET'])
def download_lane_results(file_type=None):
    """Download file kết quả lane detection (video, csv, excel, json)"""
    try:
        output_dir = "output"
        
        # Lấy file_type từ query param nếu không có trong path
        if file_type is None:
            file_type = request.args.get('file_type')
        
        # Lấy filename trực tiếp từ query param nếu có
        filename = request.args.get('filename')
        
        if filename:
            # Sử dụng filename được cung cấp trực tiếp
            file_path = os.path.join(output_dir, filename)
        else:
            # Fallback: tự generate filename từ timestamp
            timestamp = lane_detection_data.get('timestamp')
            
            if not timestamp:
                return jsonify({"error": "No results available"}), 404
            
            if file_type == "video":
                filename = f"lane_violations_{timestamp}.mp4"
            elif file_type == "csv":
                filename = f"lane_violations_stats_{timestamp}.csv"
            elif file_type == "excel":
                filename = f"lane_violations_stats_{timestamp}.xlsx"
            elif file_type == "json":
                filename = f"lane_violations_details_{timestamp}.json"
            else:
                return jsonify({"error": "Invalid file type"}), 400
            
            file_path = os.path.join(output_dir, filename)
        
        if os.path.exists(file_path):
            print(f"📥 Downloading file: {file_path}")
            return send_file(file_path, as_attachment=True, download_name=filename)
        else:
            print(f"❌ File not found: {file_path}")
            return jsonify({"error": f"File not found: {filename}"}), 404
            
    except Exception as e:
        print(f"❌ Error downloading file: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error downloading file: {str(e)}"}), 500


# ============= EXPORT EXCEL FROM DATABASE =============

@app.route("/export_lane_violations_excel", methods=['GET'])
def export_lane_violations_excel():
    """
    Xuất file Excel thống kê vi phạm làn đường từ DATABASE MySQL
    Thay vì từ video phân tích hiện tại
    
    Query parameters:
        - camera_id: ID camera (mặc định: 1 - lane detection)
        - date: Lọc theo ngày (format: YYYY-MM-DD, optional)
        - limit: Giới hạn số bản ghi (mặc định: tất cả)
    """
    try:
        if not violation_db:
            return jsonify({"error": "Database not connected"}), 500
        
        # Lấy parameters
        camera_id = request.args.get('camera_id', 1, type=int)
        date_filter = request.args.get('date')  # YYYY-MM-DD
        limit = request.args.get('limit', type=int)
        
        print(f"📊 Exporting lane violations from database (camera={camera_id}, date={date_filter}, limit={limit})...")
        
        # Truy vấn database
        query = """
            SELECT 
                v.violation_id,
                v.video_id,
                vid.video_filename,
                v.frame_number,
                v.time_in_video,
                v.violation_type,
                v.vehicle_type,
                v.confidence,
                v.bbox_x1, v.bbox_y1, v.bbox_x2, v.bbox_y2,
                v.image_path,
                v.detected_at
            FROM lane_violations v
            LEFT JOIN videos vid ON v.video_id = vid.video_id
            WHERE v.camera_id = %s
        """
        params = [camera_id]
        
        # Thêm filter theo ngày nếu có
        if date_filter:
            query += " AND DATE(v.detected_at) = %s"
            params.append(date_filter)
        
        # Sắp xếp theo thời gian mới nhất
        query += " ORDER BY v.detected_at DESC"
        
        # Thêm limit nếu có
        if limit:
            query += f" LIMIT {limit}"
        
        # Thực thi query
        results = violation_db.db.execute_query(query, tuple(params), fetch=True)
        
        if not results:
            return jsonify({"error": "No violations found in database"}), 404
        
        print(f"✅ Found {len(results)} violations in database")
        
        # Chuyển đổi sang DataFrame
        df = pd.DataFrame(results)
        
        # Định dạng lại các cột
        if 'detected_at' in df.columns:
            df['detected_at'] = pd.to_datetime(df['detected_at']).dt.strftime('%Y-%m-%d %H:%M:%S')
        
        if 'time_in_video' in df.columns:
            # Chuyển seconds thành MM:SS
            df['time_formatted'] = df['time_in_video'].apply(
                lambda x: f"{int(x // 60):02d}:{int(x % 60):02d}" if pd.notna(x) else ""
            )
        
        # Map violation_type sang tiếng Việt
        violation_type_map = {
            'motor_in_car_lane': 'Xe máy vi phạm làn ô tô',
            'car_in_motor_lane': 'Ô tô vi phạm làn xe máy'
        }
        df['violation_type_vn'] = df['violation_type'].map(violation_type_map)
        
        # Chọn và sắp xếp lại các cột
        output_columns = [
            'violation_id', 'video_filename', 'frame_number', 
            'time_formatted', 'violation_type_vn', 'vehicle_type',
            'confidence', 'detected_at'
        ]
        
        # Chỉ giữ các cột tồn tại
        output_columns = [col for col in output_columns if col in df.columns]
        df_export = df[output_columns]
        
        # Đổi tên cột sang tiếng Việt
        df_export.columns = [
            'ID Vi Phạm', 'Tên Video', 'Frame', 
            'Thời Gian (MM:SS)', 'Loại Vi Phạm', 'Loại Xe',
            'Độ Tin Cậy', 'Thời Gian Phát Hiện'
        ]
        
        # Tạo file Excel
        output_dir = "output"
        os.makedirs(output_dir, exist_ok=True)
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"lane_violations_database_{timestamp}.xlsx"
        excel_path = os.path.join(output_dir, excel_filename)
        
        # Xuất Excel với formatting
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name='Vi Phạm Làn Đường', index=False)
            
            # Lấy workbook và worksheet
            workbook = writer.book
            worksheet = writer.sheets['Vi Phạm Làn Đường']
            
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Thêm sheet thống kê
            summary_data = {
                'Chỉ Số': [
                    'Tổng Vi Phạm',
                    'Xe Máy Vi Phạm',
                    'Ô Tô Vi Phạm',
                    'Thời Gian Xuất'
                ],
                'Giá Trị': [
                    len(results),
                    len([r for r in results if r.get('violation_type') == 'motor_in_car_lane']),
                    len([r for r in results if r.get('violation_type') == 'car_in_motor_lane']),
                    datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Thống Kê', index=False)
        
        print(f"✅ Excel exported: {excel_path}")
        
        # Trả về file
        return send_file(excel_path, as_attachment=True, download_name=excel_filename)
        
    except Exception as e:
        print(f"❌ Error exporting Excel from database: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error exporting Excel: {str(e)}"}), 500


# ============= HELMET DETECTION ENDPOINTS =============

@app.route("/download_helmet_results", methods=['GET'])
@app.route("/download_helmet_results/<file_type>", methods=['GET'])
def download_helmet_results(file_type=None):
    """Download helmet detection result files (video, csv, excel, json) for the current session"""
    try:
        output_dir = "output"
        
        # Lấy file_type từ query param nếu không có trong path
        if file_type is None:
            file_type = request.args.get('file_type')
        
        # Lấy filename trực tiếp từ query param nếu có
        filename = request.args.get('filename')
        
        if filename:
            # Sử dụng filename được cung cấp trực tiếp
            file_path = os.path.join(output_dir, filename)
        else:
            # Fallback: tự generate filename từ timestamp
            timestamp = helmet_detection_data.get('timestamp')

            if not timestamp:
                return jsonify({"error": "No helmet results available"}), 404

            if file_type == "video":
                filename = f"helmet_analysis_{timestamp}.mp4"
            elif file_type == "csv":
                filename = f"helmet_violations_stats_{timestamp}.csv"
            elif file_type == "excel":
                filename = f"helmet_violations_stats_{timestamp}.xlsx"
            elif file_type == "json":
                filename = f"helmet_violations_details_{timestamp}.json"
            else:
                return jsonify({"error": "Invalid file type"}), 400

            file_path = os.path.join(output_dir, filename)

        if os.path.exists(file_path):
            print(f"📥 Downloading helmet file: {file_path}")
            return send_file(file_path, as_attachment=True, download_name=filename)
        else:
            print(f"❌ Helmet file not found: {file_path}")
            return jsonify({"error": f"File not found: {filename}"}), 404

    except Exception as e:
        print(f"❌ Error downloading helmet file: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error downloading helmet file: {str(e)}"}), 500

@app.route("/start_helmet_detection", methods=['POST'])
def start_helmet_detection():
    """Start helmet detection from web interface"""
    global helmet_detection_active, helmet_detection_data
    
    if helmet_detection_active:
        return jsonify({'error': 'Helmet detection is already running'}), 400
    
    # Check if video uploaded
    uploaded_video = session.get('uploaded_video_helmet')
    if not uploaded_video:
        return jsonify({'error': 'No video uploaded for helmet detection'}), 400
    
    print(f"🚀 Starting helmet detection for: {uploaded_video}")
    helmet_detection_active = True
    
    # Update video status to 'processing' in database
    video_id = session.get('uploaded_video_helmet_id')
    if video_db is not None and video_id is not None:
        try:
            video_db.update_video_status(video_id, 'processing')
            print(f"✅ [CAMERA 2 - START] Video status updated to processing: video_id={video_id}")
        except Exception as e:
            print(f"⚠️ [CAMERA 2 - START] Failed to update video status: {e}")
    
    # ✅ TẢI DỮ LIỆU TỪ DATABASE TRƯỚC KHI RESET
    print("🔧 [START] Loading existing data from database...")
    
    existing_without_helmet = 0
    existing_with_helmet = 0
    existing_violations = []
    
    if violation_db is not None:
        try:
            # Truy vấn TẤT CẢ vi phạm helmet từ database
            query = """
                SELECT 
                    violation_id,
                    video_id,
                    frame_number,
                    time_in_video,
                    has_helmet,
                    confidence,
                    license_plate,
                    bbox_x1, bbox_y1, bbox_x2, bbox_y2,
                    detected_at
                FROM helmet_violations
                ORDER BY detected_at DESC
            """
            db_results = violation_db.db.execute_query(query, fetch=True)
            
            if db_results:
                print(f"📊 [START] Loaded {len(db_results)} violations from database")
                
                for row in db_results:
                    has_helmet = row.get('has_helmet', True)
                    
                    # Đếm theo loại
                    if has_helmet:
                        existing_with_helmet += 1
                    else:
                        existing_without_helmet += 1
                    
                    # Lưu vào violations list
                    time_seconds = row.get('time_in_video', 0)
                    violation_info = {
                        'violation_id': row.get('violation_id'),
                        'video_id': row.get('video_id'),
                        'frame_number': row.get('frame_number'),
                        'time_seconds': time_seconds,
                        'time_formatted': f"{int(time_seconds // 60):02d}:{int(time_seconds % 60):02d}",
                        'has_helmet': has_helmet,
                        'confidence': row.get('confidence'),
                        'license_plate': row.get('license_plate'),
                        'bbox': [
                            row.get('bbox_x1'),
                            row.get('bbox_y1'),
                            row.get('bbox_x2'),
                            row.get('bbox_y2')
                        ],
                        'detected_at': row.get('detected_at').strftime('%Y-%m-%d %H:%M:%S') if row.get('detected_at') else ''
                    }
                    existing_violations.append(violation_info)
                
                print(f"✅ [START] Loaded: With helmet={existing_with_helmet}, Without helmet={existing_without_helmet}")
            else:
                print("ℹ️ [START] No existing data in database")
                
        except Exception as db_error:
            print(f"⚠️ [START] Failed to load from database: {db_error}")
            import traceback
            traceback.print_exc()
    else:
        print("ℹ️ [START] No database connection - starting fresh")
    
    # ✅ RESET CHỈ CÁC BIẾN TẠM THỜI, GIỮ NGUYÊN DỮ LIỆU TỪ DATABASE
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    helmet_detection_data = {
        'violations': existing_violations,  # ✅ GIỮ DỮ LIỆU CŨ
        'total_violations': existing_without_helmet,  # ✅ GIỮ SỐ LIỆU CŨ
        'with_helmet': existing_with_helmet,  # ✅ GIỮ SỐ LIỆU CŨ
        'without_helmet': existing_without_helmet,  # ✅ GIỮ SỐ LIỆU CŨ
        'start_time': time.time(),
        'video_writer': None,
        'output_path': None,
        'timestamp': timestamp,
        'frame_count': None,  # ✅ Set None để trigger load database trong generate_frames_helmet
        'original_fps': None,
        'output_size': None,
        'detection_cooldown': {},  # Reset cooldown cho phiên mới
        'frame_detections': set(),  # Reset frame tracking
        'temporal_votes': {},  # Reset vote memory
        'state_map': {},  # Reset state machine memory
        'recent_plates': [],  # Reset recent plates cache
        'current_with': 0,
        'current_without': 0,
        'current_total': 0
    }
    
    print(f"✅ [START COMPLETE] Starting counts: With={existing_with_helmet}, Without={existing_without_helmet}, Total violations={len(existing_violations)}")
    
    return jsonify({
        'success': True,
        'message': 'Helmet detection started successfully',
        'timestamp': timestamp
    })


@app.route("/stop_helmet_detection", methods=['POST'])
def stop_helmet_detection():
    """Stop helmet detection and export results"""
    global helmet_detection_active, helmet_detection_data
    
    if not helmet_detection_active:
        return jsonify({'error': 'Helmet detection is not running'}), 400
    
    print("🛑 Nhận yêu cầu dừng helmet detection từ web...")
    helmet_detection_active = False
    
    # Update video status to 'completed' in database
    video_id = session.get('uploaded_video_helmet_id')
    if video_db is not None and video_id is not None:
        try:
            video_db.update_video_status(video_id, 'completed')
            print(f"✅ [CAMERA 2 - STOP] Video status updated to completed: video_id={video_id}")
        except Exception as e:
            print(f"⚠️ [CAMERA 2 - STOP] Failed to update video status: {e}")
    
    # Đợi để detection dừng hoàn toàn
    time.sleep(2)
    
    try:
        # Đóng video writer nếu có
        if helmet_detection_data.get('video_writer'):
            helmet_detection_data['video_writer'].release()
            helmet_detection_data['video_writer'] = None
            print("✅ Helmet video writer released")
        
        violations_data = helmet_detection_data.get('violations', [])
        timestamp = helmet_detection_data.get('timestamp', datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
        output_path = helmet_detection_data.get('output_path', '')
        
        # Xuất CSV, Excel và JSON nếu có vi phạm
        csv_filename = None
        xlsx_filename = None
        json_filename = None
        
        output_dir = "output"
        os.makedirs(output_dir, exist_ok=True)
        
        # Dù có hay không vi phạm, vẫn tạo file thống kê để tiện tải
        df = pd.DataFrame(violations_data)
        csv_filename = os.path.join(output_dir, f"helmet_violations_stats_{timestamp}.csv")
        df.to_csv(csv_filename, index=False, encoding='utf-8-sig')
        print(f"✅ Đã xuất helmet CSV: {csv_filename}")

        # Excel
        xlsx_filename = os.path.join(output_dir, f"helmet_violations_stats_{timestamp}.xlsx")
        try:
            df.to_excel(xlsx_filename, index=False)
            print(f"✅ Đã xuất helmet Excel: {xlsx_filename}")
        except Exception as e:
            print(f"⚠️ Không thể xuất Excel: {e}")
            xlsx_filename = None
        
        # JSON chi tiết
        json_filename = os.path.join(output_dir, f"helmet_violations_details_{timestamp}.json")
        with open(json_filename, 'w', encoding='utf-8') as f:
            # Convert Decimal to float before JSON serialization
            violations_data_converted = convert_decimal_to_float(violations_data)
            json.dump({
                'video_info': {
                    'input_path': session.get('uploaded_video_helmet', ''),
                    'output_path': output_path,
                    'timestamp': timestamp
                },
                'summary': {
                    'total_violations': len(violations_data),
                    'frames_processed': helmet_detection_data.get('frame_count', 0),
                    'processing_time': time.time() - helmet_detection_data.get('start_time', time.time())
                },
                'violations': violations_data_converted
            }, f, indent=2, ensure_ascii=False)
        print(f"✅ Đã xuất helmet JSON: {json_filename}")
        
        response_data = {
            'success': True,
            'message': 'Helmet detection stopped and results exported successfully',
            'summary': {
                'total_violations': len(violations_data),
                'processing_time': time.time() - helmet_detection_data.get('start_time', time.time()),
                'frames_processed': helmet_detection_data.get('frame_count', 0)
            },
            'files': {
                'video': os.path.basename(output_path) if output_path else None,
                'csv': os.path.basename(csv_filename) if csv_filename else None,
                'excel': os.path.basename(xlsx_filename) if xlsx_filename else None,
                'json': os.path.basename(json_filename) if json_filename else None
            }
        }
        
        print(f"✅ Helmet detection đã dừng và xuất kết quả thành công")
        return jsonify(response_data)
        
    except Exception as e:
        print(f"❌ Lỗi khi xuất kết quả helmet: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error exporting helmet results: {str(e)}'}), 500


@app.route("/helmet_detection_status", methods=['GET'])
def get_helmet_detection_status():
    """
    Get current status of helmet detection with detailed stats
    Lấy dữ liệu TRỰC TIẾP TỪ DATABASE kết hợp với realtime session data
    """
    global helmet_detection_active, helmet_detection_data
    
    # Tính runtime từ session
    if helmet_detection_data.get('start_time') and isinstance(helmet_detection_data['start_time'], (int, float)):
        current_time = time.time()
        runtime_seconds = current_time - helmet_detection_data['start_time']
    else:
        runtime_seconds = 0
    
    # LẤY DỮ LIỆU TỪ DATABASE (tổng tích lũy - ALL TIME)
    db_with_helmet = 0
    db_without_helmet = 0
    db_total_detections = 0
    
    if stats_db:
        try:
            camera_id = 2  # Camera 2 = Helmet
            
            # ✅ Lấy thống kê TOÀN BỘ từ database (không truyền date)
            print(f"🔍 [DEBUG] Calling get_helmet_stats(camera_id={camera_id}) WITHOUT date filter")
            helmet_stats = stats_db.get_helmet_stats(camera_id)
            print(f"🔍 [DEBUG] Raw helmet_stats result: {helmet_stats}")
            
            if helmet_stats:
                db_without_helmet = helmet_stats.get('no_helmet_count', 0) or 0
                db_with_helmet = helmet_stats.get('with_helmet_count', 0) or 0
                db_total_detections = helmet_stats.get('total_detections', 0) or 0
            
            print(f"📊 [HELMET DB] ALL TIME: without={db_without_helmet}, with={db_with_helmet}, total={db_total_detections}")
        except Exception as e:
            print(f"⚠️ Error fetching helmet stats from DB: {e}")
            import traceback
            traceback.print_exc()
    
    # Nếu đang active, ưu tiên dữ liệu realtime từ session (đang xử lý)
    # Nếu không active, dùng dữ liệu từ database (đã xử lý xong)
    if helmet_detection_active:
        # Realtime: dùng dữ liệu từ session
        with_helmet = helmet_detection_data.get('with_helmet', 0)
        without_helmet = helmet_detection_data.get('without_helmet', 0)
        total_detections = with_helmet + without_helmet
        print(f"🔴 [ACTIVE] Using REALTIME data from session")
    else:
        # Không active: dùng dữ liệu từ database
        with_helmet = db_with_helmet
        without_helmet = db_without_helmet
        total_detections = db_total_detections
        print(f"🟢 [INACTIVE] Using DATABASE data: without={without_helmet}, with={with_helmet}, total={total_detections}")
    
    # Tính tỷ lệ vi phạm
    if total_detections > 0:
        violation_rate = round((without_helmet / total_detections) * 100, 1)
    else:
        violation_rate = 0
    
    # Current frame counts (chỉ khi đang active)
    cur_with = helmet_detection_data.get('current_with', 0) if helmet_detection_active else 0
    cur_without = helmet_detection_data.get('current_without', 0) if helmet_detection_active else 0
    cur_total = helmet_detection_data.get('current_total', 0) if helmet_detection_active else 0
    cur_rate = round((cur_without / cur_total) * 100, 1) if cur_total > 0 else 0

    # ✅ Lấy danh sách vi phạm gần đây (10 vi phạm mới nhất - chỉ vi phạm không đội mũ)
    recent_violations = []
    if violation_db:
        try:
            camera_id = 2  # Camera 2 = Helmet
            query = """
                SELECT 
                    violation_id,
                    frame_number,
                    time_in_video,
                    license_plate,
                    confidence,
                    detected_at,
                    TIME(detected_at) as time_formatted,
                    DATE(detected_at) as violation_date
                FROM helmet_violations
                WHERE camera_id = %s AND has_helmet = FALSE
                ORDER BY detected_at DESC
                LIMIT 10
            """
            print(f"🔍 Debug: Querying helmet_violations for camera_id={camera_id}")
            results = violation_db.db.execute_query(query, (camera_id,), fetch=True)
            
            if results:
                print(f"✅ Found {len(results)} helmet violations from database")
                for row in results:
                    # Format time_formatted properly (TIME() returns timedelta)
                    time_str = ''
                    if row.get('time_formatted'):
                        if isinstance(row['time_formatted'], datetime.timedelta):
                            total_seconds = int(row['time_formatted'].total_seconds())
                            hours = total_seconds // 3600
                            minutes = (total_seconds % 3600) // 60
                            seconds = total_seconds % 60
                            time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                        else:
                            time_str = str(row['time_formatted'])
                    
                    # Format violation date
                    date_str = ''
                    if row.get('violation_date'):
                        if isinstance(row['violation_date'], datetime.date):
                            date_str = row['violation_date'].strftime('%d/%m/%Y')
                        else:
                            date_str = str(row['violation_date'])
                    
                    recent_violations.append({
                        'violation_id': row['violation_id'],
                        'frame_number': row['frame_number'],
                        'time_in_video': float(row['time_in_video']) if row['time_in_video'] else 0,
                        'license_plate': row['license_plate'] or '',
                        'confidence': float(row['confidence']) if row['confidence'] else 0,
                        'detected_at': row['detected_at'].isoformat() if row['detected_at'] else '',
                        'time_formatted': time_str,
                        'date_formatted': date_str
                    })
            else:
                print(f"⚠️ No helmet violations found in database")
        except Exception as e:
            print(f"⚠️ Error fetching recent helmet violations: {e}")
            import traceback
            traceback.print_exc()

    response_data = {
        'active': helmet_detection_active,
        # Cumulative (từ database hoặc session nếu đang chạy)
        'total_violations': without_helmet,
        'with_helmet_cum': with_helmet,
        'without_helmet_cum': without_helmet,
        'total_detections_cum': total_detections,
        'violation_rate_cum': violation_rate,
        # Current (per frame - chỉ khi active)
        'with_helmet': cur_with,
        'without_helmet': cur_without,
        'total_detections': cur_total,
        'violation_rate': cur_rate,
        'frame_count': helmet_detection_data.get('frame_count', 0),
        'runtime': runtime_seconds,
        'runtime_formatted': f"{int(runtime_seconds//3600):02d}:{int((runtime_seconds%3600)//60):02d}:{int(runtime_seconds%60):02d}",
        'last_updated': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'source': 'database' if not helmet_detection_active else 'realtime',
        'recent_violations': recent_violations
    }
    
    print(f"📤 [RESPONSE] Sending helmet stats: total_violations={response_data['total_violations']}, source={response_data['source']}, recent_violations={len(recent_violations)}")
    return jsonify(response_data)


# Generate frames for red light detection (basic streaming)
def generate_placeholder_stream():
    """Generate a placeholder stream when no video is uploaded"""
    import numpy as np
    
    while True:
        # Create a black frame with text
        frame = np.zeros((480, 640, 3), dtype=np.uint8)
        
        # Add text overlay
        cv2.putText(frame, "CHUA CO VIDEO", (180, 200), 
                   cv2.FONT_HERSHEY_SIMPLEX, 1.2, (255, 255, 255), 2)
        cv2.putText(frame, "Vui long upload video", (150, 250), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.8, (200, 200, 200), 2)
        cv2.putText(frame, "de bat dau phat hien", (150, 280), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.8, (200, 200, 200), 2)
        cv2.putText(frame, "vi pham vuot den do", (150, 310), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.8, (200, 200, 200), 2)
        
        # Add loading animation (simple rotating circle)
        import time
        angle = int(time.time() * 100) % 360
        center = (320, 380)
        radius = 30
        end_point = (int(center[0] + radius * np.cos(np.radians(angle))), 
                    int(center[1] + radius * np.sin(np.radians(angle))))
        cv2.circle(frame, center, radius, (100, 100, 100), 2)
        cv2.line(frame, center, end_point, (0, 255, 0), 3)
        
        # Encode frame
        ret, buffer = cv2.imencode('.jpg', frame)
        frame_bytes = buffer.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
        
        # Small delay to prevent high CPU usage
        time.sleep(0.1)


def generate_frames_red_light(path_x):
    """Basic red light detection - fallback when advanced detection is not used"""
    cap = cv2.VideoCapture(path_x)
    
    while True:
        ret, frame = cap.read()
        if not ret:
            break
            
        # Simple frame processing - just add text overlay
        cv2.putText(frame, "Basic Red Light Detection", (50, 50), 
                   cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
        
        ret, buffer = cv2.imencode('.jpg', frame)
        frame_bytes = buffer.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
    
    cap.release()

# Generate frames for red light detection with processing (advanced streaming)
def generate_frames_red_light_advanced(path_x):
    """Real-time streaming với advanced processing và license plate detection"""
    from red_light_main import process_red_light_video_complete, generate_frames_red_light_new
    import tempfile
    import threading
    import queue
    import time
    
    # Create frame queue for streaming
    frame_queue = queue.Queue(maxsize=30)  # Buffer 30 frames
    processing_active = True
    
    def process_video_thread():
        """Background thread để process video và đưa frames vào queue"""
        nonlocal processing_active
        try:
            # Tạo temporary output file
            with tempfile.NamedTemporaryFile(suffix='.mp4', delete=False) as temp_file:
                temp_output = temp_file.name
            
            # Process video với custom frame callback
            cap = cv2.VideoCapture(path_x)
            if not cap.isOpened():
                return
                
            # Import các function cần thiết từ processRedLightVideo
            import sys
            import os
            sys.path.append(os.path.dirname(os.path.abspath(__file__)))
            
            # Import helper functions (manually implement here for web streaming)
            def draw_simulated_traffic_light_web(frame, current_light):
                """Draw simulated traffic light for web streaming"""
                height, width = frame.shape[:2]
                
                # Traffic light position (top-right corner)
                light_width = 60
                light_height = 150
                light_x = width - light_width - 20
                light_y = 20
                
                # Draw traffic light background
                cv2.rectangle(frame, (light_x, light_y), (light_x + light_width, light_y + light_height), (0, 0, 0), -1)
                cv2.rectangle(frame, (light_x, light_y), (light_x + light_width, light_y + light_height), (255, 255, 255), 2)
                
                # Light positions
                circle_radius = 18
                circle_x = light_x + light_width // 2
                red_y = light_y + 30
                yellow_y = light_y + 75
                green_y = light_y + 120
                
                # Draw inactive lights
                cv2.circle(frame, (circle_x, red_y), circle_radius, (50, 50, 50), -1)
                cv2.circle(frame, (circle_x, yellow_y), circle_radius, (50, 50, 50), -1)
                cv2.circle(frame, (circle_x, green_y), circle_radius, (50, 50, 50), -1)
                
                # Draw active light
                if current_light == "red":
                    cv2.circle(frame, (circle_x, red_y), circle_radius, (0, 0, 255), -1)
                elif current_light == "yellow":
                    cv2.circle(frame, (circle_x, yellow_y), circle_radius, (0, 255, 255), -1)
                elif current_light == "green":
                    cv2.circle(frame, (circle_x, green_y), circle_radius, (0, 255, 0), -1)
                
                # Add light status text
                cv2.putText(frame, f"Light: {current_light.upper()}", 
                           (light_x - 50, light_y - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 2)
            
            def extract_license_plate_web(frame, bbox, reader):
                """Simple license plate extraction for web streaming"""
                if reader is None:
                    return None, None
                    
                try:
                    x1, y1, x2, y2 = bbox
                    vehicle_width = x2 - x1
                    vehicle_height = y2 - y1
                    
                    if vehicle_width < 40 or vehicle_height < 30:
                        return None, None
                    
                    # Focus on bottom part of vehicle  
                    crop_h = y2 - y1
                    crop_y1 = y1 + int(crop_h * 0.6)
                    license_region = frame[crop_y1:y2, x1:x2]
                    
                    if license_region.size == 0 or license_region.shape[0] < 10:
                        return None, None
                    
                    # Simple resize and OCR
                    scale_factor = 2.0 if vehicle_width < 80 else 1.5
                    enhanced = cv2.resize(license_region, None, fx=scale_factor, fy=scale_factor, interpolation=cv2.INTER_CUBIC)
                    
                    if len(enhanced.shape) == 3:
                        enhanced = cv2.cvtColor(enhanced, cv2.COLOR_BGR2GRAY)
                    
                    # Simple OCR call
                    results = reader.readtext(enhanced, paragraph=False)
                    
                    best_text = None
                    best_confidence = 0
                    best_bbox = None
                    
                    for (bbox_coords, text, prob) in results:
                        if prob > 0.2:
                            cleaned_text = ''.join(c for c in text if c.isalnum())
                            
                            if len(cleaned_text) >= 4 and len(cleaned_text) <= 12:
                                has_letters = any(c.isalpha() for c in cleaned_text)
                                has_numbers = any(c.isdigit() for c in cleaned_text)
                                
                                if (has_letters and has_numbers) or (has_numbers and len(cleaned_text) >= 4):
                                    if prob > best_confidence:
                                        best_confidence = prob
                                        best_text = cleaned_text.upper()
                                        
                                        # Convert bbox back to original frame coordinates
                                        # bbox_coords is [(x1, y1), (x2, y1), (x2, y2), (x1, y2)]
                                        ocr_x1 = int(min([p[0] for p in bbox_coords]) / scale_factor) + x1
                                        ocr_y1 = int(min([p[1] for p in bbox_coords]) / scale_factor) + crop_y1
                                        ocr_x2 = int(max([p[0] for p in bbox_coords]) / scale_factor) + x1
                                        ocr_y2 = int(max([p[1] for p in bbox_coords]) / scale_factor) + crop_y1
                                        
                                        best_bbox = (ocr_x1, ocr_y1, ocr_x2, ocr_y2)
                    
                    return best_text, best_bbox
                    
                except Exception as e:
                    return None, None
            from ultralytics import YOLO
            import easyocr
            import numpy as np
            from collections import defaultdict
            
            # Initialize models như trong processRedLightVideo
            model = YOLO('YoloWeights/yolov8n.pt')
            vehicle_model = YOLO('model_lane/vehicle.pt')
            
            try:
                reader = easyocr.Reader(['vi', 'en'])
            except:
                reader = None
            
            # Initialize variables
            frame_count = 0
            current_light = "red"
            violation_count = 0
            processed_vehicles = set()
            frames_since_light_change = 0
            light_cycle_duration = 90
            light_patterns = ["red", "red", "yellow", "green", "red", "red"]
            current_light_index = 0
            
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            
            # Detection lines
            detection_lines = [{
                "start": {"x": int(width * 0.0), "y": int(height * 0.85)},
                "end": {"x": int(width * 0.6), "y": int(height * 0.90)}
            }]
            
            while processing_active:
                ret, frame = cap.read()
                if not ret:
                    break
                    
                frame_count += 1
                
                # Traffic light simulation
                frames_since_light_change += 1
                if frames_since_light_change >= light_cycle_duration:
                    frames_since_light_change = 0
                    current_light_index = (current_light_index + 1) % len(light_patterns)
                    current_light = light_patterns[current_light_index]
                
                # Detect vehicles and violations (simplified for streaming)
                vehicle_results = vehicle_model(frame)
                
                # Draw simulated traffic light
                draw_simulated_traffic_light_web(frame, current_light)
                
                # Process vehicles và license plates
                vehicles_with_plates = []
                for r in vehicle_results:
                    boxes = r.boxes
                    if boxes is not None:
                        for box in boxes:
                            cls = int(box.cls[0])
                            conf = float(box.conf[0])
                            
                            if cls in [0, 1, 2, 3, 4] and conf > 0.3:
                                x1, y1, x2, y2 = map(int, box.xyxy[0])
                                
                                # Extract license plate (simplified for speed)
                                license_text = None
                                license_bbox = None
                                if reader and frame_count % 10 == 0:  # Every 10 frames
                                    try:
                                        license_text, license_bbox = extract_license_plate_web(frame, [x1, y1, x2, y2], reader)
                                    except:
                                        pass
                                
                                # Check violation
                                vehicle_id = f"{x1}_{y1}_{x2}_{y2}"
                                center_y = (y1 + y2) // 2
                                
                                is_violation = (current_light == "red" and 
                                              center_y > height * 0.85 and 
                                              vehicle_id not in processed_vehicles)
                                
                                if is_violation:
                                    violation_count += 1
                                    processed_vehicles.add(vehicle_id)
                                    
                                    # Draw violation
                                    cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 3)
                                    cv2.putText(frame, "VI PHAM!", (x1, y1 - 10), 
                                              cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
                                    
                                    # Tạo PDF biên bản phạt cho vi phạm đèn đỏ
                                    try:
                                        import createBB_red_light
                                        from PIL import Image
                                        import tempfile
                                        
                                        # Lưu ảnh vi phạm
                                        os.makedirs("data_vuot_den_do", exist_ok=True)
                                        cv2.imwrite(f"data_vuot_den_do/{violation_count}.jpg", frame)
                                        
                                        # Tạo PDF biên bản phạt
                                        stt_BB_red_light = f'BienBanNopPhatVuotDenDo/{violation_count}.pdf'
                                        frame_pil = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
                                        temp_image = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
                                        frame_pil.save(temp_image.name)
                                        
                                        # Cập nhật thông tin biên bản với biển số xe
                                        examBB = createBB_red_light.infoObject()
                                        if license_text:
                                            examBB['license_plate'] = license_text
                                        createBB_red_light.bienBanNopPhat(examBB, temp_image.name,
                                                                         f"data_vuot_den_do/{violation_count}.jpg", stt_BB_red_light)
                                        temp_image.close()
                                        print(f"Created PDF violation report: {stt_BB_red_light}")
                                    except Exception as e:
                                        print(f"Error creating PDF for violation {violation_count}: {e}")
                                else:
                                    # Normal vehicle
                                    color = (0, 255, 0) if current_light != "red" else (255, 255, 0)
                                    cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                                
                                # Draw license plate if detected
                                if license_text and license_bbox:
                                    # Draw license plate bounding box
                                    lp_x1, lp_y1, lp_x2, lp_y2 = license_bbox
                                    cv2.rectangle(frame, (lp_x1, lp_y1), (lp_x2, lp_y2), (0, 255, 0), 2)
                                    
                                    # Draw license plate text above the license plate box
                                    label = f"LP: {license_text}"
                                    label_size = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)[0]
                                    
                                    # Position above the license plate box
                                    text_x = max(0, lp_x1)
                                    text_y = max(label_size[1] + 5, lp_y1 - 5)
                                    
                                    # Ensure text stays within frame
                                    if text_x + label_size[0] > width:
                                        text_x = width - label_size[0] - 5
                                    if text_y < label_size[1] + 5:
                                        text_y = lp_y2 + label_size[1] + 10
                                    
                                    # Draw background rectangle for better visibility
                                    cv2.rectangle(frame, 
                                                (text_x, text_y - label_size[1] - 2), 
                                                (text_x + label_size[0] + 4, text_y + 2), 
                                                (0, 0, 0), -1)
                                    
                                    # Draw text
                                    cv2.putText(frame, label, (text_x + 2, text_y), 
                                              cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                                elif license_text:
                                    # Fallback: Draw text above vehicle if no license bbox
                                    label = f"LP: {license_text}"
                                    label_size = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)[0]
                                    
                                    # Position above the vehicle box
                                    text_x = max(0, x1)
                                    text_y = max(label_size[1] + 5, y1 - 5)
                                    
                                    # Draw background rectangle for better visibility
                                    cv2.rectangle(frame, 
                                                (text_x, text_y - label_size[1] - 2), 
                                                (text_x + label_size[0] + 4, text_y + 2), 
                                                (0, 0, 0), -1)
                                    
                                    # Draw text
                                    cv2.putText(frame, label, (text_x + 2, text_y), 
                                              cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                
                # Draw detection lines
                for line in detection_lines:
                    start_x = int(line["start"]["x"])
                    start_y = int(line["start"]["y"])
                    end_x = int(line["end"]["x"])
                    end_y = int(line["end"]["y"])
                    
                    line_color = (0, 0, 255) if current_light == "red" else (255, 255, 255)
                    cv2.line(frame, (start_x, start_y), (end_x, end_y), line_color, 3)
                
                # Draw info panel
                cv2.rectangle(frame, (10, 10), (400, 120), (0, 0, 0), -1)
                cv2.rectangle(frame, (10, 10), (400, 120), (255, 255, 255), 2)
                
                status_color = (0, 0, 255) if current_light == "red" else (0, 255, 0) if current_light == "green" else (0, 255, 255)
                cv2.putText(frame, f"Den giao thong: {current_light.upper()}", (20, 35),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.6, status_color, 2)
                # cv2.putText(frame, f"Vi pham: {violation_count}", (20, 60),
                #            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
                # cv2.putText(frame, f"Frame: {frame_count}", (20, 85),
                #            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
                cv2.putText(frame, "LIVE STREAM", (20, 105),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)
                
                # Add frame to queue (non-blocking)
                try:
                    frame_queue.put(frame, block=False)
                except queue.Full:
                    # Skip frame if queue is full
                    pass
                    
            cap.release()
            
        except Exception as e:
            print(f"Error in processing thread: {e}")
        finally:
            processing_active = False
    
    # Start background processing thread
    processing_thread = threading.Thread(target=process_video_thread, daemon=True)
    processing_thread.start()
    
    # Stream frames from queue
    try:
        while processing_active or not frame_queue.empty():
            try:
                # Get frame from queue
                frame = frame_queue.get(timeout=1.0)
                
                # Encode frame
                ret, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
                if ret:
                    frame_bytes = buffer.tobytes()
                    yield (b'--frame\r\n'
                           b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
                
            except queue.Empty:
                # No frame available, continue
                continue
                
    except GeneratorExit:
        # Client disconnected
        processing_active = False
    finally:
        processing_active = False


# Clear uploaded video
@app.route("/clear_upload", methods=['POST'])
def clear_upload():
    try:
        if 'uploaded_video' in session:
            filepath = session['uploaded_video']
            
            # Thử xóa file với retry logic
            if os.path.exists(filepath):
                max_retries = 3
                retry_delay = 0.5
                
                for attempt in range(max_retries):
                    try:
                        os.remove(filepath)
                        print(f"✅ Deleted video file: {filepath}")
                        break
                    except PermissionError as e:
                        if attempt < max_retries - 1:
                            print(f"⚠️ File in use, retrying in {retry_delay}s... (attempt {attempt + 1}/{max_retries})")
                            time.sleep(retry_delay)
                            retry_delay *= 2
                        else:
                            print(f"⚠️ Could not delete file (still in use): {filepath}")
                    except Exception as e:
                        print(f"⚠️ Error deleting file: {e}")
                        break
            
            session.pop('uploaded_video', None)
            session.pop('uploaded_video_id', None)
        
        return jsonify({'success': True, 'message': 'Upload cleared successfully'})
    except Exception as e:
        print(f"❌ Error in clear_upload: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# Clear uploaded video for helmet detection
@app.route("/clear_upload_helmet", methods=['POST'])
def clear_upload_helmet():
    global helmet_detection_active, helmet_detection_data
    
    try:
        # Dừng helmet detection nếu đang chạy
        if helmet_detection_active:
            helmet_detection_active = False
            print("🛑 Stopping helmet detection before clearing upload...")
            
            # Đóng video writer nếu đang mở
            if helmet_detection_data.get('video_writer'):
                try:
                    helmet_detection_data['video_writer'].release()
                    helmet_detection_data['video_writer'] = None
                    print("✅ Video writer released")
                except Exception as e:
                    print(f"⚠️ Error releasing video writer: {e}")
            
            # Đợi một chút để đảm bảo VideoCapture đã đóng
            time.sleep(0.5)
        
        if 'uploaded_video_helmet' in session:
            filepath = session['uploaded_video_helmet']
            
            # Thử xóa file với retry logic
            if os.path.exists(filepath):
                max_retries = 3
                retry_delay = 0.5
                
                for attempt in range(max_retries):
                    try:
                        os.remove(filepath)
                        print(f"✅ Deleted helmet video file: {filepath}")
                        break
                    except PermissionError as e:
                        if attempt < max_retries - 1:
                            print(f"⚠️ File in use, retrying in {retry_delay}s... (attempt {attempt + 1}/{max_retries})")
                            time.sleep(retry_delay)
                            retry_delay *= 2  # Exponential backoff
                        else:
                            print(f"⚠️ Could not delete file (still in use): {filepath}")
                            print(f"   Error: {e}")
                            # Đánh dấu để xóa sau (deferred deletion)
                            # File sẽ được xóa khi process kết thúc hoặc khi VideoCapture đóng
                    except Exception as e:
                        print(f"⚠️ Error deleting file: {e}")
                        break
            
            session.pop('uploaded_video_helmet', None)
            session.pop('uploaded_video_helmet_id', None)
        
        return jsonify({'success': True, 'message': 'Upload cleared successfully'})
    except Exception as e:
        print(f"❌ Error in clear_upload_helmet: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# Clear uploaded video for lane detection
@app.route("/clear_upload_lane", methods=['POST'])
def clear_upload_lane():
    global lane_detection_active, lane_detection_data
    
    try:
        # Dừng lane detection nếu đang chạy
        if lane_detection_active:
            lane_detection_active = False
            print("🛑 Stopping lane detection before clearing upload...")
            
            # Đóng video writer nếu đang mở
            if lane_detection_data.get('video_writer'):
                try:
                    lane_detection_data['video_writer'].release()
                    lane_detection_data['video_writer'] = None
                    print("✅ Video writer released")
                except Exception as e:
                    print(f"⚠️ Error releasing video writer: {e}")
            
            # Đợi một chút để đảm bảo VideoCapture đã đóng
            time.sleep(0.5)
        
        if 'uploaded_video_lane' in session:
            filepath = session['uploaded_video_lane']
            
            # Thử xóa file với retry logic
            if os.path.exists(filepath):
                max_retries = 3
                retry_delay = 0.5
                
                for attempt in range(max_retries):
                    try:
                        os.remove(filepath)
                        print(f"✅ Deleted lane video file: {filepath}")
                        break
                    except PermissionError as e:
                        if attempt < max_retries - 1:
                            print(f"⚠️ File in use, retrying in {retry_delay}s... (attempt {attempt + 1}/{max_retries})")
                            time.sleep(retry_delay)
                            retry_delay *= 2
                        else:
                            print(f"⚠️ Could not delete file (still in use): {filepath}")
                    except Exception as e:
                        print(f"⚠️ Error deleting file: {e}")
                        break
            
            session.pop('uploaded_video_lane', None)
            session.pop('current_video_id_lane', None)
        
        return jsonify({'success': True, 'message': 'Upload cleared successfully'})
    except Exception as e:
        print(f"❌ Error in clear_upload_lane: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# API endpoint for helmet violation statistics
@app.route("/api/helmet_violations", methods=['GET'])
def get_helmet_violations():
    # In production, this would fetch from database or real-time detection
    # For now, return sample data
    violations = session.get('helmet_violations', [])
    return jsonify({
        'total_count': len(violations),
        'violations': violations[-10:]  # Last 10 violations
    })


# Check processing status
@app.route("/api/processing_status", methods=['GET'])
def get_processing_status():
    processing_info = session.get('helmet_processing', {})
    if not processing_info:
        return jsonify({'status': 'no_job'})
    
    response = {
        'status': processing_info.get('status', 'unknown'),
        'job_id': processing_info.get('job_id')
    }
    
    if processing_info.get('status') == 'completed':
        response['stats'] = processing_info.get('stats', {})
        response['output_path'] = processing_info.get('output_path')
    elif processing_info.get('status') == 'error':
        response['error'] = processing_info.get('error', 'Unknown error')
    
    return jsonify(response)


# New API endpoint for real-time progress tracking
@app.route("/api/helmet_progress", methods=['GET'])
def get_helmet_progress():
    """Get real-time progress of helmet detection processing"""
    # For original testHelmet.py, just return basic progress
    import glob
    violation_files = glob.glob('data_xe_vp_bh/*.jpg')
    progress = {
        'violations': len(violation_files),
        'total_frames': 0,
        'current_frame': 0,
        'status': 'completed',
        'estimated_time_left': 0,
        'formatted_time_left': "Hoàn thành"
    }
    
    return jsonify(progress)


# Reset progress tracking
@app.route("/api/helmet_progress/reset", methods=['POST'])
def reset_helmet_progress():
    """Reset progress tracking"""
    reset_processing_progress()
    return jsonify({'success': True, 'message': 'Progress reset successfully'})


def reset_processing_progress():
    """Reset processing progress for helmet detection"""
    # Clear session data
    if 'helmet_processing' in session:
        session.pop('helmet_processing', None)
    if 'pending_helmet_process' in session:
        session.pop('pending_helmet_process', None)
    if 'uploaded_video_helmet' in session:
        session.pop('uploaded_video_helmet', None)
    
    # You can add additional cleanup here if needed
    print("Processing progress reset")


# API endpoints for red light violations
@app.route("/api/red_light_violations", methods=['GET'])
def get_red_light_violations():
    """Get red light violation statistics"""
    import glob
    
    # Count violation images
    violation_files = glob.glob('data_vuot_den_do/*.jpg')
    fine_documents = glob.glob('BienBanNopPhatVuotDenDo/*.pdf')
    
    # Get recent violations
    recent_violations = []
    for file_path in sorted(violation_files, reverse=True)[:10]:
        filename = os.path.basename(file_path)
        # Extract timestamp from filename if possible
        parts = filename.split('_')
        if len(parts) >= 3:
            recent_violations.append({
                'filename': filename,
                'timestamp': parts[1] + '_' + parts[2].split('.')[0] if len(parts) > 2 else 'unknown',
                'file_path': file_path
            })
    
    return jsonify({
        'total_violations': len(violation_files),
        'total_fines': len(fine_documents),
        'recent_violations': recent_violations,
        'detection_method': 'Advanced' if session.get('red_light_advanced', False) else 'Basic'
    })


@app.route("/api/set_detection_method", methods=['POST'])
def set_detection_method():
    """Set red light detection method preference"""
    data = request.get_json()
    method = data.get('method', 'original')
    
    session['red_light_advanced'] = (method == 'advanced')
    
    return jsonify({
        'success': True,
        'method': method,
        'message': f'Detection method set to {method}'
    })


@app.route("/api/session_status", methods=['GET'])
def get_session_status():
    """Debug endpoint to check session status"""
    return jsonify({
        'session_keys': list(session.keys()),
        'uploaded_video': session.get('uploaded_video', 'NOT_FOUND'),
        'uploaded_video_exists': os.path.exists(session.get('uploaded_video', '')) if session.get('uploaded_video') else False,
        'red_light_advanced': session.get('red_light_advanced', False),
        'session_id': session.get('_session_id', 'NO_ID')
    })


# =============================================================================
# NEW DATABASE API ENDPOINTS
# =============================================================================

@app.route('/api/upload_video_db', methods=['POST'])
def upload_video_db():
    """
    Upload video and save to database
    Form data: 'video' file, 'camera_id' (1, 2, or 3)
    """
    if 'video' not in request.files:
        return jsonify({'error': 'No video file'}), 400
    
    if 'camera_id' not in request.form:
        return jsonify({'error': 'No camera_id specified'}), 400
    
    file = request.files['video']
    camera_id = int(request.form['camera_id'])
    
    if file.filename == '':
        return jsonify({'error': 'Empty filename'}), 400
    
    if camera_id not in [1, 2, 3]:
        return jsonify({'error': 'Invalid camera_id. Must be 1, 2, or 3'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type'}), 400
    
    try:
        # Save file
        from werkzeug.utils import secure_filename
        filename = secure_filename(file.filename)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"cam{camera_id}_{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        print(f"✅ File saved: {filepath}")
        
        # Get video properties
        cap = cv2.VideoCapture(filepath)
        if cap.isOpened():
            fps = int(cap.get(cv2.CAP_PROP_FPS))
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            duration = frame_count / fps if fps > 0 else 0
            file_size = os.path.getsize(filepath) / (1024 * 1024)
            cap.release()
            
            props = {
                'fps': fps,
                'resolution': f"{width}x{height}",
                'duration_seconds': duration,
                'file_size_mb': round(file_size, 2)
            }
        else:
            props = {}
        
        # Insert to database if available
        video_id = None
        if video_db:
            try:
                video_id = video_db.insert_video(
                    camera_id=camera_id,
                    video_filename=filename,
                    video_path=filepath,
                    file_size_mb=props.get('file_size_mb'),
                    duration_seconds=int(props.get('duration_seconds', 0)),
                    fps=props.get('fps'),
                    resolution=props.get('resolution')
                )
                print(f"✅ Video saved to database with ID: {video_id}")
            except Exception as e:
                print(f"⚠️ Database insert failed: {e}")
        
        # Store in session for processing
        if camera_id == 1:
            session['uploaded_video_lane'] = filepath
        elif camera_id == 2:
            session['uploaded_video_helmet'] = filepath
        else:
            session['uploaded_video'] = filepath
        
        return jsonify({
            'success': True,
            'video_id': video_id,
            'camera_id': camera_id,
            'camera_name': f"Camera {camera_id}",
            'filename': filename,
            'properties': props,
            'message': f'Video uploaded successfully for Camera {camera_id}'
        })
        
    except Exception as e:
        print(f"❌ Upload error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/stats/overall')
def get_overall_stats_api():
    """Get overall statistics from database with detailed breakdown"""
    try:
        if stats_db:
            overall_stats = stats_db.get_overall_stats()
            today = datetime.date.today()
            
            detailed_stats = []
            total_all_time = 0
            total_today = 0
            
            for camera in overall_stats:
                camera_id = camera['camera_id']
                camera_type = camera['camera_type']
                
                # Get detailed stats per camera
                if camera_type == 'lane':
                    # All time
                    lane_stats = stats_db.get_lane_stats(camera_id)
                    motor_all = sum(s['count'] for s in lane_stats if s['violation_type'] == 'motor_in_car_lane')
                    car_all = sum(s['count'] for s in lane_stats if s['violation_type'] == 'car_in_motor_lane')
                    
                    # Today
                    lane_today = stats_db.get_lane_stats(camera_id, date=today)
                    motor_today = sum(s['count'] for s in lane_today if s['violation_type'] == 'motor_in_car_lane')
                    car_today = sum(s['count'] for s in lane_today if s['violation_type'] == 'car_in_motor_lane')
                    
                    detailed_stats.append({
                        'camera_id': camera_id,
                        'camera_name': camera['camera_name'],
                        'camera_type': camera_type,
                        'total_all_time': motor_all + car_all,
                        'total_today': motor_today + car_today,
                        'breakdown': {
                            'motor_violations': motor_all,
                            'car_violations': car_all,
                            'motor_today': motor_today,
                            'car_today': car_today
                        }
                    })
                    total_all_time += motor_all + car_all
                    total_today += motor_today + car_today
                    
                elif camera_type == 'helmet':
                    # All time
                    helmet_stats = stats_db.get_helmet_stats(camera_id)
                    no_helmet_all = helmet_stats['no_helmet_count'] if helmet_stats and helmet_stats.get('no_helmet_count') is not None else 0
                    
                    # Today
                    helmet_today = stats_db.get_helmet_stats(camera_id, date=today)
                    no_helmet_today = helmet_today['no_helmet_count'] if helmet_today and helmet_today.get('no_helmet_count') is not None else 0
                    
                    detailed_stats.append({
                        'camera_id': camera_id,
                        'camera_name': camera['camera_name'],
                        'camera_type': camera_type,
                        'total_all_time': no_helmet_all or 0,
                        'total_today': no_helmet_today or 0,
                        'breakdown': {
                            'no_helmet': no_helmet_all or 0,
                            'no_helmet_today': no_helmet_today or 0
                        }
                    })
                    total_all_time += (no_helmet_all or 0)
                    total_today += (no_helmet_today or 0)
                    
                elif camera_type == 'red_light':
                    # All time
                    redlight_stats = stats_db.get_red_light_stats(camera_id)
                    violations_all = redlight_stats['violation_count'] if redlight_stats and redlight_stats.get('violation_count') is not None else 0
                    
                    # Today
                    redlight_today = stats_db.get_red_light_stats(camera_id, date=today)
                    violations_today = redlight_today['violation_count'] if redlight_today and redlight_today.get('violation_count') is not None else 0
                    
                    detailed_stats.append({
                        'camera_id': camera_id,
                        'camera_name': camera['camera_name'],
                        'camera_type': camera_type,
                        'total_all_time': violations_all or 0,
                        'total_today': violations_today or 0,
                        'breakdown': {
                            'red_light_violations': violations_all or 0,
                            'red_light_today': violations_today or 0
                        }
                    })
                    total_all_time += (violations_all or 0)
                    total_today += (violations_today or 0)
            
            return jsonify({
                'success': True,
                'summary': {
                    'total_violations_all_time': total_all_time,
                    'total_violations_today': total_today
                },
                'cameras': detailed_stats,
                'source': 'database'
            })
    except Exception as e:
        print(f"⚠️ Database error: {e}")
        import traceback
        traceback.print_exc()
    
    # Fallback
    return jsonify({
        'success': True,
        'summary': {
            'total_violations_all_time': 0,
            'total_violations_today': 0
        },
        'cameras': [
            {'camera_id': 1, 'camera_name': 'Camera 1', 'camera_type': 'lane', 
             'total_all_time': 0, 'total_today': 0, 'breakdown': {}},
            {'camera_id': 2, 'camera_name': 'Camera 2', 'camera_type': 'helmet', 
             'total_all_time': 0, 'total_today': 0, 'breakdown': {}},
            {'camera_id': 3, 'camera_name': 'Camera 3', 'camera_type': 'red_light', 
             'total_all_time': 0, 'total_today': 0, 'breakdown': {}}
        ],
        'source': 'fallback'
    })


@app.route('/api/stats/overall/<period>')
def get_overall_stats_by_period_api(period):
    """Get overall statistics filtered by time period"""
    try:
        if not stats_db:
            return jsonify({
                'success': False,
                'error': 'Database not available'
            }), 503
        
        today = datetime.date.today()
        start_date = None
        end_date = today
        
        # Calculate date range based on period
        if period == 'today':
            start_date = today
            end_date = today
        elif period == 'week':
            start_date = today - datetime.timedelta(days=7)
        elif period == 'month':
            start_date = today - datetime.timedelta(days=30)
        elif period == 'year':
            start_date = datetime.date(today.year, 1, 1)
        elif period == 'all':
            start_date = None
            end_date = None
        else:
            return jsonify({
                'success': False,
                'error': f'Invalid period: {period}'
            }), 400
        
        overall_stats = stats_db.get_overall_stats()
        detailed_stats = []
        total_violations = 0
        
        for camera in overall_stats:
            camera_id = camera['camera_id']
            camera_type = camera['camera_type']
            
            if camera_type == 'lane':
                if period == 'today' or (start_date and end_date and start_date == end_date):
                    # Single date - use existing method
                    lane_stats = stats_db.get_lane_stats(camera_id, date=start_date)
                else:
                    # Date range - use new method
                    lane_stats = stats_db.get_lane_stats_by_period(camera_id, start_date, end_date)
                
                motor_count = sum(s['count'] for s in lane_stats if s['violation_type'] == 'motor_in_car_lane')
                car_count = sum(s['count'] for s in lane_stats if s['violation_type'] == 'car_in_motor_lane')
                total = motor_count + car_count
                
                detailed_stats.append({
                    'camera_id': camera_id,
                    'camera_name': camera['camera_name'],
                    'camera_type': camera_type,
                    'total': total,
                    'breakdown': {
                        'motor_violations': motor_count,
                        'car_violations': car_count
                    }
                })
                total_violations += total
                
            elif camera_type == 'helmet':
                if period == 'today' or (start_date and end_date and start_date == end_date):
                    # Single date - use existing method
                    helmet_stats = stats_db.get_helmet_stats(camera_id, date=start_date)
                else:
                    # Date range - use new method
                    helmet_stats = stats_db.get_helmet_stats_by_period(camera_id, start_date, end_date)
                
                no_helmet = helmet_stats['no_helmet_count'] if helmet_stats and helmet_stats.get('no_helmet_count') is not None else 0
                
                detailed_stats.append({
                    'camera_id': camera_id,
                    'camera_name': camera['camera_name'],
                    'camera_type': camera_type,
                    'total': no_helmet or 0,
                    'breakdown': {
                        'no_helmet': no_helmet or 0
                    }
                })
                total_violations += (no_helmet or 0)
                
            elif camera_type == 'red_light':
                if period == 'today' or (start_date and end_date and start_date == end_date):
                    # Single date - use existing method
                    redlight_stats = stats_db.get_red_light_stats(camera_id, date=start_date)
                else:
                    # Date range - use new method
                    redlight_stats = stats_db.get_red_light_stats_by_period(camera_id, start_date, end_date)
                
                violations = redlight_stats['violation_count'] if redlight_stats and redlight_stats.get('violation_count') is not None else 0
                
                detailed_stats.append({
                    'camera_id': camera_id,
                    'camera_name': camera['camera_name'],
                    'camera_type': camera_type,
                    'total': violations or 0,
                    'breakdown': {
                        'red_light_violations': violations or 0
                    }
                })
                total_violations += (violations or 0)
        
        # Calculate breakdown by type
        lane_total = sum(c['total'] for c in detailed_stats if c['camera_type'] == 'lane')
        helmet_total = sum(c['total'] for c in detailed_stats if c['camera_type'] == 'helmet')
        redlight_total = sum(c['total'] for c in detailed_stats if c['camera_type'] == 'red_light')
        
        return jsonify({
            'success': True,
            'period': period,
            'summary': {
                'total_violations': total_violations,
                'lane_violations': lane_total,
                'helmet_violations': helmet_total,
                'redlight_violations': redlight_total
            },
            'cameras': detailed_stats,
            'source': 'database'
        })
        
    except Exception as e:
        print(f"❌ Error getting stats by period: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/stats/<int:camera_id>')
def get_camera_stats_api(camera_id):
    """Get statistics for specific camera"""
    try:
        if stats_db:
            if camera_id == 1:
                stats = stats_db.get_lane_stats(camera_id)
            elif camera_id == 2:
                stats = stats_db.get_helmet_stats(camera_id)
            elif camera_id == 3:
                stats = stats_db.get_red_light_stats(camera_id)
            else:
                return jsonify({'error': 'Invalid camera_id'}), 400
            
            return jsonify({
                'camera_id': camera_id,
                'camera_name': f"Camera {camera_id}",
                'stats': stats,
                'source': 'database'
            })
    except Exception as e:
        print(f"⚠️ Database error: {e}")
    
    return jsonify({
        'camera_id': camera_id,
        'stats': None,
        'source': 'fallback'
    })


@app.route('/api/stats/trend/<period>')
def get_trend_stats_api(period):
    """
    Get trend statistics for charts
    Periods: today, week, month, year, all
    """
    try:
        if not stats_db:
            return jsonify({
                'success': False,
                'error': 'Database not available'
            }), 503
        
        today = datetime.date.today()
        
        if period == 'today':
            # Get hourly breakdown for today
            query = """
                SELECT 
                    HOUR(detected_at) as hour,
                    COUNT(*) as count
                FROM (
                    SELECT detected_at FROM lane_violations WHERE DATE(detected_at) = %s
                    UNION ALL
                    SELECT detected_at FROM helmet_violations WHERE DATE(detected_at) = %s
                    UNION ALL
                    SELECT detected_at FROM red_light_violations WHERE DATE(detected_at) = %s
                ) as all_violations
                GROUP BY HOUR(detected_at)
                ORDER BY hour
            """
            results = stats_db.db.execute_query(query, (today, today, today), fetch=True)
            
            # Fill in missing hours
            hourly_data = [0] * 24
            if results:
                for row in results:
                    hourly_data[row['hour']] = row['count']
            
            # Group by 3-hour intervals
            interval_data = []
            for i in range(0, 24, 3):
                interval_data.append(sum(hourly_data[i:i+3]))
            
            return jsonify({
                'success': True,
                'period': period,
                'labels': ['0h', '3h', '6h', '9h', '12h', '15h', '18h', '21h'],
                'data': interval_data
            })
            
        elif period == 'week':
            # Get last 7 days data
            query = """
                SELECT 
                    DATE(detected_at) as date,
                    COUNT(*) as count
                FROM (
                    SELECT detected_at FROM lane_violations WHERE detected_at >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
                    UNION ALL
                    SELECT detected_at FROM helmet_violations WHERE detected_at >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
                    UNION ALL
                    SELECT detected_at FROM red_light_violations WHERE detected_at >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
                ) as all_violations
                GROUP BY DATE(detected_at)
                ORDER BY date
            """
            results = stats_db.db.execute_query(query, fetch=True)
            
            # Create labels and data
            labels = []
            data = []
            for i in range(6, -1, -1):
                date = today - datetime.timedelta(days=i)
                labels.append(date.strftime('%d/%m'))
                
                # Find data for this date
                count = 0
                if results:
                    for row in results:
                        if row['date'] == date:
                            count = row['count']
                            break
                data.append(count)
            
            return jsonify({
                'success': True,
                'period': period,
                'labels': labels,
                'data': data
            })
            
        elif period == 'month':
            # Get last 30 days, grouped by week
            query = """
                SELECT 
                    WEEK(detected_at, 1) as week,
                    COUNT(*) as count
                FROM (
                    SELECT detected_at FROM lane_violations WHERE detected_at >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
                    UNION ALL
                    SELECT detected_at FROM helmet_violations WHERE detected_at >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
                    UNION ALL
                    SELECT detected_at FROM red_light_violations WHERE detected_at >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
                ) as all_violations
                GROUP BY WEEK(detected_at, 1)
                ORDER BY week
            """
            results = stats_db.db.execute_query(query, fetch=True)
            
            labels = ['Tuần 1', 'Tuần 2', 'Tuần 3', 'Tuần 4', 'Tuần 5']
            data = [row['count'] if results and len(results) > i else 0 for i, row in enumerate(results or [])]
            
            # Ensure we have 5 data points
            while len(data) < 5:
                data.append(0)
            
            return jsonify({
                'success': True,
                'period': period,
                'labels': labels[:len(data)],
                'data': data[:5]
            })
        
        elif period == 'year':
            # Get data for current year, grouped by month
            query = """
                SELECT 
                    MONTH(detected_at) as month,
                    COUNT(*) as count
                FROM (
                    SELECT detected_at FROM lane_violations 
                    WHERE YEAR(detected_at) = YEAR(CURDATE())
                    UNION ALL
                    SELECT detected_at FROM helmet_violations 
                    WHERE YEAR(detected_at) = YEAR(CURDATE())
                    UNION ALL
                    SELECT detected_at FROM red_light_violations 
                    WHERE YEAR(detected_at) = YEAR(CURDATE())
                ) as all_violations
                GROUP BY MONTH(detected_at)
                ORDER BY month
            """
            results = stats_db.db.execute_query(query, fetch=True)
            
            # Create labels and data for 12 months
            labels = ['T1', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'T8', 'T9', 'T10', 'T11', 'T12']
            data = [0] * 12
            if results:
                for row in results:
                    month_idx = row['month'] - 1
                    if 0 <= month_idx < 12:
                        data[month_idx] = row['count']
            
            return jsonify({
                'success': True,
                'period': period,
                'labels': labels,
                'data': data
            })
        
        elif period == 'all':
            # Get data for all time, grouped by year
            query = """
                SELECT 
                    YEAR(detected_at) as year,
                    COUNT(*) as count
                FROM (
                    SELECT detected_at FROM lane_violations
                    UNION ALL
                    SELECT detected_at FROM helmet_violations
                    UNION ALL
                    SELECT detected_at FROM red_light_violations
                ) as all_violations
                GROUP BY YEAR(detected_at)
                ORDER BY year
            """
            results = stats_db.db.execute_query(query, fetch=True)
            
            # Create labels and data
            labels = []
            data = []
            if results:
                for row in results:
                    labels.append(str(row['year']))
                    data.append(row['count'])
            
            # If no data, return empty
            if not labels:
                labels = ['Năm hiện tại']
                data = [0]
            
            return jsonify({
                'success': True,
                'period': period,
                'labels': labels,
                'data': data
            })
        
        else:
            return jsonify({
                'success': False,
                'error': f'Invalid period: {period}'
            }), 400
            
    except Exception as e:
        print(f"❌ Trend stats error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/stats/breakdown')
def get_breakdown_stats_api():
    """
    Get detailed breakdown by camera and violation type
    """
    try:
        if not stats_db:
            return jsonify({
                'success': False,
                'error': 'Database not available'
            }), 503
        
        today = datetime.date.today()
        
        # Get stats for each camera
        cameras_data = []
        
        # Camera 1 - Lane
        lane_query = """
            SELECT 
                COUNT(*) as total_all_time,
                SUM(CASE WHEN DATE(detected_at) = %s THEN 1 ELSE 0 END) as total_today
            FROM lane_violations
            WHERE camera_id = 1
        """
        lane_result = stats_db.db.execute_query(lane_query, (today,), fetch=True)
        
        # Camera 2 - Helmet
        helmet_query = """
            SELECT 
                COUNT(*) as total_all_time,
                SUM(CASE WHEN DATE(detected_at) = %s THEN 1 ELSE 0 END) as total_today
            FROM helmet_violations
            WHERE camera_id = 2
        """
        helmet_result = stats_db.db.execute_query(helmet_query, (today,), fetch=True)
        
        # Camera 3 - Red Light
        redlight_query = """
            SELECT 
                COUNT(*) as total_all_time,
                SUM(CASE WHEN DATE(detected_at) = %s THEN 1 ELSE 0 END) as total_today
            FROM red_light_violations
            WHERE camera_id = 3
        """
        redlight_result = stats_db.db.execute_query(redlight_query, (today,), fetch=True)
        
        lane_data = lane_result[0] if lane_result else {'total_all_time': 0, 'total_today': 0}
        helmet_data = helmet_result[0] if helmet_result else {'total_all_time': 0, 'total_today': 0}
        redlight_data = redlight_result[0] if redlight_result else {'total_all_time': 0, 'total_today': 0}
        
        return jsonify({
            'success': True,
            'cameras': [
                {
                    'camera_id': 1,
                    'camera_name': 'Camera 1 (Làn Đường)',
                    'total_all_time': int(lane_data['total_all_time'] or 0),
                    'total_today': int(lane_data['total_today'] or 0)
                },
                {
                    'camera_id': 2,
                    'camera_name': 'Camera 2 (Mũ Bảo Hiểm)',
                    'total_all_time': int(helmet_data['total_all_time'] or 0),
                    'total_today': int(helmet_data['total_today'] or 0)
                },
                {
                    'camera_id': 3,
                    'camera_name': 'Camera 3 (Đèn Đỏ)',
                    'total_all_time': int(redlight_data['total_all_time'] or 0),
                    'total_today': int(redlight_data['total_today'] or 0)
                }
            ],
            'source': 'database'
        })
        
    except Exception as e:
        print(f"❌ Breakdown stats error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/stats/week-trend')
def get_week_trend_api():
    """
    Get 7-day trend data broken down by violation type
    """
    try:
        if not stats_db:
            return jsonify({
                'success': False,
                'error': 'Database not available'
            }), 503
        
        today = datetime.date.today()
        
        # Prepare date range
        dates = [(today - datetime.timedelta(days=i)).strftime('%Y-%m-%d') for i in range(6, -1, -1)]
        labels = [(today - datetime.timedelta(days=i)).strftime('%d/%m') for i in range(6, -1, -1)]
        
        # Get lane violations
        lane_query = """
            SELECT DATE(detected_at) as date, COUNT(*) as count
            FROM lane_violations
            WHERE detected_at >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
            GROUP BY DATE(detected_at)
        """
        lane_results = stats_db.db.execute_query(lane_query, fetch=True) or []
        lane_dict = {str(row['date']): row['count'] for row in lane_results}
        lane_data = [lane_dict.get(date, 0) for date in dates]
        
        # Get helmet violations
        helmet_query = """
            SELECT DATE(detected_at) as date, COUNT(*) as count
            FROM helmet_violations
            WHERE detected_at >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
            GROUP BY DATE(detected_at)
        """
        helmet_results = stats_db.db.execute_query(helmet_query, fetch=True) or []
        helmet_dict = {str(row['date']): row['count'] for row in helmet_results}
        helmet_data = [helmet_dict.get(date, 0) for date in dates]
        
        # Get red light violations
        redlight_query = """
            SELECT DATE(detected_at) as date, COUNT(*) as count
            FROM red_light_violations
            WHERE detected_at >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
            GROUP BY DATE(detected_at)
        """
        redlight_results = stats_db.db.execute_query(redlight_query, fetch=True) or []
        redlight_dict = {str(row['date']): row['count'] for row in redlight_results}
        redlight_data = [redlight_dict.get(date, 0) for date in dates]
        
        return jsonify({
            'success': True,
            'labels': labels,
            'datasets': {
                'lane': lane_data,
                'helmet': helmet_data,
                'redlight': redlight_data
            },
            'source': 'database'
        })
        
    except Exception as e:
        print(f"❌ Week trend error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/export/csv')
def export_csv_report():
    """
    Export traffic violations report as CSV
    """
    try:
        from io import StringIO
        import csv
        
        # Get overall stats
        if stats_db:
            overall_stats = stats_db.get_overall_stats()
            today = datetime.date.today()
            
            # Create CSV content
            output = StringIO()
            writer = csv.writer(output)
            
            # Header
            writer.writerow(['BÁO CÁO VI PHẠM GIAO THÔNG'])
            writer.writerow([f'Ngày xuất: {datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'])
            writer.writerow([])
            
            # Summary
            writer.writerow(['TỔNG QUAN'])
            writer.writerow(['Loại Vi Phạm', 'Tổng Cộng', 'Hôm Nay'])
            
            total_all = 0
            total_today = 0
            
            for camera in overall_stats:
                camera_name = f"Camera {camera['camera_id']} - {camera['camera_type']}"
                all_time = camera.get('total_violations', 0)
                today_count = camera.get('today_violations', 0)
                
                writer.writerow([camera_name, all_time, today_count])
                total_all += all_time
                total_today += today_count
            
            writer.writerow([])
            writer.writerow(['Tổng Cộng', total_all, total_today])
            
            # Get CSV content
            csv_content = output.getvalue()
            output.close()
            
            # Create response
            response = app.response_class(
                response=csv_content,
                status=200,
                mimetype='text/csv',
                headers={
                    'Content-Disposition': f'attachment; filename=traffic_violations_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
                }
            )
            return response
        else:
            return jsonify({'error': 'Database not available'}), 503
            
    except Exception as e:
        print(f"❌ CSV export error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/excel')
def export_excel_report():
    """
    Export traffic violations report as Excel
    Requires openpyxl: pip install openpyxl
    """
    try:
        from io import BytesIO
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return jsonify({
                'error': 'openpyxl not installed. Run: pip install openpyxl'
            }), 500
        
        if not stats_db:
            return jsonify({'error': 'Database not available'}), 503
        
        # Get data
        overall_stats = stats_db.get_overall_stats()
        today = datetime.date.today()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Báo Cáo Vi Phạm"
        
        # Styles
        header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        title_font = Font(bold=True, size=16)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:D1')
        cell = ws['A1']
        cell.value = 'BÁO CÁO VI PHẠM GIAO THÔNG'
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Date
        ws.merge_cells('A2:D2')
        cell = ws['A2']
        cell.value = f'Ngày xuất: {datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
        cell.alignment = Alignment(horizontal='center')
        
        # Headers
        ws['A4'] = 'Camera'
        ws['B4'] = 'Loại Vi Phạm'
        ws['C4'] = 'Tổng Cộng'
        ws['D4'] = 'Hôm Nay'
        
        for cell in ['A4', 'B4', 'C4', 'D4']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
            ws[cell].border = border
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        row = 5
        total_all = 0
        total_today = 0
        
        for camera in overall_stats:
            ws[f'A{row}'] = f"Camera {camera['camera_id']}"
            ws[f'B{row}'] = camera['camera_type'].title()
            ws[f'C{row}'] = camera.get('total_violations', 0)
            ws[f'D{row}'] = camera.get('today_violations', 0)
            
            total_all += camera.get('total_violations', 0)
            total_today += camera.get('today_violations', 0)
            
            for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
                ws[cell].border = border
                ws[cell].alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
        
        # Total row
        ws[f'A{row}'] = 'TỔNG CỘNG'
        ws[f'B{row}'] = ''
        ws[f'C{row}'] = total_all
        ws[f'D{row}'] = total_today
        
        for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
            ws[cell].font = Font(bold=True)
            ws[cell].border = border
            ws[cell].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = app.response_class(
            response=output.getvalue(),
            status=200,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=traffic_violations_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            }
        )
        return response
        
    except Exception as e:
        print(f"❌ Excel export error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/pdf')
def export_pdf_report():
    """
    Export traffic violations report as PDF
    Requires reportlab: pip install reportlab
    """
    try:
        from io import BytesIO
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            return jsonify({
                'error': 'reportlab not installed. Run: pip install reportlab'
            }), 500
        
        if not stats_db:
            return jsonify({'error': 'Database not available'}), 503
        
        # Get data
        overall_stats = stats_db.get_overall_stats()
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#FF6B00'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Title
        title = Paragraph("BÁO CÁO VI PHẠM GIAO THÔNG", title_style)
        elements.append(title)
        
        # Date
        date_text = f"Ngày xuất: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        date_para = Paragraph(date_text, styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 0.5*inch))
        
        # Table data
        data = [['Camera', 'Loại Vi Phạm', 'Tổng Cộng', 'Hôm Nay']]
        
        total_all = 0
        total_today = 0
        
        for camera in overall_stats:
            data.append([
                f"Camera {camera['camera_id']}",
                camera['camera_type'].title(),
                str(camera.get('total_violations', 0)),
                str(camera.get('today_violations', 0))
            ])
            total_all += camera.get('total_violations', 0)
            total_today += camera.get('today_violations', 0)
        
        data.append(['TỔNG CỘNG', '', str(total_all), str(total_today)])
        
        # Create table
        table = Table(data, colWidths=[2*inch, 2*inch, 1.5*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B00')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E0E0E0')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Create response
        response = app.response_class(
            response=buffer.getvalue(),
            status=200,
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename=traffic_violations_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            }
        )
        return response
        
    except Exception as e:
        print(f"❌ PDF export error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============================================================================
# GEMINI AI CHATBOT ENDPOINTS
# ============================================================================

@app.route('/api/chat', methods=['POST'])
def chat_with_ai():
    """
    Endpoint chính để chat với AI
    
    Request body:
        {
            "message": "Hôm nay có bao nhiêu vi phạm?"
        }
    
    Response:
        {
            "success": true,
            "response": "...",
            "timestamp": "..."
        }
    """
    try:
        if not CHATBOT_AVAILABLE or not chatbot:
            return jsonify({
                'success': False,
                'error': 'Chatbot không khả dụng. Vui lòng kiểm tra cấu hình GEMINI_API_KEY.',
                'response': 'Xin lỗi, chatbot hiện không hoạt động. Vui lòng liên hệ quản trị viên.'
            }), 503
        
        data = request.get_json()
        if not data or 'message' not in data:
            return jsonify({
                'success': False,
                'error': 'Missing message in request body'
            }), 400
        
        user_message = data['message'].strip()
        if not user_message:
            return jsonify({
                'success': False,
                'error': 'Message cannot be empty'
            }), 400
        
        # Gọi chatbot xử lý
        result = chatbot.chat(user_message)
        
        return jsonify({
            'success': result['success'],
            'response': result['response'],
            'error': result.get('error'),
            'timestamp': datetime.datetime.now().isoformat()
        })
        
    except Exception as e:
        print(f"❌ Chat API error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e),
            'response': 'Đã xảy ra lỗi khi xử lý câu hỏi. Vui lòng thử lại.'
        }), 500


@app.route('/api/chat/suggestions', methods=['GET'])
def get_chat_suggestions():
    """
    Lấy danh sách câu hỏi gợi ý
    
    Response:
        {
            "success": true,
            "suggestions": [...]
        }
    """
    try:
        if not CHATBOT_AVAILABLE or not chatbot:
            return jsonify({
                'success': True,
                'suggestions': [
                    "Hôm nay có bao nhiêu vi phạm?",
                    "Thống kê vi phạm vượt đèn đỏ",
                    "Vượt đèn đỏ phạt bao nhiêu?",
                    "Không đội mũ bảo hiểm phạt thế nào?"
                ]
            })
        
        suggestions = chatbot.suggest_questions()
        return jsonify({
            'success': True,
            'suggestions': suggestions
        })
        
    except Exception as e:
        print(f"❌ Suggestions API error: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/chat/quick-stats', methods=['GET'])
def get_quick_stats_for_chat():
    """
    Lấy thống kê nhanh để hiển thị trong chat
    
    Response:
        {
            "success": true,
            "stats": {...}
        }
    """
    try:
        if chatbot:
            stats = chatbot.get_quick_stats()
            return jsonify({
                'success': True,
                'stats': stats
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Chatbot not available'
            }), 503
            
    except Exception as e:
        print(f"❌ Quick stats API error: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route("/red_light_detection_status", methods=['GET'])
def get_red_light_detection_status():
    """
    Lấy thống kê thời gian thực vi phạm đèn đỏ TRỰC TIẾP TỪ DATABASE
    
    Response:
        {
            "success": true,
            "active": false,
            "stats": {
                "total_violations": ...,
                "total_violations_today": ...,
                "unique_vehicles": ...,
                "unique_vehicles_today": ...,
                "runtime_seconds": ...
            },
            "recent_violations": [...]
        }
    """
    try:
        if not stats_db:
            return jsonify({
                'success': False,
                'error': 'Database not available',
                'stats': {
                    'total_violations': 0,
                    'total_violations_today': 0,
                    'unique_vehicles': 0,
                    'unique_vehicles_today': 0,
                    'runtime_seconds': 0
                },
                'recent_violations': []
            })
        
        today = datetime.date.today()
        today_str = today.strftime('%Y-%m-%d')  # ✅ Convert to string format
        camera_id = 3  # Camera 3 = Red Light
        
        # Lấy tổng vi phạm ALL TIME từ database
        redlight_stats_all = stats_db.get_red_light_stats(camera_id)
        total_violations_all = redlight_stats_all['violation_count'] if redlight_stats_all else 0
        unique_vehicles_all = redlight_stats_all['unique_vehicles'] if redlight_stats_all else 0
        
        # Lấy vi phạm HÔM NAY từ database
        redlight_stats_today = stats_db.get_red_light_stats(camera_id, date=today)
        total_violations_today = redlight_stats_today['violation_count'] if redlight_stats_today else 0
        unique_vehicles_today = redlight_stats_today['unique_vehicles'] if redlight_stats_today else 0
        
        # ✅ Lấy danh sách vi phạm gần đây (10 vi phạm mới nhất - TOÀN BỘ)
        recent_violations = []
        try:
            query = """
                SELECT 
                    violation_id,
                    frame_number,
                    time_in_video,
                    license_plate,
                    confidence,
                    detected_at,
                    TIME(detected_at) as time_formatted,
                    DATE(detected_at) as violation_date
                FROM red_light_violations
                WHERE camera_id = %s
                ORDER BY detected_at DESC
                LIMIT 10
            """
            print(f"🔍 Debug: Querying ALL red_light_violations for camera_id={camera_id}")
            results = violation_db.db.execute_query(query, (camera_id,), fetch=True)
            
            if results:
                print(f"✅ Found {len(results)} red light violations from database")
                for row in results:
                    # Format time_formatted properly (TIME() returns timedelta)
                    time_str = ''
                    if row.get('time_formatted'):
                        if isinstance(row['time_formatted'], datetime.timedelta):
                            total_seconds = int(row['time_formatted'].total_seconds())
                            hours = total_seconds // 3600
                            minutes = (total_seconds % 3600) // 60
                            seconds = total_seconds % 60
                            time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                        else:
                            time_str = str(row['time_formatted'])
                    
                    # Format violation date
                    date_str = ''
                    if row.get('violation_date'):
                        if isinstance(row['violation_date'], datetime.date):
                            date_str = row['violation_date'].strftime('%d/%m/%Y')
                        else:
                            date_str = str(row['violation_date'])
                    
                    recent_violations.append({
                        'violation_id': row['violation_id'],
                        'frame_number': row['frame_number'],
                        'time_in_video': float(row['time_in_video']) if row['time_in_video'] else 0,
                        'license_plate': row['license_plate'] or 'Unknown',
                        'confidence': float(row['confidence']) if row['confidence'] else 0,
                        'detected_at': row['detected_at'].isoformat() if row['detected_at'] else '',
                        'time_formatted': time_str,
                        'date_formatted': date_str
                    })
            else:
                print(f"⚠️ No red light violations found in database")
        except Exception as e:
            print(f"⚠️ Error fetching recent violations: {e}")
            import traceback
            traceback.print_exc()
        
        # Tính runtime (giả sử session bắt đầu từ vi phạm đầu tiên trong ngày)
        runtime_seconds = 0
        video_id = session.get('current_video_id_redlight')
        if video_id and video_db:
            try:
                video_info = video_db.get_video_info(video_id)
                if video_info:
                    # Tính thời gian từ khi bắt đầu processing
                    if video_info.get('processing_started_at'):
                        start_time = video_info['processing_started_at']
                        if isinstance(start_time, datetime.datetime):
                            runtime_seconds = (datetime.datetime.now() - start_time).total_seconds()
            except Exception as e:
                print(f"⚠️ Error calculating runtime: {e}")
        
        response_data = {
            'success': True,
            'active': False,  # Red light không có realtime detection như lane/helmet
            'stats': {
                'total_violations': total_violations_all,
                'total_violations_today': total_violations_today,
                'unique_vehicles': unique_vehicles_all,
                'unique_vehicles_today': unique_vehicles_today,
                'runtime_seconds': runtime_seconds
            },
            'recent_violations': recent_violations,
            'source': 'database'
        }
        
        print(f"📊 Sending red light stats: {total_violations_today} violations today, {len(recent_violations)} recent violations")
        return jsonify(response_data)
        
    except Exception as e:
        print(f"❌ Red light detection status error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e),
            'stats': {
                'total_violations': 0,
                'total_violations_today': 0,
                'unique_vehicles': 0,
                'unique_vehicles_today': 0,
                'runtime_seconds': 0
            },
            'recent_violations': []
        }), 500


@app.route('/api/chat/status', methods=['GET'])
def get_chatbot_status():
    """
    Kiểm tra trạng thái chatbot
    
    Response:
        {
            "available": true,
            "ready": true,
            "message": "..."
        }
    """
    return jsonify({
        'available': CHATBOT_AVAILABLE,
        'ready': chatbot is not None,
        'message': 'Chatbot sẵn sàng' if chatbot else 'Chatbot chưa được khởi tạo'
    })


if __name__ == "__main__":
    print("=" * 70)
    print("🚦 TRAFFIC MONITORING SYSTEM")
    print("=" * 70)
    print("\n📹 Camera Configuration:")
    print("   Camera 1: Lane Detection (Phát hiện làn đường)")
    print("   Camera 2: Helmet Detection (Phát hiện mũ bảo hiểm)")
    print("   Camera 3: Red Light Detection (Phát hiện vượt đèn đỏ)")
    print("\n")
    
    # Auto-initialize database with configured password
    print("🔄 Initializing database connection...")
    try:
        # Sử dụng password từ config
        password = '12345678'  # MySQL root password
        if init_database(password):
            print("✅ Database connected! Using MySQL for data storage.")
            print(f"   - Host: localhost")
            print(f"   - User: root")
            print(f"   - Database: traffic_monitoring")
        else:
            print("⚠️ Database connection failed. Running without database.")
            print("   Tip: Kiểm tra MySQL service đang chạy và password đúng")
    except Exception as e:
        print(f"⚠️ Database initialization error: {e}")
        print("⚠️ Running without database support.")
        import traceback
        traceback.print_exc()
    
    print("\n🌐 Starting Flask server...")
    # Chỉ mở browser một lần trong process chính (tránh mở 2 tab khi dùng reloader)
    import os
    if os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        webbrowser.open('http://127.0.0.1:8000/')
    app.run(host="0.0.0.0", port=8000, debug=True, use_reloader=True)

